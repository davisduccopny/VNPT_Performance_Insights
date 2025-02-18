import streamlit as st
import pandas as pd
import base64
from streamlit_option_menu import option_menu
import numpy as np
import io
from openpyxl.styles import Font, Border, Side, PatternFill, Alignment
from openpyxl import Workbook
import openpyxl
import time
import datetime
import PROJECTS.module_view as module_view
import PROJECTS.config as module_config
import streamlit.components.v1 as components
import LDP_MODULE.ldp_view as ldp_view

# PART LOGIN 
if not st.session_state.get("is_logged_in", False):
    with st.spinner("🔐 Đang chuyển hướng đến trang đăng nhập..."):
        time.sleep(2)
    st.session_state.is_logged_in = False
    st.session_state.role_access_admin = False
    st.session_state.line_access = None
    st.switch_page("main.py")
    st.stop() 
# PART SET CONFIG
with open('src/style.css', encoding="utf-8")as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html = True)
with open('src/style_general.css', encoding="utf-8")as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html = True)
with open('src/style_ldp/style_view.css', encoding="utf-8")as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html = True)

class VIEW_LDP_DATA_PROCESS():
    def __init__(self):
        # define line array
        self.array_line_selectbox = line_after_load.drop_duplicates()
        self.arr_line_select = {row["ma_line"]: row["ten_line"] for _, row in self.array_line_selectbox.iterrows()}
        self.selected_line_key =list(self.arr_line_select.keys())
        # end define line array
        
    def ui_info_no_search(self,selected,header_page):
        container_title_table_view = st.container(key="container_title_table_view_ldp")
        ctn_fisrt_main_table_view = st.container(key="ctn_fisrt_main_table_view_ldp")
        
        with container_title_table_view:
            col_title_dmanage_expand = st.columns([2,2.2])
            col_title_dmanage_expand[0].markdown(f"""<h3 style='text-align: left;'>{header_page}</h3>""", unsafe_allow_html=True)
            with col_title_dmanage_expand[1]:
                cols_button_view_table = st.columns([1,1,1])
                with cols_button_view_table[0]:
                   with st.popover("Chế độ",use_container_width=True, icon=":material/tune:"):
                        radio_type_view_table = st.radio("Chọn chế độ",["Đơn","Kết hợp"], key="radio_type_view_table_ldp",horizontal=True)
                        form_view_service = st.selectbox("Chọn form dịch vụ phù hợp:",["Thu gọn","Đầy đủ"], key="form_view_service_ldp")
                with cols_button_view_table[1]:
                    export_excel_table_file = st.button("Xuất Excel", key="button_export_excel_ldp", icon=":material/cloud_download:", use_container_width=True)
        
        
        with ctn_fisrt_main_table_view:
            if radio_type_view_table == "Đơn":
                cols_view_body = ctn_fisrt_main_table_view.columns(4)
            else:
                cols_view_body = ctn_fisrt_main_table_view.columns(2)
            with cols_view_body[0]:
                year_et = st.empty()
                selected_year = year_et.selectbox("Chọn năm",list(range(2021,2030)),index=list(range(2021,2030)).index(datetime.datetime.now().year), key="year_viewldp_select") 
                selected_year = int(selected_year)
            with cols_view_body[2] if radio_type_view_table == "Đơn" else cols_view_body[1]:
                line_access_et = st.empty()
                selected_line_access = line_access_et.selectbox("Chọn line",options=self.selected_line_key, 
                                                format_func= lambda x: self.arr_line_select[x] if x else ""
                                                , key="line_viewldp_select")
            with cols_view_body[3] if radio_type_view_table == "Đơn" else cols_view_body[1]:
                month_et = st.empty()
                selected_month = month_et.selectbox("Chọn tháng",list(range(1,13)), key="month_viewldp_select")
                selected_month = int(selected_month)
                    

            if selected == "Dạng bảng tổng hợp":
                loaidoanhthu_et = cols_view_body[1].empty()
                selected_loaidoanhthu = loaidoanhthu_et.selectbox("Chọn loại doanh thu",["hiện hữu","mới trong tháng","mới trong năm","tổng doanh thu"], key="loaidoanhthu_viewldp_select")
                
                if radio_type_view_table == "Kết hợp":
                    loaidoanhthu_et.empty()
            else:
                if radio_type_view_table == "Kết hợp":
                    selected_loaidoanhthu = st.selectbox("Chọn loại doanh thu",["Hiện hữu","Phát triển mới"], key="loaidoanhthu_viewldp_select")
                else:
                    selected_loaidoanhthu = cols_view_body[1].selectbox("Chọn loại doanh thu",["Hiện hữu","Phát triển mới"], key="loaidoanhthu_viewldp_select") 
                   
            if radio_type_view_table == "Kết hợp":
                line_access_et.empty()
        
        with container_title_table_view:
                with col_title_dmanage_expand[1]:
                    with cols_button_view_table[2]:
                        if radio_type_view_table == "Đơn":
                            submit_view_option = st.button("Xem", key="submit_view_option_ldp", type="primary" ,icon=":material/visibility:", use_container_width=True)
                        else:
                            submit_view_option = st.button("Tạo file", key="submit_view_option_ldp", type="primary" ,icon=":material/send_money:", use_container_width=True)
       
        
        return export_excel_table_file,submit_view_option,radio_type_view_table,form_view_service,selected_year,selected_month,selected_line_access,selected_loaidoanhthu
    
    def filter_thuchien_for_month_view(self,thuchien_after_load,selected_month, selected_year,line,selected_loaidoanhthu,radio_data_selected_kind):
        if selected_month is not None:
            thuchien_after_filter = thuchien_after_load[(thuchien_after_load["type_process"] == radio_data_selected_kind) &
                                                        (thuchien_after_load["line"] == line) &
                                                        (thuchien_after_load["thang"] == selected_month) &
                                                        (thuchien_after_load["year_insert"] == selected_year) &
                                                        (thuchien_after_load["loaidoanhthu"] == selected_loaidoanhthu)]
            thuchien_after_filter = thuchien_after_filter[["nhom_dv", "doanhthu"]]
            thuchien_after_filter["nhom_dv"] = thuchien_after_filter["nhom_dv"].str.strip()
        else:
            thuchien_after_filter = thuchien_after_load[(thuchien_after_load["type_process"] == radio_data_selected_kind) &
                                                        (thuchien_after_load["line"] == line) &
                                                        (thuchien_after_load["year_insert"] == selected_year) &
                                                        (thuchien_after_load["loaidoanhthu"] == selected_loaidoanhthu)]
            thuchien_after_filter = thuchien_after_filter[["nhom_dv", "doanhthu", "thang","year_insert"]]
            thuchien_after_filter["nhom_dv"] = thuchien_after_filter["nhom_dv"].str.strip()
        return thuchien_after_filter

    def filter_kehoach_for_month_view(self,kehoach_after_load, dichvu_after_load,line, selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind):
        if selected_month is not None:
            filtered_kehoach = kehoach_after_load[
                (kehoach_after_load["type_process"] == radio_data_selected_kind) &
                (kehoach_after_load["year_insert"] == int(selected_year)) &
                (kehoach_after_load["line"] == line) &
                (kehoach_after_load["loaidoanhthu"] == selected_loaidoanhthu)
            ].copy()
            filtered_kehoach.loc[:, "ten_dv"] = filtered_kehoach["id_dv_606"].map(dict(zip(dichvu_after_load["ma_dv_id66"], dichvu_after_load["ten_dv"])))
            filtered_kehoach = filtered_kehoach[["ten_dv",f"t{selected_month}"]]
        else:
            filtered_kehoach = kehoach_after_load[
                (kehoach_after_load["type_process"] == radio_data_selected_kind) &
                (kehoach_after_load["year_insert"] == int(selected_year)) &
                (kehoach_after_load["line"] == line) &
                (kehoach_after_load["loaidoanhthu"] == selected_loaidoanhthu)
            ].copy()
            filtered_kehoach.loc[:, "ten_dv"] = filtered_kehoach["id_dv_606"].map(dict(zip(dichvu_after_load["ma_dv_id66"], dichvu_after_load["ten_dv"])))
            filtered_kehoach = filtered_kehoach[["ten_dv", "t1", "t2", "t3", "t4", "t5", "t6", "t7", "t8", "t9", "t10", "t11", "t12"]]
            filtered_kehoach.rename(columns={"year_insert": "Năm"}, inplace=True)
            
        return filtered_kehoach
    def filter_thuchien_for_year_view(self,thuchien_after_load,selected_month, selected_year,line,selected_loaidoanhthu,radio_data_selected_kind):
        general_filter_for_month = thuchien_after_load[(thuchien_after_load["line"] == line) &
                                                        (thuchien_after_load["type_process"] == radio_data_selected_kind) &
                                                        (thuchien_after_load["year_insert"] == selected_year) &
                                                        (thuchien_after_load["loaidoanhthu"] == selected_loaidoanhthu)]
        if selected_month is not None:
            
            # Không lũy kế
            thuchien_after_filter = general_filter_for_month[general_filter_for_month["thang"] == selected_month]
            thuchien_after_filter = thuchien_after_filter[["nhom_dv", "doanhthu"]]
            thuchien_after_filter["nhom_dv"] = thuchien_after_filter["nhom_dv"].str.strip()
            # Lũy kế
            thuchien_after_filter_luyke = general_filter_for_month[general_filter_for_month["thang"] <= selected_month]
            thuchien_after_filter_luyke = thuchien_after_filter_luyke[["nhom_dv", "doanhthu"]]
            thuchien_after_filter_luyke["nhom_dv"] = thuchien_after_filter_luyke["nhom_dv"].str.strip()
            if not thuchien_after_filter_luyke.empty:
                thuchien_after_filter_luyke['doanhthu'] = thuchien_after_filter_luyke.groupby('nhom_dv')['doanhthu'].cumsum()
            thuchien_after_filter_luyke = thuchien_after_filter_luyke.drop_duplicates(subset='nhom_dv', keep='last')
            
        else:
            # Không lũy kế
            thuchien_after_filter = general_filter_for_month
            thuchien_after_filter = thuchien_after_filter[["nhom_dv", "doanhthu", "thang","year_insert"]]
            thuchien_after_filter["nhom_dv"] = thuchien_after_filter["nhom_dv"].str.strip()
            # Lũy kế
            thuchien_after_filter_luyke = general_filter_for_month
            thuchien_after_filter_luyke = thuchien_after_filter_luyke[["nhom_dv", "doanhthu", "thang","year_insert"]]
            thuchien_after_filter_luyke["nhom_dv"] = thuchien_after_filter_luyke["nhom_dv"].str.strip()

            
            cumulative_sum_per_service = {nhom_dv: [] for nhom_dv in thuchien_after_filter_luyke['nhom_dv'].unique()}
            for index, row in thuchien_after_filter_luyke.iterrows():
                nhom_dv = row['nhom_dv']
                if len(cumulative_sum_per_service[nhom_dv]) == 0:
                    cumulative_sum_per_service[nhom_dv].append(row['doanhthu'])
                else:
                    cumulative_sum_per_service[nhom_dv].append(cumulative_sum_per_service[nhom_dv][-1] + row['doanhthu'])
            thuchien_after_filter_luyke['doanhthu'] = [cumulative_sum_per_service[row['nhom_dv']].pop(0) for index, row in thuchien_after_filter_luyke.iterrows()]

        return thuchien_after_filter,thuchien_after_filter_luyke

    def filter_kehoach_for_year_view(self,kehoach_after_load, dichvu_after_load,line, selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind):
        general_filter_for_month =  kehoach_after_load[
                (kehoach_after_load["type_process"] == radio_data_selected_kind) &
                (kehoach_after_load["year_insert"] == int(selected_year)) &
                (kehoach_after_load["line"] == line) &
                (kehoach_after_load["loaidoanhthu"] == selected_loaidoanhthu)
            ]
        if selected_month is not None:
            # Không lũy kế
            filtered_kehoach =general_filter_for_month.copy()
            filtered_kehoach.loc[:, "ten_dv"] = filtered_kehoach["id_dv_606"].map(dict(zip(dichvu_after_load["ma_dv_id66"], dichvu_after_load["ten_dv"])))
            filtered_kehoach = filtered_kehoach[["ten_dv",f"t{selected_month}"]]
            # Lũy kế
            previous_month_cols = [f't{i}' for i in range(1, selected_month)]
            filtered_kehoach_luyke =general_filter_for_month.copy()
            filtered_kehoach_luyke.loc[:, "ten_dv"] = filtered_kehoach_luyke["id_dv_606"].map(dict(zip(dichvu_after_load["ma_dv_id66"], dichvu_after_load["ten_dv"])))
            filtered_kehoach_luyke = filtered_kehoach_luyke[["ten_dv",f"t{selected_month}", "year_insert"] + previous_month_cols]
            
            for col in previous_month_cols:
                filtered_kehoach_luyke[f't{selected_month}'] += filtered_kehoach_luyke[col]
            filtered_kehoach_luyke = filtered_kehoach_luyke[["ten_dv", f"t{selected_month}"]]
        else:
            # Không lũy kế
            filtered_kehoach =general_filter_for_month.copy()
            filtered_kehoach.loc[:, "ten_dv"] = filtered_kehoach["id_dv_606"].map(dict(zip(dichvu_after_load["ma_dv_id66"], dichvu_after_load["ten_dv"])))
            filtered_kehoach = filtered_kehoach[["ten_dv", "t1", "t2", "t3", "t4", "t5", "t6", "t7", "t8", "t9", "t10", "t11", "t12"]]
            filtered_kehoach.rename(columns={"year_insert": "Năm"}, inplace=True)
            # Lũy kế
            filtered_kehoach_luyke =general_filter_for_month.copy()
            filtered_kehoach_luyke.loc[:, "ten_dv"] = filtered_kehoach_luyke["id_dv_606"].map(dict(zip(dichvu_after_load["ma_dv_id66"], dichvu_after_load["ten_dv"])))
            filtered_kehoach_luyke = filtered_kehoach_luyke[["ten_dv", "t1", "t2", "t3", "t4", "t5", "t6", "t7", "t8", "t9", "t10", "t11", "t12"]]
            filtered_kehoach_luyke.rename(columns={"year_insert": "Năm"}, inplace=True)
            month_columns = [col for col in filtered_kehoach_luyke.columns if (col.startswith("t") and col != "ten_dv")]
            filtered_kehoach_luyke[month_columns] =filtered_kehoach_luyke[month_columns].cumsum(axis=1)
        return filtered_kehoach,filtered_kehoach_luyke
    def filter_kehoach_year_sum(self,kehoach_after_load, dichvu_after_load,line, selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind):
        
        previous_month_cols = [f't{i}' for i in range(1, 12) if i != selected_month]
        filtered_kehoach_luyke = kehoach_after_load[
            (kehoach_after_load["type_process"] == radio_data_selected_kind) &
            (kehoach_after_load["year_insert"] == int(selected_year)) &
            (kehoach_after_load["line"] == line) &
            (kehoach_after_load["loaidoanhthu"] == selected_loaidoanhthu)
        ].copy()
        filtered_kehoach_luyke.loc[:, "ten_dv"] = filtered_kehoach_luyke["id_dv_606"].map(dict(zip(dichvu_after_load["ma_dv_id66"], dichvu_after_load["ten_dv"])))
        filtered_kehoach_luyke = filtered_kehoach_luyke[["ten_dv",f"t{selected_month}", "year_insert"] + previous_month_cols]
        for col in previous_month_cols:
            filtered_kehoach_luyke[f't{selected_month}'] += filtered_kehoach_luyke[col]
        filtered_kehoach_luyke = filtered_kehoach_luyke[["ten_dv", f"t{selected_month}"]]
        
        return filtered_kehoach_luyke
    
class GENERATE_VIEW():
    def __init__(self):
        self.class_data_option_process = VIEW_LDP_DATA_PROCESS()
    def highlight_numbers(self,val):
        """Định dạng màu sắc cho số âm/dương"""
        try:
            val = float(val)
        except ValueError:
            return "color: black;"
        if val > 0:
            return "color: green;" 
        elif val < 0:
            return "color: red;"    
        else:
            return "color: black;"
    def highlight_numbers_percentage(self,val):
        """Định dạng màu sắc cho số âm/dương"""
        try:
            val = float(val.strip('%'))
        except ValueError:
            return "color: black;"
        if (val - 100) > 0 and val != 100:
            return "color: green;" 
        elif (val - 100) < 0 and val != 100:
            return "color: red;"    
        else:
            return "color: black;"  
    def highlight_nonzero(self,val):
            if val == 0:
                 color = '#A6988D' 
            elif val == "0%":
                color = '#A6988D' 
            else:
                color = '#0D0D0D'  
            return f'color: {color}'
    def add_arrow(self,val):
        """Thêm mũi tên cho số âm/dương"""
        try:
            val = float(val)
        except ValueError:
            return "color: black;"
        if val > 0:
            return f"↑ {val}"
        elif val < 0:
            return f"↓ {val}" 
        else:
            return val
    def add_arrow_percentage(self,val):
        """Thêm mũi tên vào giá trị phần trăm"""
        try:
            val = float(val.strip('%'))
        except ValueError:
            return f"{val}%"  # Return as is if conversion fails
        if (val - 100) > 0 and val != 100:
            return f"↑ {val:.1f}%"  # Mũi tên lên và định dạng %.1f
        elif (val - 100) < 0 and val != 100:
            return f"↓ {val:.1f}%"  # Mũi tên xuống
        else:
            return f"{val:.1f}%"
    def download_excel_single(self,file_name,selected_year,selected_month,selected_loaidoanhthu,line,radio_type_view_table,form_view_service,radio_data_kind):
        with st.spinner("Đang tạo file..."):    
            if radio_type_view_table == "Đơn":
                ma_line = [line,]
                ten_line = [line_after_load[line_after_load["ma_line"] == line]["ten_line"].values[0],]
            else:
                ma_line = line_after_load["ma_line"].unique()
                ten_line = line_after_load["ten_line"].unique()
            arr_hierachy_1, arr_hierachy_2 = ldp_view.build_hierarchy(dichvu_after_load)
            if form_view_service == "Thu gọn":
                count_row_form_3 = len(arr_hierachy_2) + 2
                array_bold_raw = ldp_view.flag_row_form_bold(arr_hierachy_2)
                array_bold_raw_form_1 = [x + 4 for x in array_bold_raw]
                array_bold_raw_form_2 = [x + 4 + count_row_form_3 for x in array_bold_raw]
                array_bold = [array_bold_raw_form_1[0] - 1] + [array_bold_raw_form_2[0] - 1] + array_bold_raw_form_1 + array_bold_raw_form_2
            else:
                count_row_form_3 = len(arr_hierachy_1) + 2
                array_bold_raw = ldp_view.flag_row_form_bold(arr_hierachy_1)
                array_bold_raw_form_1 = [x + 4 for x in array_bold_raw]
                array_bold_raw_form_2 = [x + 4 + count_row_form_3 for x in array_bold_raw]
                
                array_bold = [array_bold_raw_form_1[0] - 1] + [array_bold_raw_form_2[0] - 1] + array_bold_raw_form_1 + array_bold_raw_form_2
            output_all_tt = io.BytesIO()
            wb = Workbook()
            Count_process_status_single = 0
            progress_bar_single = st.empty()
            progress_bar_single.progress(Count_process_status_single)
            for name, sheetname in zip(ma_line, ten_line):
                Count_process_status_single += 0.5
                ws = wb.create_sheet(title=name)
                ws['A1'] = "Phòng"
                ws['B1'] = "DN3"
                ws['A2'] = "Line"
                ws['B2'] = sheetname
                df_form1,df_form2,df_form3 = self.map_data_for_month_view(name,selected_month,selected_year,selected_loaidoanhthu,form_view_service,radio_data_kind)
                percentage_success_status = int((Count_process_status_single/len(ma_line))*100)
                progress_bar_single.progress(percentage_success_status,f"Hoàn thành {percentage_success_status}%")
                try:
                    if df_form1 is None or df_form3 is None:
                        st.warning(f"Không có dữ liệu để xuất file - {sheetname}")
                        continue
                    data = [df_form3.columns.tolist()] + df_form3.values.tolist()
                    for r_idx, row in enumerate(data, 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx + 3, column=c_idx, value=value)
                            cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

                    data = [df_form2.columns.tolist()] + df_form2.values.tolist()
                    for r_idx, row in enumerate(data, 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx + 3, column=c_idx + 15, value=value)
                            cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

                    data = [df_form1.columns.tolist()] + df_form1.values.tolist()
                    thin_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                    for r_idx, row in enumerate(data, start=1):
                        for c_idx, value in enumerate(row, start=1):
                            cell = ws.cell(row=r_idx + 3 + count_row_form_3, column=c_idx, value=value)
                            cell.font = Font(name='Times New Roman', size=12)
                            cell.border = thin_border
                            try:
                                if c_idx in [8,11]: 
                                    val = float(value)
                                    if val != 0:
                                        cell.font = Font(color="00FF00" if val > 0 else "FF0000")
                                        cell.value = f"{'↑' if val > 0 else '↓'} {abs(val)}"
                                    else:
                                        cell.value = val
                                    cell.alignment = openpyxl.styles.Alignment(horizontal='right')

                                elif c_idx in [5,7,10]:  
                                    val = float(value.strip('%'))
                                    if val != 0 and val != 100:
                                        cell.font = Font(color="00FF00" if (val - 100) > 0 else "FF0000")
                                        cell.value = f"{'↑' if (val - 100) > 0 else '↓'} {abs(val):.1f}%"
                                    else:
                                        cell.value = value
                                    cell.alignment = openpyxl.styles.Alignment(horizontal='right')
                            except (ValueError, AttributeError):
                                pass
                    for col_range in [(3, 15), (16, 30)]:
                        start_col, end_col = col_range
                        start_cell = ws.cell(row=3, column=start_col)
                        end_cell = ws.cell(row=3, column=end_col)
                        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
                        
                        # Thiết lập lại border cho các ô trong vùng merge
                        for row in ws.iter_rows(min_row=3, max_row=3, min_col=start_col, max_col=end_col):
                            for cell in row:
                                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))

                    # Gán giá trị cho ô merged và thiết lập font, đường viền
                    for col_idx, text in [(3, 'KẾ HOẠCH'), (16, 'THỰC HIỆN')]:
                        cell = ws.cell(row=3, column=col_idx, value=text)
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell.font = Font(name='Times New Roman', size=12, bold=True)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                        # In đậm các dòng cần thiết
                    bold_rows_need_thuchien = array_bold
                    for row_idx in bold_rows_need_thuchien:
                        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_idx, max_row=row_idx):
                            for cell in col:
                                cell.font = Font(bold=True, color=cell.font.color)
                    # Tự động điều chỉnh độ rộng của các cột dựa trên nội dung của ô
                    for column_cells in ws.columns:
                        length = max(len(str(cell.value)) for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
                except (ValueError, AttributeError):
                    continue
                Count_process_status_single += 0.5
                percentage_success_status = int((Count_process_status_single/len(ma_line))*100)
                progress_bar_single.progress(percentage_success_status,f"Hoàn thành {percentage_success_status}%")
            wb.save(output_all_tt)
            output_all_tt.seek(0)
            excel_base64 = base64.b64encode(output_all_tt.read()).decode()

            # Create a download link using st.download_button
            download_link_all_tt = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{file_name}.xlsx">Tải xuống kết quả</a>'
            st.success("Tạo file thành công.")
            progress_bar_single.empty()
            st.markdown(download_link_all_tt, unsafe_allow_html=True)
    
    def download_excel_year_view(self,file_name,selected_year,selected_month,line,radio_type_view_table,form_view_service,radio_data_selected_kind):
        with st.spinner("Đang tạo file..."):
            start_time_performance = time.time()
            if radio_type_view_table == "Đơn":
                ma_line = [line,]
                ten_line = [line_after_load[line_after_load["ma_line"] == line]["ten_line"].values[0],]
            else:
                ma_line = line_after_load["ma_line"].unique() 
                ten_line = line_after_load["ten_line"].unique()
            arr_hierachy_1, arr_hierachy_2 = ldp_view.build_hierarchy(dichvu_after_load)
            if form_view_service == "Thu gọn":
                count_row_form_3 = len(arr_hierachy_2) + 6
                array_bold_raw = ldp_view.flag_row_form_bold(arr_hierachy_2)
                
                array_bold_raw_form_1 = [x + 5 for x in array_bold_raw]
                array_bold_raw_form_2 = [x + 4 + count_row_form_3 for x in array_bold_raw]
                array_bold_raw_form_3 = [x + 4 + count_row_form_3 * 2 for x in array_bold_raw]
                array_bold_raw_form_4 = [x + 4 + count_row_form_3 * 3 for x in array_bold_raw]
                array_bold_raw_form_5 = [x + 4 + count_row_form_3 * 4 for x in array_bold_raw]
                array_bold_raw_form_6 = [x + 4 + count_row_form_3 * 5 for x in array_bold_raw]
                array_bold_raw_form_7 = [x + 4 + count_row_form_3 * 6 for x in array_bold_raw]
                array_bold_raw_form_8 = [x + 4 + count_row_form_3 * 7 for x in array_bold_raw]
                array_bold = [array_bold_raw_form_1[0] - 1] + [array_bold_raw_form_2[0] - 1] + array_bold_raw_form_1 + array_bold_raw_form_2 + [array_bold_raw_form_3[0] - 1] + [array_bold_raw_form_4[0] - 1] + [array_bold_raw_form_5[0] - 1] + [array_bold_raw_form_6[0] - 1] + [array_bold_raw_form_7[0] - 1] + [array_bold_raw_form_8[0] - 1] + array_bold_raw_form_3 + array_bold_raw_form_4 + array_bold_raw_form_5 + array_bold_raw_form_6 + array_bold_raw_form_7 + array_bold_raw_form_8
            else:
                count_row_form_3 = len(arr_hierachy_1) + 6
                array_bold_raw = ldp_view.flag_row_form_bold(arr_hierachy_1)
                array_bold_raw_form_1 = [x + 5 for x in array_bold_raw]
                array_bold_raw_form_2 = [x + 4 + count_row_form_3 for x in array_bold_raw]
                array_bold_raw_form_3 = [x + 4 + count_row_form_3 * 2 for x in array_bold_raw]
                array_bold_raw_form_4 = [x + 4 + count_row_form_3 * 3 for x in array_bold_raw]
                array_bold_raw_form_5 = [x + 4 + count_row_form_3 * 4 for x in array_bold_raw]
                array_bold_raw_form_6 = [x + 4 + count_row_form_3 * 5 for x in array_bold_raw]
                array_bold_raw_form_7 = [x + 4 + count_row_form_3 * 6 for x in array_bold_raw]
                array_bold_raw_form_8 = [x + 4 + count_row_form_3 * 7 for x in array_bold_raw]
                
                array_bold = [array_bold_raw_form_1[0] - 1] + [array_bold_raw_form_2[0] - 1] + array_bold_raw_form_1 + array_bold_raw_form_2 + [array_bold_raw_form_3[0] - 1] + [array_bold_raw_form_4[0] - 1] + [array_bold_raw_form_5[0] - 1] + [array_bold_raw_form_6[0] - 1] + [array_bold_raw_form_7[0] - 1] + [array_bold_raw_form_8[0] - 1] + array_bold_raw_form_3 + array_bold_raw_form_4 + array_bold_raw_form_5 + array_bold_raw_form_6 + array_bold_raw_form_7 + array_bold_raw_form_8
                
            output_all_tt = io.BytesIO()
            wb = Workbook()
            Count_process_status = 0
            progress_bar = st.empty()
            progress_bar.progress(0)

            for name, sheetname in zip(ma_line, ten_line):
                Count_process_status += 0.5
                ws = wb.create_sheet(title=name)
                ws['A1'] = "Phòng"
                ws['B1'] = "DN3"
                ws['A2'] = "Line"
                ws['B2'] = sheetname
                try: 
                    dataframes = []
                    for ldt in ["tổng doanh thu","hiện hữu","mới trong tháng","mới trong năm"]:
                        try:
                            df_1,df_2,df_3,df_1_luyke,df_2_luyke,df_3_luyke,df_4,df_4_luyke = self.map_data_for_year_view(name,selected_month,selected_year,ldt,form_view_service,radio_data_selected_kind)
                            if ldt == "tổng doanh thu" and (df_1 is not None):
                                dataframes += [df_1,df_1_luyke,df_2,df_2_luyke,df_3,df_3_luyke,df_4,df_4_luyke]
                            elif ldt != "tổng doanh thu" and (df_1 is not None):
                                dataframes += [df_1.iloc[:,2:],df_1_luyke.iloc[:,2:],df_2.iloc[:,2:],df_2_luyke.iloc[:,2:],df_3.iloc[:,2:],df_3_luyke.iloc[:,2:],df_4.iloc[:,2:],df_4_luyke.iloc[:,2:]]
                        except:
                            continue
                    percentage_status_updated = int((Count_process_status / len(ma_line)) * 100) if Count_process_status > 0 else 0
                    progress_bar.progress(value=percentage_status_updated,text=f"Hoàn thành {percentage_status_updated}%")
                    
                    def add_data_to_sheet(ws, dataframe, start_row, start_col, offset=0):
                        data = [dataframe.columns.tolist()] + dataframe.values.tolist()
                        for r_idx, row in enumerate(data, 0): 
                            for c_idx, value in enumerate(row, 0): 
                                cell = ws.cell(row=r_idx + start_row, column=c_idx + start_col, value=value)
                                cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                    def merge_columns(ws, col_ranges, row):
                        for col_range in col_ranges:
                            start_col, end_col = col_range
                            ws.merge_cells(start_row=row, start_column=start_col, end_row=row, end_column=end_col)
                            for cell in ws.iter_rows(min_row=row, max_row=row, min_col=start_col, max_col=end_col):
                                for c in cell:
                                    c.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                    def set_merged_cell(ws, col_info, row):
                        for col_idx, text, color in col_info:
                            cell = ws.cell(row=row, column=col_idx, value=text)
                            cell.alignment = Alignment(horizontal='center')
                            cell.font = Font(name='Times New Roman', size=13, bold=True)  # Thiết lập font chữ và cỡ chữ
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                            cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
                    
                    start_row = [
                        5,
                        count_row_form_3 + 4,
                        count_row_form_3 * 2 + 4,
                        count_row_form_3 * 3 + 4,
                        count_row_form_3 * 4 + 4,
                        count_row_form_3 * 5 + 4,
                        count_row_form_3 * 6 + 4,
                        count_row_form_3 * 7 + 4
                    ]
                    column_offsets_first_row = [1, 14, 25, 36]
                    column_offsets_first_row_luyke = [1, 11, 19, 27]
                    column_offsets_other_rows = [1, 15, 27, 39]
                    array_first_form = [0, 8, 16, 24]
                    array_first_form_luyke = [ 1,9,17,25]

                    for idx, df in enumerate(dataframes):
                        row_idx = idx % 8 
                        col_idx = idx // 8  

                        start_cell_row = start_row[row_idx] 
                        
                        if idx in array_first_form:
                            start_cell_col = column_offsets_first_row[col_idx]
                        elif idx in array_first_form_luyke:
                            start_cell_col = column_offsets_first_row_luyke[col_idx]
                        else:
                            start_cell_col = column_offsets_other_rows[col_idx]
                        
                        add_data_to_sheet(ws, df, start_cell_row, start_cell_col)
                        
                    Stringdde = ["I. PHÂN TÍCH","II. PHÂN TÍCH LŨY KẾ","III. THỰC HIỆN"
                                 ,"IV. THỰC HIỆN LŨY KẾ","V. KẾ HOẠCH","VI. KẾ HOẠCH LŨY KẾ","VII. % THỰC HIỆN","VIII. % THỰC HIỆN LŨY KẾ"]
                    for idx, row in enumerate(start_row):
                        ws[f'A{row - 2}'] = Stringdde[idx]
                        ws[f'A{row - 2}'].font = Font(name='Times New Roman', size=14, bold=True,color="FC0F03")
                        if idx == 0:
                            merge_columns(ws, [(3, column_offsets_first_row[1]-1), (column_offsets_first_row[1], column_offsets_first_row[2]-1), 
                                            (column_offsets_first_row[2], column_offsets_first_row[3]-1), 
                                            (column_offsets_first_row[3], column_offsets_first_row[3] -1 + 11)], row-1)
                            set_merged_cell(ws, [(3, 'TỔNG DOANH THU', '2F75B5'),
                                                (column_offsets_first_row[1], 'HIỆN HỮU', '8932E5'),
                                                (column_offsets_first_row[2], 'MỚI TRONG NĂM', '548235'),
                                                (column_offsets_first_row[3], 'MỚI TRONG THÁNG', 'F4B084')], row-1)
                        elif idx == 1:
                            merge_columns(ws, [(3, column_offsets_first_row_luyke[1]-1), (column_offsets_first_row_luyke[1], column_offsets_first_row_luyke[2]-1),
                                                (column_offsets_first_row_luyke[2], column_offsets_first_row_luyke[3]-1), 
                                                (column_offsets_first_row_luyke[3], column_offsets_first_row_luyke[3] -1 + 8)], row-1)
                            set_merged_cell(ws, [(3, 'TỔNG DOANH THU', '2F75B5'),
                                                (column_offsets_first_row_luyke[1], 'HIỆN HỮU', '8932E5'),
                                                (column_offsets_first_row_luyke[2], 'MỚI TRONG NĂM', '548235'),
                                                (column_offsets_first_row_luyke[3], 'MỚI TRONG THÁNG', 'F4B084')], row-1)
                            
                        else:
                            merge_columns(ws, [(3, column_offsets_other_rows[1]-1), (column_offsets_other_rows[1], column_offsets_other_rows[2]-1),
                                                    (column_offsets_other_rows[2], column_offsets_other_rows[3]-1), 
                                                    (column_offsets_other_rows[3], column_offsets_other_rows[3] -1 + 12)], row-1)
                            set_merged_cell(ws, [(3, 'TỔNG DOANH THU', '2F75B5'),
                                                (column_offsets_other_rows[1], 'HIỆN HỮU', '8932E5'),
                                                (column_offsets_other_rows[2], 'MỚI TRONG NĂM', '548235'),
                                                (column_offsets_other_rows[3], 'MỚI TRONG THÁNG', 'F4B084')], row-1)
                    
                    
                    bold_rows_need_thuchien = array_bold
                    for row_idx in bold_rows_need_thuchien:
                        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_idx, max_row=row_idx):
                            for cell in col:
                                cell.font = Font(name='Times New Roman',bold=True,size=12)
                    for column_cells in ws.columns:
                        length = max(len(str(cell.value)) for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length + 5
                    
                except (ValueError, AttributeError) as e:
                    st.error(f"Error: {e}")
                    continue
                
                Count_process_status += 0.5
                percentage_status_updated = int((Count_process_status / len(ma_line)) * 100) if Count_process_status > 0 else 0
                progress_bar.progress(value=percentage_status_updated,text=f"Hoàn thành {percentage_status_updated}%")
                
            wb.save(output_all_tt)
            output_all_tt.seek(0)
            excel_base64 = base64.b64encode(output_all_tt.read()).decode()
            download_link_all_tt = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{file_name}.xlsx">Tải xuống kết quả</a>'
            st.success("Tạo file thành công.")
            st.markdown(download_link_all_tt, unsafe_allow_html=True)
            progress_bar.empty()
            end_time_performance = time.time()
            elapsed_time = end_time_performance - start_time_performance
            st.write(f"Thời gian thực hiện: {elapsed_time:.4f} giây")
    
    def map_data_for_month_view(self,line,selected_month,selected_year,selected_loaidoanhthu,form_view_service,radio_data_selected_kind):
        filtered_kehoach = self.class_data_option_process.filter_kehoach_for_month_view(kehoach_after_load,dichvu_after_load,line,selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind)
        filtered_thuchien_ht = self.class_data_option_process.filter_thuchien_for_month_view(thuchien_after_load,selected_month,selected_year,line,selected_loaidoanhthu,radio_data_selected_kind)
        filtered_thuchien_t1 = self.class_data_option_process.filter_thuchien_for_month_view(thuchien_after_load,1,selected_year,line,selected_loaidoanhthu,radio_data_selected_kind)
        filtered_thuchien_kytruoc = self.class_data_option_process.filter_thuchien_for_month_view(thuchien_after_load,selected_month -1,selected_year,line,selected_loaidoanhthu,radio_data_selected_kind)
        form_2_filter = self.class_data_option_process.filter_thuchien_for_month_view(thuchien_after_load,None,selected_year,line,selected_loaidoanhthu,radio_data_selected_kind)
        form3_kehoach_after_filter_no_month = self.class_data_option_process.filter_kehoach_for_month_view(kehoach_after_load,dichvu_after_load,line,None,selected_year,selected_loaidoanhthu,radio_data_selected_kind)
        df_hierachy_1, df_hierachy_2 = ldp_view.build_hierarchy(dichvu_after_load)
        if form_view_service == "Thu gọn":
            ma_dv_show, ten_dv_show = ldp_view.sorted_service_from_db(df_hierachy_2)
        else:
            ma_dv_show, ten_dv_show = ldp_view.sorted_service_from_db(df_hierachy_1)
    # FORM 1
        df_form_form_1,kehoach_col_map, thuchien_col_map = ldp_view.create_dataframe_thuchien_tt(selected_month,ma_dv_show,ten_dv_show)
        df_form_form_1 = ldp_view.map_kehoach_to_thuchien_tt(filtered_kehoach,df_form_form_1,kehoach_col_map,selected_month)
        df_form_form_1 = ldp_view.map_thuchien_tt_for_form(df_form_form_1[thuchien_col_map],filtered_thuchien_ht,df_form_form_1)
        df_form_form_1 = ldp_view.map_thuchien_tt_for_form(df_form_form_1["THỰC HIỆN T01"],filtered_thuchien_t1,df_form_form_1)
        df_form_form_1 = ldp_view.map_thuchien_tt_for_form(df_form_form_1["KỲ TRƯỚC"],filtered_thuchien_kytruoc,df_form_form_1)
        df_form_form_1 = ldp_view.update_summary_rows(df_form_form_1)
        df_form_form_1 = ldp_view.update_form_1_table(df_form_form_1,thuchien_col_map,kehoach_col_map)
    # FORM 2
        df_form2_new = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        df_form2_new = ldp_view.map_thuchien_to_form_2(form_2_filter,df_form2_new)
        df_form2_new = ldp_view.update_summary_rows(df_form2_new)
    # FORM 3
        df_form3_kehoach = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        df_form3_kehoach = ldp_view.map_kehoach_to_form_3(form3_kehoach_after_filter_no_month,df_form3_kehoach)
        df_form3_kehoach = ldp_view.update_summary_rows(df_form3_kehoach)
    
        if filtered_thuchien_ht.empty:
            st.warning("Không có dữ liệu thực hiện")
            return None,None,None
        elif filtered_kehoach.empty:
            st.warning("Không có dữ liệu kế hoạch")
            return None,None,None
        else:
            df_form3_kehoach.iloc[:,2:]= df_form3_kehoach.iloc[:,2:].astype(float).round(0) 
            df_form_form_1.iloc[:, 2:] = (df_form_form_1.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').fillna(0).round(0))
            df_form_form_1['% THỰC HIỆN'] = df_form_form_1['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
            df_form_form_1['% VỚI KỲ TRƯỚC'] = df_form_form_1['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
            df_form_form_1['% THỰC HIỆN T01'] = df_form_form_1['% THỰC HIỆN T01'].apply(lambda x: '{:.0f}%'.format(x))
            df_form2_new.iloc[:,2:]= df_form2_new.iloc[:,2:].astype(float).round(0)
            df_form2_new = df_form2_new.fillna(0).infer_objects(copy=False)
            return df_form_form_1,df_form2_new,df_form3_kehoach
    
    def map_data_for_year_view(self,line,selected_month,selected_year,selected_loaidoanhthu,form_view_service,radio_data_selected_kind):
        def single_filter_data_for_year(line,selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind):
            df_thuchien_single_nomonth,df_thuchien_luyke_nomonth = self.class_data_option_process.filter_thuchien_for_year_view(thuchien_after_load,None,selected_year,line,selected_loaidoanhthu,radio_data_selected_kind)
            df_thuchien_single,df_thuchien_luyke = self.class_data_option_process.filter_thuchien_for_year_view(thuchien_after_load,selected_month,selected_year,line,selected_loaidoanhthu,radio_data_selected_kind)
            df_kehoach_single_nomonth,df_kehoach_single_luyke_nomonth = self.class_data_option_process.filter_kehoach_for_year_view(kehoach_after_load,dichvu_after_load,line,None,selected_year,selected_loaidoanhthu,radio_data_selected_kind)
            df_kehoach_single,df_kehoach_single_luyke = self.class_data_option_process.filter_kehoach_for_year_view(kehoach_after_load,dichvu_after_load,line,selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind)
            # Thang truoc
            df_thuchien_single_kytruoc,df_thuchien_luyke_kytruoc = self.class_data_option_process.filter_thuchien_for_year_view(thuchien_after_load,selected_month - 1,selected_year,line,selected_loaidoanhthu,radio_data_selected_kind)
            # Cung ky
            df_thuchien_single_cungky,df_thuchien_luyke_cungky = self.class_data_option_process.filter_thuchien_for_year_view(thuchien_after_load,selected_month,selected_year - 1,line,selected_loaidoanhthu,radio_data_selected_kind)
            return df_thuchien_single_nomonth,df_thuchien_luyke_nomonth,df_thuchien_single,df_thuchien_luyke,df_kehoach_single_nomonth,df_kehoach_single_luyke_nomonth,df_kehoach_single,df_kehoach_single_luyke,df_thuchien_single_kytruoc,df_thuchien_luyke_kytruoc,df_thuchien_single_cungky,df_thuchien_luyke_cungky
        filter_df_year_sum_kehoach = self.class_data_option_process.filter_kehoach_year_sum(kehoach_after_load,dichvu_after_load,line,selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind)
        df_hierachy_1, df_hierachy_2 = ldp_view.build_hierarchy(dichvu_after_load)
        if form_view_service == "Thu gọn":
            ma_dv_show, ten_dv_show = ldp_view.sorted_service_from_db(df_hierachy_2)
        else:
            ma_dv_show, ten_dv_show = ldp_view.sorted_service_from_db(df_hierachy_1)
        if selected_loaidoanhthu != "tổng doanh thu":
            df_thuchien_single_nomonth,df_thuchien_luyke_nomonth,df_thuchien_single,df_thuchien_luyke,df_kehoach_single_nomonth,df_kehoach_single_luyke_nomonth,df_kehoach_single,df_kehoach_single_luyke,df_thuchien_single_kytruoc,df_thuchien_luyke_kytruoc,df_thuchien_single_cungky,df_thuchien_luyke_cungky = single_filter_data_for_year(line,selected_month,selected_year,selected_loaidoanhthu,radio_data_selected_kind)
        else:
            df_thuchien_single_nomonth_hh,df_thuchien_luyke_nomonth_hh,df_thuchien_single_hh,df_thuchien_luyke_hh,df_kehoach_single_nomonth_hh,df_kehoach_single_luyke_nomonth_hh,df_kehoach_single_hh,df_kehoach_single_luyke_hh,df_thuchien_single_kytruoc_hh,df_thuchien_luyke_kytruoc_hh,df_thuchien_single_cungky_hh,df_thuchien_luyke_cungky_hh = single_filter_data_for_year(line,selected_month,selected_year,"hiện hữu",radio_data_selected_kind)
            df_thuchien_single_nomonth_mtn,df_thuchien_luyke_nomonth_mtn,df_thuchien_single_mtn,df_thuchien_luyke_mtn,df_kehoach_single_nomonth_mtn,df_kehoach_single_luyke_nomonth_mtn,df_kehoach_single_mtn,df_kehoach_single_luyke_mtn,df_thuchien_single_kytruoc_mtn,df_thuchien_luyke_kytruoc_mtn,df_thuchien_single_cungky_mtn,df_thuchien_luyke_cungky_mtn = single_filter_data_for_year(line,selected_month,selected_year,"mới trong năm",radio_data_selected_kind)
            # FILTER FORM 1
            df_thuchien_single,df_thuchien_single_kytruoc,df_kehoach_single,df_thuchien_single_cungky = ldp_view.sum_hienhuu_moitrongnam(df_thuchien_single_hh,df_thuchien_single_mtn,df_thuchien_single_kytruoc_hh,df_thuchien_single_kytruoc_mtn,df_kehoach_single_hh,df_kehoach_single_mtn,df_thuchien_single_cungky_hh,df_thuchien_single_cungky_mtn)
            df_thuchien_luyke,df_thuchien_luyke_kytruoc,df_kehoach_single_luyke,df_thuchien_luyke_cungky = ldp_view.sum_hienhuu_moitrongnam(df_thuchien_luyke_hh,df_thuchien_luyke_mtn,df_thuchien_luyke_kytruoc_hh,df_thuchien_luyke_kytruoc_mtn,df_kehoach_single_luyke_hh,df_kehoach_single_luyke_mtn,df_thuchien_luyke_cungky_hh,df_thuchien_luyke_cungky_mtn)
            # FILTER FORM 2
            def sum_form_2_year_view(df_thuchien_single_nomonth_hh,df_thuchien_single_nomonth_mtn):
                df_thuchien_single_nomonth = pd.merge(
                        df_thuchien_single_nomonth_hh, 
                        df_thuchien_single_nomonth_mtn, 
                        on=['nhom_dv', 'thang', 'year_insert'], 
                        how='outer',
                        suffixes=('_hienhuu', '_moitrongnam')
                        )
                df_thuchien_single_nomonth["doanhthu_hienhuu"] = df_thuchien_single_nomonth["doanhthu_hienhuu"].fillna(0)
                df_thuchien_single_nomonth["doanhthu_moitrongnam"] = df_thuchien_single_nomonth["doanhthu_moitrongnam"].fillna(0)
                df_thuchien_single_nomonth["doanhthu"] = (df_thuchien_single_nomonth["doanhthu_hienhuu"] + df_thuchien_single_nomonth["doanhthu_moitrongnam"])
                df_thuchien_single_nomonth = df_thuchien_single_nomonth.groupby(['nhom_dv', 'thang', 'year_insert'], as_index=False)['doanhthu'].sum()
                return df_thuchien_single_nomonth
            df_thuchien_single_nomonth = sum_form_2_year_view(df_thuchien_single_nomonth_hh,df_thuchien_single_nomonth_mtn)
            df_thuchien_luyke_nomonth = sum_form_2_year_view(df_thuchien_luyke_nomonth_hh,df_thuchien_luyke_nomonth_mtn)
            # FILTER FORM 3
            month_columns_form3 = [col for col in df_kehoach_single_nomonth_hh.columns if (col.startswith("t") and col != "ten_dv")]
            df_kehoach_single_nomonth_hh[month_columns_form3] = df_kehoach_single_nomonth_hh[month_columns_form3].values + df_kehoach_single_nomonth_mtn[month_columns_form3].values
            df_kehoach_single_luyke_nomonth_hh[month_columns_form3] = df_kehoach_single_luyke_nomonth_hh[month_columns_form3].values + df_kehoach_single_luyke_nomonth_mtn[month_columns_form3].values
            df_kehoach_single_nomonth = df_kehoach_single_nomonth_hh
            df_kehoach_single_luyke_nomonth = df_kehoach_single_luyke_nomonth_hh
        # FORM 1 PHAN TICH
        df_form_form_1,df_form_form_1_luyke,kehoach_col_map, thuchien_col_map = ldp_view.create_dataframe_form_1_year_view(selected_month,ma_dv_show,ten_dv_show)
        df_form_form_1 = ldp_view.map_kehoach_to_thuchien_tt(df_kehoach_single,df_form_form_1,kehoach_col_map,selected_month)
        df_form_form_1 = ldp_view.map_kehoach_to_thuchien_tt(filter_df_year_sum_kehoach,df_form_form_1,"KẾ HOẠCH NĂM",selected_month)
        df_form_form_1 = ldp_view.map_thuchien_tt_for_form(df_form_form_1[thuchien_col_map],df_thuchien_single,df_form_form_1)
        df_form_form_1 = ldp_view.map_thuchien_tt_for_form(df_form_form_1["KỲ TRƯỚC"],df_thuchien_single_kytruoc,df_form_form_1)
        df_form_form_1 = ldp_view.map_thuchien_tt_for_form(df_form_form_1["CÙNG KỲ"],df_thuchien_single_cungky,df_form_form_1)
        df_form_form_1 = ldp_view.update_summary_rows(df_form_form_1)
        df_form_form_1 = ldp_view.update_form_1_year_view(df_form_form_1,thuchien_col_map,kehoach_col_map)
        df_form_form_1_luyke = ldp_view.map_kehoach_to_thuchien_tt(df_kehoach_single_luyke,df_form_form_1_luyke,kehoach_col_map,selected_month)
        df_form_form_1_luyke = ldp_view.map_kehoach_to_thuchien_tt(filter_df_year_sum_kehoach,df_form_form_1_luyke,"KẾ HOẠCH NĂM",selected_month)
        df_form_form_1_luyke = ldp_view.map_thuchien_tt_for_form(df_form_form_1_luyke[thuchien_col_map],df_thuchien_luyke,df_form_form_1_luyke)
        df_form_form_1_luyke = ldp_view.map_thuchien_tt_for_form(df_form_form_1_luyke["CÙNG KỲ"],df_thuchien_luyke_cungky,df_form_form_1_luyke)
        df_form_form_1_luyke = ldp_view.update_summary_rows(df_form_form_1_luyke)
        df_form_form_1_luyke = ldp_view.update_form_1_year_view(df_form_form_1_luyke,thuchien_col_map,kehoach_col_map)
        # FORM 2 THUCHIEN
        df_form2_new = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        df_form2_new = ldp_view.map_thuchien_to_form_2(df_thuchien_single_nomonth,df_form2_new)
        df_form2_new = ldp_view.update_summary_rows(df_form2_new)
        df_form2_new_luyke = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        df_form2_new_luyke = ldp_view.map_thuchien_to_form_2(df_thuchien_luyke_nomonth,df_form2_new_luyke)
        df_form2_new_luyke = ldp_view.update_summary_rows(df_form2_new_luyke)
        # FORM 3 KEHOACH
        df_form3_kehoach = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        df_form3_kehoach = ldp_view.map_kehoach_to_form_3(df_kehoach_single_nomonth,df_form3_kehoach)
        df_form3_kehoach = ldp_view.update_summary_rows(df_form3_kehoach)
        df_form3_kehoach_luyke = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        df_form3_kehoach_luyke = ldp_view.map_kehoach_to_form_3(df_kehoach_single_luyke_nomonth,df_form3_kehoach_luyke)
        df_form3_kehoach_luyke = ldp_view.update_summary_rows(df_form3_kehoach_luyke)
        # FORM 4 % THUC HIEN
        df_form4_thuchien = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        df_form4_thuchien_luyke = ldp_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
        columns_percentage = [col for col in df_form4_thuchien.columns if col.startswith("Tháng")]
        
        df_form4_thuchien[columns_percentage] = (df_form2_new[columns_percentage].div(df_form3_kehoach[columns_percentage].replace(0, np.nan)) * 100).fillna(0)
        df_form4_thuchien_luyke[columns_percentage] = (df_form2_new_luyke[columns_percentage].div(df_form3_kehoach_luyke[columns_percentage].replace(0, np.nan)) * 100).fillna(0)
        if df_thuchien_single.empty:
            st.warning("Không có dữ liệu thực hiện")
            return None,None,None,None,None,None,None,None
        elif df_kehoach_single.empty:
            st.warning("Không có dữ liệu kế hoạch")
            return None,None,None,None,None,None,None,None
        else:
            df_form3_kehoach.iloc[:,2:]= df_form3_kehoach.iloc[:,2:].astype(float).round(2)
            df_form3_kehoach_luyke.iloc[:,2:]= df_form3_kehoach_luyke.iloc[:,2:].astype(float).round(2)
            df_form_form_1.iloc[:, 2:] = (df_form_form_1.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').fillna(0).round(2))
            df_form_form_1['% THỰC HIỆN'] = df_form_form_1['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
            df_form_form_1['% VỚI KỲ TRƯỚC'] = df_form_form_1['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
            df_form_form_1_luyke.iloc[:, 2:] = (df_form_form_1_luyke.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').fillna(0).round(2))
            df_form_form_1_luyke['% THỰC HIỆN'] = df_form_form_1_luyke['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
            df_form2_new.iloc[:,2:]= df_form2_new.iloc[:,2:].astype(float).round(2)
            df_form2_new = df_form2_new.fillna(0).infer_objects(copy=False)
            df_form2_new_luyke.iloc[:,2:]= df_form2_new_luyke.iloc[:,2:].astype(float).round(2)
            df_form2_new_luyke = df_form2_new_luyke.fillna(0).infer_objects(copy=False)
            df_form4_thuchien = df_form4_thuchien.fillna(0).infer_objects(copy=False)
            df_form4_thuchien_luyke = df_form4_thuchien_luyke.fillna(0).infer_objects(copy=False)
            df_form4_thuchien = df_form4_thuchien.replace([np.inf, -np.inf], 0)
            df_form4_thuchien_luyke = df_form4_thuchien_luyke.replace([np.inf, -np.inf], 0)
            df_form4_thuchien[columns_percentage] = df_form4_thuchien[columns_percentage].map(lambda x: '{:.0f}%'.format(x) if x != 0 else 0)
            df_form4_thuchien_luyke[columns_percentage] = df_form4_thuchien_luyke[columns_percentage].map(lambda x: '{:.0f}%'.format(x) if x != 0 else 0)
            return df_form_form_1,df_form2_new,df_form3_kehoach,df_form_form_1_luyke,df_form2_new_luyke,df_form3_kehoach_luyke,df_form4_thuchien,df_form4_thuchien_luyke

        
        
    def month_table_view_gen(self,export_excel_table_file,submit_view_option,radio_type_view_table,form_view_service,selected_year,selected_month,selected_line_access,selected_loaidoanhthu,radio_data_selected_kind):
        if radio_type_view_table == "Đơn":
            if submit_view_option:
                with st.spinner("Đang tải dữ liệu..."):
                    df_1,df_2,df_3 = self.map_data_for_month_view(selected_line_access,selected_month,selected_year,selected_loaidoanhthu,form_view_service,radio_data_selected_kind)
                    # Áp dụng định dạng vào bảng
                    if df_1 is not None:
                        styled_df = df_1.style \
                            .map(self.highlight_nonzero) \
                            .map(self.highlight_numbers, subset=["+/- VỚI KỲ TRƯỚC","+ /- VỚI T01"]) \
                            .format(self.add_arrow, subset=["+/- VỚI KỲ TRƯỚC","+ /- VỚI T01"]) \
                            .map(self.highlight_numbers_percentage, subset=["% THỰC HIỆN","% THỰC HIỆN T01","% VỚI KỲ TRƯỚC"]) \
                            .format(self.add_arrow_percentage, subset=["% THỰC HIỆN","% THỰC HIỆN T01","% VỚI KỲ TRƯỚC"]) \
                            .format("{:.2f}", subset=[f"KẾ HOẠCH T0{selected_month}",f"THỰC HIỆN T0{selected_month}","THỰC HIỆN T01","KỲ TRƯỚC"] if selected_month != 1 else [f"KẾ HOẠCH T0{selected_month}",f"THỰC HIỆN T{selected_month}","THỰC HIỆN T01","KỲ TRƯỚC"])
                            
                        df_2 = df_2.style.map(self.highlight_nonzero) \
                            .format("{:.2f}",subset= df_2.columns[2:])
                        df_3 = df_3.style.map(self.highlight_nonzero) \
                            .format("{:.2f}",subset= df_3.columns[2:])
                            
                        container_df_view_type_table = st.container(key="container_df_view_type_table")
                        with container_df_view_type_table:
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Kểt quả phân tích</h6>""", unsafe_allow_html=True)
                            st.dataframe(styled_df)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số thực hiện các tháng</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_2)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số kế hoạch</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_3)
                    else:
                        st.warning("Vui lòng kiểm tra lại dữ liệu") 
            if export_excel_table_file:
                file_name = f"Kết quả tháng {selected_month} năm {selected_year} line {selected_line_access}_{selected_loaidoanhthu}"
                self.download_excel_single(file_name,selected_year,selected_month,selected_loaidoanhthu,selected_line_access,radio_type_view_table,form_view_service,radio_data_selected_kind)
        else:
            if submit_view_option or export_excel_table_file:
                file_name = f"Kết quả tháng {selected_month} năm {selected_year} kết hợp_{selected_loaidoanhthu}"
                self.download_excel_single(file_name,selected_year,selected_month,selected_loaidoanhthu,None,radio_type_view_table,form_view_service,radio_data_selected_kind)
    def year_table_view_gen(self,export_excel_table_file,submit_view_option,radio_type_view_table,form_view_service,selected_year,selected_month,selected_line_access,selected_loaidoanhthu,radio_data_selected_kind):
        if radio_type_view_table == "Đơn":
            if submit_view_option:
                with st.spinner("Đang tải dữ liệu..."):
                    df_1,df_2,df_3,df_1_luyke,df_2_luyke,df_3_luyke,df_4,df_4_luyke = self.map_data_for_year_view(selected_line_access,selected_month,selected_year,selected_loaidoanhthu,form_view_service,radio_data_selected_kind)
                    if df_1 is not None:
                        container_df_view_type_table = st.container(key="container_df_view_type_table")
                        with container_df_view_type_table:
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Kểt quả phân tích</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_1)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Kểt quả phân tích lũy kế</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_1_luyke)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số thực hiện các tháng</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_2)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số thực hiện các tháng lũy kế</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_2_luyke)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số kế hoạch</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_3)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số kế hoạch lũy kế</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_3_luyke)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">% Thực hiện</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_4)
                            st.markdown("""<h6 style="text-align:center;font-weight:bold;">% Thực hiện lũy kế</h6>""", unsafe_allow_html=True)
                            st.dataframe(df_4_luyke)
                    else:
                        st.warning("Vui lòng kiểm tra lại dữ liệu")
            if export_excel_table_file:
                file_name = f"Kết quả năm {selected_year} line {selected_line_access}_{selected_loaidoanhthu}"
                self.download_excel_year_view(file_name,selected_year,selected_month,selected_line_access,radio_type_view_table,form_view_service,radio_data_selected_kind)

            
        else: 
            if submit_view_option or export_excel_table_file:
                file_name = f"Kết quả năm {selected_year} kết hợp"
                self.download_excel_year_view(file_name,selected_year,selected_month,selected_line_access,radio_type_view_table,form_view_service,radio_data_selected_kind)

class MAIN_VIEW():
    def __init__(self):
        self.class_data_option_process = VIEW_LDP_DATA_PROCESS()
    def sidebar_viewldp(self):
        with st.sidebar:
            selected = option_menu(
                    menu_title= None,  # required
                    options=["Dạng bảng theo tháng", "Dạng bảng tổng hợp"],  # required
                    icons=["calendar2-range-fill", "calendar3"],  
                    menu_icon= None,  
                    default_index=0,  
                    orientation="vertical",  
                    key="menu_sidebar_delete",
                    styles={
                    "container": {
                        "padding": "0px 5px", 
                        "max-width": "100%",
                        "margin": "0px auto",  
                        "border": "None",
                        "border-radius": "20px",
                        "background-color": "rgb(120 ,189, 243)",
                    },
                    "icon": {
                        "color": "#fff",  
                        "font-size": "0.8rem",
                        "font-weight": "bold",
                    },
                    "nav-link": {
                        "font-size": "0.8rem", 
                        "text-align": "left",  
                        "--hover-color": "#54a7ef",
                        "font-weight": "bold",
                    },
                    "nav-link-selected": {
                        "border-radius": "15px",
                        "background-color": "#7FC8F8005E7C", 
                        "font-size": "0.8rem",
                        "font-family": "Tahoma, Geneva, sans-serif",
                        
                        
                    }
                }
                    )
        return selected
    def run_view(self):
        selected = self.sidebar_viewldp()
        if selected == "Dạng bảng theo tháng":
            header_page = "Dạng bảng theo tháng"
            radio_data_selected_kind = "LINE"
        elif selected == "Dạng bảng tổng hợp":
            header_page = "Dạng bảng tổng hợp"
            radio_data_selected_kind = "LDPVNPT"   
        export_excel_table_file,submit_view_option,radio_type_view_table,form_view_service,selected_year,selected_month,selected_line_access,selected_loaidoanhthu = self.class_data_option_process.ui_info_no_search(selected,header_page)
        if selected == "Dạng bảng theo tháng":
            GENERATE_VIEW().month_table_view_gen(export_excel_table_file,submit_view_option,radio_type_view_table,form_view_service,selected_year,selected_month,selected_line_access,selected_loaidoanhthu,radio_data_selected_kind)
        else:
            GENERATE_VIEW().year_table_view_gen(export_excel_table_file,submit_view_option,radio_type_view_table,form_view_service,selected_year,selected_month,selected_line_access,selected_loaidoanhthu,radio_data_selected_kind)
                
thuchien_after_load, kehoach_after_load, nhanvien_after_load, dichvu_after_load,line_after_load = ldp_view.load_data_ldp()
MAIN_VIEW().run_view()
module_config.add_sidebar_footer()