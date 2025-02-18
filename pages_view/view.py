import streamlit as st
import pandas as pd
import base64
from streamlit_option_menu import option_menu
import numpy as np
import io
from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook
import openpyxl
import time
import PROJECTS.module_view as module_view
import PROJECTS.config as module_config
import streamlit.components.v1 as components

# PART LOGIN 
if not st.session_state.get("is_logged_in", False):
    with st.spinner("🔐 Đang chuyển hướng đến trang đăng nhập..."):
        time.sleep(2)
    st.session_state.is_logged_in = False
    st.session_state.role_access_admin = False
    st.session_state.line_access = None
    st.switch_page("main.py")
    st.stop() 
# PART SET CONFIG
with open('src/style.css', encoding="utf-8")as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html = True) 
with open('src/style_general.css', encoding="utf-8")as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html = True)
with open('src/style_view.css', encoding="utf-8")as f:
    st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html = True)


# PART DESIGN FRONTEND
class DESIGN_FRONTEND():
    def __init__(self):
        pass
    def ui_info(self):
        st.header("")
        st.markdown("## **Báo cáo theo dịch vụ và nhân viên**")
    def ui_info_no_search(self):
        container_title_table_view = st.container(key="container_title_table_view")
        with container_title_table_view:
            col_title_dmanage_expand = st.columns([2,2.2])
            col_title_dmanage_expand[0].markdown(f"""<h3 style='text-align: left; padding:0'>Xem dạng bảng</h3>""", unsafe_allow_html=True)
            with col_title_dmanage_expand[1]:
                cols_button_view_table = st.columns([1,1,1])
                with cols_button_view_table[0]:
                   with st.popover("Chế độ",use_container_width=True, icon=":material/tune:"):
                        radio_type_view_table = st.radio("Chọn chế độ",["Đơn","Kết hợp"], key="radio_type_view_table")
                with cols_button_view_table[1]:
                    export_excel_table_file = st.button("Xuất Excel", key="button_export_excel", icon=":material/cloud_download:", use_container_width=True)
                with cols_button_view_table[2]:
                    if radio_type_view_table == "Đơn":
                        submit_view_option = st.button("Xem", key="submit_view_option", type="primary" ,icon=":material/visibility:", use_container_width=True)
                    else:
                        submit_view_option = st.button("Tạo file", key="submit_view_option", type="primary" ,icon=":material/send_money:", use_container_width=True)
        return radio_type_view_table,export_excel_table_file,submit_view_option
    def streamlit_menu_sidebar(self):
        # st.sidebar.write("Chế độ xem:")
        # with st.sidebar:
        selected = option_menu(
            menu_title= None,  # required
            options=["Board", "Table"],  # required
            icons=["clipboard2-data", "table"],  # optional
            menu_icon= None,  # optional
            default_index=0,  # optional
            orientation="horizontal",
            key="menu_option",
            styles={
        "container": {
            "padding": "0px 5px", 
            "max-width": "30%",
            "margin": "0px auto",  
            "border": "None",
            "border-radius": "20px",
            "background-color": "rgb(120 ,189, 243)",
            "font-weight": "600"
        },
        "icon": {
            "color": "#fff",  
            "font-size": "12px",
            "font-weight": "bold",
        },
        "nav-link": {
            "font-size": "12px", 
            "text-align": "left",  
            "--hover-color": "#54a7ef",
            "font-weight": "bold",
        },
        "nav-link-selected": {
            "border-radius": "15px",
            "background-color": "#7FC8F8005E7C", 
            "font-size": "12px",
            "font-family": "Tahoma, Geneva, sans-serif",
            
        }
        }
        )
        return selected
    def sidebar_option(self,cols_head_select_option_revenue):
        selected = cols_head_select_option_revenue[0].pills(" ",["👩‍💻 Nhân viên", "🌁 Dịch vụ"],default="👩‍💻 Nhân viên",key="pills_em_select_dashboard")
        st_toggle_check = st.sidebar.empty()
        if selected == "👩‍💻 Nhân viên":
            if st_toggle_check.toggle("Chọn nhiều tháng", False, key="select_multiple_month"):
                selected_months = st.sidebar.multiselect("Chọn tháng", range(1, 13), default=[1], help="Chọn nhiều tháng để xem dữ liệu")
            else:
                selected_months = st.sidebar.multiselect("Chọn tháng", range(1, 13), default=[1], max_selections=1, help="Bạn chỉ được chọn 1 tháng")
        else:
            st_toggle_check.empty()
            selected_months = st.sidebar.multiselect("Chọn tháng", range(1, 13), default=[1], max_selections=1, help="Bạn chỉ được chọn 1 tháng")
        selected_loaidoanhthu = st.sidebar.selectbox("Chọn loại doanh thu", thuchien_after_load["loaidoanhthu"].unique())
        line_nv = st.session_state.line_access
        if (selected == "👩‍💻 Nhân viên"):
            nv_mapping = dict(zip(nhanvien_after_load["ma_nv"], nhanvien_after_load["ten_nv"]))
            selected_nv_name = st.sidebar.selectbox("Chọn nhân viên", nhanvien_after_load["ten_nv"].unique())
            selected_data_kind = nhanvien_after_load[nhanvien_after_load["ten_nv"] == selected_nv_name]["ma_nv"].values[0]
        else:
            filtered_dichvu = dichvu_after_load[dichvu_after_load["danh_muc_tt"].notna()]
            filtered_dichvu = filtered_dichvu[~filtered_dichvu["danh_muc_tt"].isin(["1", "1.1.01", "1.1.02", "1.1.03", "1.1.05", "1.1.06", "1.1.07"])]
            selected_dv = st.sidebar.selectbox("Chọn dịch vụ", filtered_dichvu["ten_dv"].unique())
            dv_mapping = dict(zip(dichvu_after_load["ten_dv"], dichvu_after_load["ma_dv_id66"]))
            selected_data_kind = dv_mapping[selected_dv]
        selected_year = st.sidebar.selectbox("Chọn năm", kehoach_after_load["year_insert"].astype(int).unique())
        return selected,selected_months,selected_year,selected_loaidoanhthu,line_nv,selected_data_kind
    def table_sidebar_option(self,radio_type_view_table,cols_head_select_option_revenue):
        selected = cols_head_select_option_revenue[0].pills("🌀",["👩‍💻 Nhân viên", "🌁 Dịch vụ"], key="table_option_view_pills",default="👩‍💻 Nhân viên")
        container_sidebar_table = st.sidebar.container(key="container_sidebar_table")
        with container_sidebar_table:
            selected_months = st.selectbox("Chọn tháng",range(1,13), key="selected_months_table")
            selected_loaidoanhthu = st.selectbox("Chọn loại doanh thu", thuchien_after_load["loaidoanhthu"].unique(), key="loaidoanhthu_table_option")
            line_nv = st.session_state.line_access
            empty_select_box = st.empty()
            if radio_type_view_table == "Đơn":
                if (selected == "👩‍💻 Nhân viên"):
                    nv_mapping = dict(zip(nhanvien_after_load["ma_nv"], nhanvien_after_load["ten_nv"]))
                    selected_nv_name = empty_select_box.selectbox("Chọn nhân viên", nhanvien_after_load["ten_nv"].unique(), key="selected_manv_table")
                    selected_data_kind = nhanvien_after_load[nhanvien_after_load["ten_nv"] == selected_nv_name]["ma_nv"].values[0]
                else:
                    filtered_dichvu = dichvu_after_load[dichvu_after_load["danh_muc_tt"].notna()]
                    filtered_dichvu = filtered_dichvu[~filtered_dichvu["danh_muc_tt"].isin(["1", "1.1.01", "1.1.02", "1.1.03", "1.1.05", "1.1.06", "1.1.07"])]
                    selected_dv = empty_select_box.selectbox("Chọn dịch vụ", filtered_dichvu["ten_dv"].unique(), key="selected_dv_table")
                    dv_mapping = dict(zip(dichvu_after_load["ten_dv"], dichvu_after_load["ma_dv_id66"]))
                    selected_data_kind = dv_mapping[selected_dv]
            else:
                empty_select_box.empty()
                selected_data_kind = None
            selected_year = st.selectbox("Chọn năm", kehoach_after_load["year_insert"].astype(int).unique(), key="year_insert_table_view")
        return selected,selected_months,selected_year,selected_loaidoanhthu,line_nv,selected_data_kind
    def employee_design_frontend(self, selected_months,line_nv, selected_year, selected_loaidoanhthu, selected_ma_nv):
        # Lọc dữ liệu
        with st.spinner("💱 Đang tải dữ liệu..."):
            filtered_thuchien = thuchien_after_load[
            (thuchien_after_load["IDnhanvien"] == selected_ma_nv) &
            (thuchien_after_load["thang"].isin(selected_months)) &
            (thuchien_after_load["year_insert"] == int(selected_year)) &
            (thuchien_after_load["loaidoanhthu"] == selected_loaidoanhthu) &
            (thuchien_after_load["line"] == line_nv)
            ]
            filtered_thuchien.loc[:, "ten_dv"] = filtered_thuchien["nhom_dv"].map(dichvu_after_load.set_index("ma_dv_id66")["ten_dv"])

            filtered_kehoach = kehoach_after_load[
            (kehoach_after_load["ma_nv"] == selected_ma_nv) &
            (kehoach_after_load["year_insert"] == int(selected_year)) &
            (kehoach_after_load["line"] == line_nv) &
            (kehoach_after_load["loaidoanhthu"] == selected_loaidoanhthu)
            ]
            container_main_vertical_block = st.container(key="container_main_vertical_block")
            with container_main_vertical_block:
                # CONTAINER 1
                container_header_metric = st.container(key="container_header_service")
                with container_header_metric:
                    metric1,metric2,metric3= module_view.container_header_metric(filtered_thuchien,filtered_kehoach,selected_months)
                # CONTAINER 2
                container_first_employee = st.container(border=False)
                with container_first_employee:
                    cols_first_employee = st.columns([15,30,15])
                    with cols_first_employee[0]:
                        # Biểu đồ tròn
                        module_view.container_first_piechart(thuchien_after_load,nhanvien_after_load,line_nv, selected_months, selected_year)
                        
                with cols_first_employee[1]:
                        # TreeMap
                        module_view.container_first_treemap(filtered_thuchien)
                with cols_first_employee[2]:
                        # Process Column
                    module_view.container_first_process_column(dichvu_after_load,filtered_thuchien,filtered_kehoach, selected_months)
                # CONTAINER 3
                container_second_employee = st.container(border=False)
                with container_second_employee:
                    cols_second_employee = st.columns([15,30,15])
                    with cols_second_employee[0]:
                        # Biểu đồ Donut
                        module_view.container_second_donut_chart(metric1, metric2)
                    with cols_second_employee[1]:
                        # Biểu đồ cột
                        module_view.container_second_barchart(thuchien_after_load,line_nv,selected_ma_nv,selected_loaidoanhthu,selected_year)
                    with cols_second_employee[2]:
                        # DESCRIPTION DASHBOARD
                        st.markdown("""
                        <h6 style="text-align:center;font-weight:bold;"> Mô tả dashboard 🤔 </h6>
                        <p style="text-align:justify;"> Dashboard này giúp bạn theo dõi:</p>
                        <ul style="text-align:left;">
                            <li>Doanh thu của nhân viên theo từng dịch vụ, từng tháng.</li>
                            <li>Tỷ lệ hoàn thành so với kế hoạch.</li>
                            <li>Biểu đồ cột thể hiện doanh thu qua các tháng</li>
                        </ul>
                        """, unsafe_allow_html=True)

                # CONTAINER 4
                container_third_employee = st.container(border=False)
                with container_third_employee:
                    cols_third_employee = st.columns([46,15])
                    with cols_third_employee[0]:
                        # Biểu đồ cột
                        module_view.container_third_barchart(filtered_thuchien,dichvu_after_load)

                    with cols_third_employee[1]:
                        st.markdown("""
                        <h6 style="text-align:center;font-weight:bold;"> Hơn nữa 🤔 </h6>
                        <p style="text-align:justify;"> Dashboard này còn có các biểu đồ:</p>
                        <ul style="text-align:left;">
                            <li>Biểu đồ tròn tỷ lệ doanh thu giữa các thành viên trong line</li>
                            <li>Biểu đồ treemap thể hiện tỷ lệ doanh thu của các dịch vụ của nhân viên đó.</li>
                            <li>Biểu đồ cột ghép - thể hiện doanh thu theo các tháng của các nhóm dịch vụ</li>
                        </ul>
                        """, unsafe_allow_html=True)
                
    def service_design_frontend(self, selected_months,line_nv, selected_year, selected_loaidoanhthu, selected_ma_dv):
        with st.spinner("💱 Đang tải dữ liệu..."):
            selected_ten_dv = dichvu_after_load[dichvu_after_load["ma_dv_id66"] == selected_ma_dv]["ten_dv"].values[0]
            filtered_thuchien_dv = thuchien_after_load[
            (thuchien_after_load["nhom_dv"] == selected_ma_dv) &
            (thuchien_after_load["thang"].isin(selected_months)) &
            (thuchien_after_load["year_insert"] == selected_year) &
            (thuchien_after_load["loaidoanhthu"] == selected_loaidoanhthu) &
            (thuchien_after_load["line"] == line_nv)
        ]

            filtered_kehoach_dv = kehoach_after_load[
                (kehoach_after_load["id_dv_606"] == selected_ma_dv) &
                (kehoach_after_load["year_insert"] == int(selected_year)) &
                (kehoach_after_load["line"] == line_nv)&
                (kehoach_after_load["loaidoanhthu"] == selected_loaidoanhthu)
            ]
            # CONTAINER MAIN
            container_main_service_design = st.container(key="container_main_service_design")
            with container_main_service_design:
                # CONTAINER 1
                container_header_services = st.container(key="container_header_services")
                with container_header_services:
                    module_view.container_services_header(filtered_thuchien_dv,filtered_kehoach_dv,selected_months,selected_ten_dv,selected_ma_dv,selected_loaidoanhthu,selected_year,thuchien_after_load,kehoach_after_load,line_nv)
                # CONTAINER 2
                container_first_services = st.container(key="container_first_services")
                with container_first_services:
                    cols_first_services = st.columns([15, 30])
                    with cols_first_services[0]:
                        module_view.container_first_services_piechart(thuchien_after_load,dichvu_after_load,line_nv,selected_months, selected_year)
                    with cols_first_services[1]:
                        module_view.container_first_services_barchart(thuchien_after_load,selected_ma_dv,selected_loaidoanhthu,selected_year,line_nv)
                # CONTAINER 3
                container_second_services = st.container(key="container_second_services")
                with container_second_services:
                    cols_second_services = st.columns([46, 15])
                    with cols_second_services[0]:
                        module_view.container_second_table_services(thuchien_after_load,nhanvien_after_load,selected_months,selected_ma_dv,selected_loaidoanhthu,selected_year,line_nv)
                    
                    with cols_second_services[1]:
                        image_dashboard = module_config.get_relative_file_path("../src/dashboard.png")
                        st.markdown(f"""
                        <h6 style="text-align:center;font-weight:bold;"> Dashboard mô tả dịch vụ</h6>
                        <img src="data:image/png;base64,{image_dashboard}" style="display:block; margin-left:auto; margin-right:auto; width:30%;">
                        <ul style="text-align:left;">
                            <li>Các metric </li>
                            <li>Biểu đồ tròn tỷ lệ dịch vụ với nhau trong tháng</li>
                            <li>Biểu đồ cột ghép - thể hiện doanh thu và tỷ lệ tăng trưởng</li>
                            <li>Bảng đóng góp của từng thành viên trong line đối với dịch vụ tháng đó.</li>
                        </ul>
                        """, unsafe_allow_html=True)
    def table_employee_design(self,selected_option, selected_months,line_nv, selected_year, selected_loaidoanhthu, selected_data_kind,radio_type_view_table):
        if radio_type_view_table == "Đơn":
            container_info = st.container(key="container_info_table_view")
            with container_info:
                cols_info = st.columns([0.75,1,0.75])
                with cols_info[0]:
                    container_column_table_view_1 = st.container(key="container_column_table_view_1")
                    if selected_option == "👩‍💻 Nhân viên":
                        container_column_table_view_1.markdown(f"""###### Tên nhân viên: <span style="color:#044B7F;">{nhanvien_after_load[nhanvien_after_load['ma_nv'] == selected_data_kind]['ten_nv'].values[0]}</span>""", unsafe_allow_html=True)
                    else:
                        container_column_table_view_1.markdown(f"""###### Tên dịch vụ: <span style="color:#044B7F;">{dichvu_after_load[dichvu_after_load['ma_dv_id66'] == selected_data_kind]['ten_dv'].values[0]}</span>""", unsafe_allow_html=True)
                    container_column_table_view_1.markdown(f"**Năm:** {selected_year}")
                    
                with cols_info[1]:
                    container_column_table_view_2 = st.container(key="container_column_table_view_2")
                    container_column_table_view_2.markdown(f"""###### **Line:**  <span style="color:#044B7F;">{line_after_load[line_after_load["ma_line"] == line_nv]["ten_line"].values[0]}</span>""",unsafe_allow_html=True)
                    container_column_table_view_2.markdown(f"**Tháng chọn:** {selected_months}")
                    
                with cols_info[2]:
                    container_column_table_view_3 = st.container(key="container_column_table_view_3")
                    if selected_option == "👩‍💻 Nhân viên":
                        container_column_table_view_3.markdown(f"""###### **Mã nhân viên:**  <span style="color:#044B7F;">{selected_data_kind}</span>""", unsafe_allow_html=True)
                    else:
                        container_column_table_view_3.markdown(f"""###### **Mã dịch vụ:**  <span style="color:#044B7F;">{selected_data_kind}</span>""", unsafe_allow_html=True)
                    container_column_table_view_3.markdown(f"**Loại doanh thu:** {selected_loaidoanhthu}")
        
class BACKEND_TABLE_VIEW():
    def __init__(self):
        pass
    def highlight_numbers(self,val):
        """Định dạng màu sắc cho số âm/dương"""
        try:
            val = float(val)
        except ValueError:
            return "color: black;"
        if val > 0:
            return "color: green;" 
        elif val < 0:
            return "color: red;"    
        else:
            return "color: black;"
    def highlight_numbers_percentage(self,val):
        """Định dạng màu sắc cho số âm/dương"""
        try:
            val = float(val.strip('%'))
        except ValueError:
            return "color: black;"
        if (val - 100) > 0 and val != 100:
            return "color: green;" 
        elif (val - 100) < 0 and val != 100:
            return "color: red;"    
        else:
            return "color: black;"  
    def highlight_nonzero(self,val):
            if val == 0:
                 color = '#A6988D' 
            elif val == "0%":
                color = '#A6988D' 
            else:
                color = '#0D0D0D'  
            return f'color: {color}'
    def add_arrow(self,val):
        """Thêm mũi tên cho số âm/dương"""
        try:
            val = float(val)
        except ValueError:
            return "color: black;"
        if val > 0:
            return f"↑ {val}"
        elif val < 0:
            return f"↓ {val}" 
        else:
            return val
    def add_arrow_percentage(self,val):
        """Thêm mũi tên vào giá trị phần trăm"""
        try:
            val = float(val.strip('%'))
        except ValueError:
            return f"{val}%"  # Return as is if conversion fails
        if (val - 100) > 0 and val != 100:
            return f"↑ {val:.1f}%"  # Mũi tên lên và định dạng %.1f
        elif (val - 100) < 0 and val != 100:
            return f"↓ {val:.1f}%"  # Mũi tên xuống
        else:
            return f"{val:.1f}%"  
    def download_xlsx_file_from_form(self, styled_df):
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            styled_df.to_excel(writer, index=False)
        output.seek(0)
        return output
    def table_generate_backend(self, selected_months,line_nv, selected_year, selected_loaidoanhthu, selected_data_kind,selected_option, radio_type_view_table, action_check_view, action_check_export):
        def action_view_for_loop_func(selected_data_kind):
            if selected_option == "👩‍💻 Nhân viên":
            #    PART GET DATA
                filtered_kehoach_employee = module_view.filter_kehoach_by_employee(kehoach_after_load,dichvu_after_load,selected_data_kind,line_nv,selected_months,selected_year,selected_loaidoanhthu)
                filtered_thuchien_employee_ht = module_view.filter_data_thuchien_by_employee(thuchien_after_load,selected_data_kind,selected_months,selected_year,line_nv,selected_loaidoanhthu)
                filtered_thuchien_employee_t1 = module_view.filter_data_thuchien_by_employee(thuchien_after_load,selected_data_kind,1,selected_year,line_nv,selected_loaidoanhthu)
                filtered_thuchien_employee_kytruoc = module_view.filter_data_thuchien_by_employee(thuchien_after_load,selected_data_kind,selected_months -1,selected_year,line_nv,selected_loaidoanhthu)
                form_2_filter_employee = module_view.filter_data_thuchien_by_employee(thuchien_after_load,selected_data_kind,None,selected_year,line_nv,selected_loaidoanhthu)
                form3_kehoach_after_filter_no_month = module_view.filter_kehoach_by_employee(kehoach_after_load,dichvu_after_load,selected_data_kind,line_nv,None,selected_year,selected_loaidoanhthu)
                df_hierachy_1, df_hierachy_2 = module_view.build_hierarchy(dichvu_after_load)
                ma_dv_show, ten_dv_show = module_view.sorted_service_from_db(df_hierachy_2)
            # FORM 1
                df_form_employee_form_1,kehoach_col_map, thuchien_col_map = module_view.create_dataframe_thuchien_tt(selected_months,ma_dv_show,ten_dv_show)
                df_form_employee_form_1 = module_view.map_kehoach_to_thuchien_tt(filtered_kehoach_employee,df_form_employee_form_1,kehoach_col_map,selected_months)
                df_form_employee_form_1 = module_view.map_thuchien_tt_for_form(df_form_employee_form_1[thuchien_col_map],filtered_thuchien_employee_ht,df_form_employee_form_1)
                df_form_employee_form_1 = module_view.map_thuchien_tt_for_form(df_form_employee_form_1["THỰC HIỆN T01"],filtered_thuchien_employee_t1,df_form_employee_form_1)
                df_form_employee_form_1 = module_view.map_thuchien_tt_for_form(df_form_employee_form_1["KỲ TRƯỚC"],filtered_thuchien_employee_kytruoc,df_form_employee_form_1)
                df_form_employee_form_1 = module_view.update_summary_rows(df_form_employee_form_1)
                df_form_employee_form_1 = module_view.update_form_1_table(df_form_employee_form_1,thuchien_col_map,kehoach_col_map)
            # FORM 2
                df_form2_employee_new = module_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
                df_form2_employee_new = module_view.map_thuchien_to_form_2(form_2_filter_employee,df_form2_employee_new)
                df_form2_employee_new = module_view.update_summary_rows(df_form2_employee_new)
            # FORM 3
                df_form3_employee_kehoach = module_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
                df_form3_employee_kehoach = module_view.map_kehoach_to_form_3_employee(form3_kehoach_after_filter_no_month,df_form3_employee_kehoach)
                df_form3_employee_kehoach = module_view.update_summary_rows(df_form3_employee_kehoach)
            
                if filtered_thuchien_employee_ht.empty:
                    st.warning("Không có dữ liệu thực hiện")
                    return None,None,None
                else:
                    df_form3_employee_kehoach.iloc[:,2:]= df_form3_employee_kehoach.iloc[:,2:].astype(float).round(0) 
                    df_form_employee_form_1.iloc[:, 2:] = (df_form_employee_form_1.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').fillna(0).round(0))
                    df_form_employee_form_1['% THỰC HIỆN'] = df_form_employee_form_1['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
                    df_form_employee_form_1['% VỚI KỲ TRƯỚC'] = df_form_employee_form_1['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
                    df_form_employee_form_1['% THỰC HIỆN T01'] = df_form_employee_form_1['% THỰC HIỆN T01'].apply(lambda x: '{:.0f}%'.format(x))
                    df_form2_employee_new.iloc[:,2:]= df_form2_employee_new.iloc[:,2:].astype(float).round(0)
                    df_form2_employee_new=df_form2_employee_new.fillna(0)
                    return df_form_employee_form_1,df_form2_employee_new,df_form3_employee_kehoach
            else:
                # GET DATA 
                # selected_ten_dv = dichvu_after_load[dichvu_after_load["ma_dv_id66"] == selected_data_kind]["ten_dv"].values[0]
                filtered_kehoach_service = module_view.filter_kehoach_by_service(kehoach_after_load,nhanvien_after_load,line_nv,selected_data_kind,selected_months,selected_year,selected_loaidoanhthu)
                filtered_thuchien_service_ht = module_view.filter_thuchien_by_service(thuchien_after_load,selected_data_kind,selected_months,selected_year,line_nv,selected_loaidoanhthu)
                filtered_thuchien_service_t1 = module_view.filter_thuchien_by_service(thuchien_after_load,selected_data_kind,1,selected_year,line_nv,selected_loaidoanhthu)
                filtered_thuchien_service_kytruoc = module_view.filter_thuchien_by_service(thuchien_after_load,selected_data_kind,selected_months -1,selected_year,line_nv,selected_loaidoanhthu)
                form_2_filter_service = module_view.filter_thuchien_by_service(thuchien_after_load,selected_data_kind,None,selected_year,line_nv,selected_loaidoanhthu)
                form3_kehoach_after_filter_no_month_service = module_view.filter_kehoach_by_service(kehoach_after_load,nhanvien_after_load,line_nv,selected_data_kind,None,selected_year,selected_loaidoanhthu)
                # FORM 1
                df_form_service_form_1,kehoach_col_map, thuchien_col_map = module_view.create_thuchien_service_dataframe_form_1(selected_months,nhanvien_after_load)
                df_form_service_form_1 = module_view.map_kehoach_to_thuchien_service(filtered_kehoach_service,df_form_service_form_1,kehoach_col_map,selected_months)
                df_form_service_form_1 = module_view.map_thuchien_service_for_form(df_form_service_form_1[thuchien_col_map],filtered_thuchien_service_ht,df_form_service_form_1)
                df_form_service_form_1 = module_view.map_thuchien_service_for_form(df_form_service_form_1["THỰC HIỆN T01"],filtered_thuchien_service_t1,df_form_service_form_1)
                df_form_service_form_1 = module_view.map_thuchien_service_for_form(df_form_service_form_1["KỲ TRƯỚC"],filtered_thuchien_service_kytruoc,df_form_service_form_1)
                df_form_service_form_1 = module_view.update_summary_rows_service(df_form_service_form_1)
                df_form_service_form_1 = module_view.update_form_1_table(df_form_service_form_1,thuchien_col_map,kehoach_col_map)
                # FORM 2
                form2_data_service = module_view.create_data_service_form_2(nhanvien_after_load)
                form2_data_service = module_view.map_thuchien_service_to_form_2(form_2_filter_service,form2_data_service)
                form2_data_service = module_view.update_summary_rows_service(form2_data_service)
                # FORM 3
                form3_data_service = module_view.create_data_service_form_2(nhanvien_after_load)
                form3_data_service = module_view.map_kehoach_to_form_3(form3_kehoach_after_filter_no_month_service,form3_data_service)
                form3_data_service = module_view.update_summary_rows_service(form3_data_service)
                
                if filtered_thuchien_service_ht.empty:
                    st.warning("Không có dữ liệu thực hiện")
                    return None,None,None
                else:
                    form3_data_service.iloc[:,2:]= form3_data_service.iloc[:,2:].astype(float).round(0) 
                    df_form_service_form_1.iloc[:, 2:] = (df_form_service_form_1.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').fillna(0).round(0))
                    df_form_service_form_1['% THỰC HIỆN'] = df_form_service_form_1['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
                    df_form_service_form_1['% VỚI KỲ TRƯỚC'] = df_form_service_form_1['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
                    df_form_service_form_1['% THỰC HIỆN T01'] = df_form_service_form_1['% THỰC HIỆN T01'].apply(lambda x: '{:.0f}%'.format(x))
                    form2_data_service.iloc[:,2:]= form2_data_service.iloc[:,2:].astype(float).round(0)
                    form2_data_service=form2_data_service.fillna(0)
                    return df_form_service_form_1,form2_data_service,form3_data_service                
        if radio_type_view_table == "Đơn":
                    df_1,df_2,df_3 = action_view_for_loop_func(selected_data_kind)
                    if action_check_view:
                        with st.spinner("💱 Đang tải dữ liệu..."):
                            # Áp dụng định dạng vào bảng
                            if df_1 is not None:
                                styled_df = df_1.style \
                                    .map(self.highlight_nonzero) \
                                    .map(self.highlight_numbers, subset=["+/- VỚI KỲ TRƯỚC","+ /- VỚI T01"]) \
                                    .format( self.add_arrow,subset=["+/- VỚI KỲ TRƯỚC","+ /- VỚI T01"]) \
                                    .map(self.highlight_numbers_percentage, subset=["% THỰC HIỆN","% THỰC HIỆN T01","% VỚI KỲ TRƯỚC"]) \
                                    .format(self.add_arrow_percentage, subset=["% THỰC HIỆN","% THỰC HIỆN T01","% VỚI KỲ TRƯỚC"]) \
                                    .format("{:.2f}", subset=[f"KẾ HOẠCH T0{selected_months}",f"THỰC HIỆN T0{selected_months}","THỰC HIỆN T01","KỲ TRƯỚC"] if selected_months != 1 else [f"KẾ HOẠCH T0{selected_months}",f"THỰC HIỆN T{selected_months}","THỰC HIỆN T01","KỲ TRƯỚC"])
                                df_2 = df_2.style.map(self.highlight_nonzero) \
                                    .format("{:.2f}",subset= df_2.columns[2:])
                                df_3 = df_3.style.map(self.highlight_nonzero) \
                                    .format("{:.2f}",subset= df_3.columns[2:])
                                container_df_view_type_table = st.container(key="container_df_view_type_table")
                                with container_df_view_type_table:
                                    st.markdown("""<h6 style="text-align:center;font-weight:bold;">Kểt quả phân tích</h6>""", unsafe_allow_html=True)
                                    st.dataframe(styled_df)
                                    st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số thực hiện các tháng</h6>""", unsafe_allow_html=True)
                                    st.dataframe(df_2)
                                    st.markdown("""<h6 style="text-align:center;font-weight:bold;">Số kế hoạch</h6>""", unsafe_allow_html=True)
                                    st.dataframe(df_3)
                            else:
                                st.warning("Vui lòng kiểm tra lại dữ liệu")
                    if action_check_export:
                        with st.spinner("💱 Đang tạo file..."):
                            if df_1 is not None:
                                if selected_option == "👩‍💻 Nhân viên":
                                    self.download_excel_single(df_1,df_2,df_3, f"Ketqua_nhanvien_thang_{selected_months}_{selected_data_kind}_{selected_year}.xlsx")
                                else:
                                    self.download_excel_single(df_1,df_2,df_3, f"Ketqua_dichvu_thang_{selected_months}_{selected_data_kind}_{selected_year}.xlsx")
                            else:
                                st.warning("Vui lòng kiểm tra lại dữ liệu")    
        else:
            if action_check_export or action_check_view:
                with st.spinner("Đang tạo file..."):
                    line_nv_for_file_down = line_after_load[line_after_load["ma_line"] == line_nv]["ten_line"].values[0]
                    if selected_option == "👩‍💻 Nhân viên":
                        ten_array = nhanvien_after_load["ten_nv"].unique()
                        ma_array = nhanvien_after_load["ma_nv"].unique()
                        cell_name = "Tên nhân viên"
                    else:
                        filtered_dichvu_down_all = dichvu_after_load[dichvu_after_load["danh_muc_tt"].notna()]
                        filtered_dichvu_down_all = filtered_dichvu_down_all[~filtered_dichvu_down_all["danh_muc_tt"].isin(["1", "1.1.01", "1.1.02", "1.1.03", "1.1.05", "1.1.06", "1.1.07"])]
                        ten_array = filtered_dichvu_down_all["ten_dv"].unique()
                        ma_array = filtered_dichvu_down_all["ma_dv_id66"].unique()
                        cell_name = "Tên dịch vụ"
                    check_status_empty = False
                    output_all_tt = io.BytesIO()
                    wb = Workbook()
                    for name, sheet_name in zip(ten_array, ma_array):
                        ws = wb.create_sheet(title=sheet_name)
                        ws['A1'] = "Phòng"
                        ws['B1'] = "DN3"
                        ws['A2'] = cell_name
                        ws['B2'] = name 
                        if selected_option == "👩‍💻 Nhân viên":
                        #    PART GET DATA
                            filtered_kehoach_employee = module_view.filter_kehoach_by_employee(kehoach_after_load,dichvu_after_load,sheet_name,line_nv,selected_months,selected_year,selected_loaidoanhthu)
                            filtered_thuchien_employee_ht = module_view.filter_data_thuchien_by_employee(thuchien_after_load,sheet_name,selected_months,selected_year,line_nv,selected_loaidoanhthu)
                            filtered_thuchien_employee_t1 = module_view.filter_data_thuchien_by_employee(thuchien_after_load,sheet_name,1,selected_year,line_nv,selected_loaidoanhthu)
                            filtered_thuchien_employee_kytruoc = module_view.filter_data_thuchien_by_employee(thuchien_after_load,sheet_name,selected_months -1,selected_year,line_nv,selected_loaidoanhthu)
                            form_2_filter_employee = module_view.filter_data_thuchien_by_employee(thuchien_after_load,sheet_name,None,selected_year,line_nv,selected_loaidoanhthu)
                            form3_kehoach_after_filter_no_month = module_view.filter_kehoach_by_employee(kehoach_after_load,dichvu_after_load,sheet_name,line_nv,None,selected_year,selected_loaidoanhthu)
                            df_hierachy_1, df_hierachy_2 = module_view.build_hierarchy(dichvu_after_load)
                            ma_dv_show, ten_dv_show = module_view.sorted_service_from_db(df_hierachy_2)
                        # FORM 1
                            df_form_employee_form_1,kehoach_col_map, thuchien_col_map = module_view.create_dataframe_thuchien_tt(selected_months,ma_dv_show,ten_dv_show)
                            df_form_employee_form_1 = module_view.map_kehoach_to_thuchien_tt(filtered_kehoach_employee,df_form_employee_form_1,kehoach_col_map,selected_months)
                            df_form_employee_form_1 = module_view.map_thuchien_tt_for_form(df_form_employee_form_1[thuchien_col_map],filtered_thuchien_employee_ht,df_form_employee_form_1)
                            df_form_employee_form_1 = module_view.map_thuchien_tt_for_form(df_form_employee_form_1["THỰC HIỆN T01"],filtered_thuchien_employee_t1,df_form_employee_form_1)
                            df_form_employee_form_1 = module_view.map_thuchien_tt_for_form(df_form_employee_form_1["KỲ TRƯỚC"],filtered_thuchien_employee_kytruoc,df_form_employee_form_1)
                            df_form_employee_form_1 = module_view.update_summary_rows(df_form_employee_form_1)
                            df_form_employee_form_1 = module_view.update_form_1_table(df_form_employee_form_1,thuchien_col_map,kehoach_col_map)
                        # FORM 2
                            df_form2_employee_new = module_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
                            df_form2_employee_new = module_view.map_thuchien_to_form_2(form_2_filter_employee,df_form2_employee_new)
                            df_form2_employee_new = module_view.update_summary_rows(df_form2_employee_new)
                        # FORM 3
                            df_form3_employee_kehoach = module_view.create_dataframe_form_2(ma_dv_show,ten_dv_show)
                            df_form3_employee_kehoach = module_view.map_kehoach_to_form_3_employee(form3_kehoach_after_filter_no_month,df_form3_employee_kehoach)
                            df_form3_employee_kehoach = module_view.update_summary_rows(df_form3_employee_kehoach)
                        
                            if filtered_thuchien_employee_ht.empty:
                                st.warning(sheet_name + "Không có dữ liệu thực hiện")
                                continue
                            else:
                                df_form3_employee_kehoach.iloc[:,2:]= df_form3_employee_kehoach.iloc[:,2:].astype(float).round(0) 
                                df_form_employee_form_1.iloc[:, 2:] = (df_form_employee_form_1.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').fillna(0).round(0))
                                df_form_employee_form_1['% THỰC HIỆN'] = df_form_employee_form_1['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
                                df_form_employee_form_1['% VỚI KỲ TRƯỚC'] = df_form_employee_form_1['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
                                df_form_employee_form_1['% THỰC HIỆN T01'] = df_form_employee_form_1['% THỰC HIỆN T01'].apply(lambda x: '{:.0f}%'.format(x))
                                df_form2_employee_new.iloc[:,2:]= df_form2_employee_new.iloc[:,2:].astype(float).round(0)
                                df_form2_employee_new=df_form2_employee_new.fillna(0)
                                df_1,df_2,df_3 = df_form_employee_form_1,df_form2_employee_new,df_form3_employee_kehoach
                                
                                file_name = f"Ketqua_tonghop_nhanvien_{selected_months}_{selected_year}_{selected_loaidoanhthu}_{line_nv_for_file_down}.xlsx"
                        else:
                            # GET DATA 
                            # selected_ten_dv = dichvu_after_load[dichvu_after_load["ma_dv_id66"] == sheet_name]["ten_dv"].values[0]
                            filtered_kehoach_service = module_view.filter_kehoach_by_service(kehoach_after_load,nhanvien_after_load,line_nv,sheet_name,selected_months,selected_year,selected_loaidoanhthu)
                            filtered_thuchien_service_ht = module_view.filter_thuchien_by_service(thuchien_after_load,sheet_name,selected_months,selected_year,line_nv,selected_loaidoanhthu)
                            filtered_thuchien_service_t1 = module_view.filter_thuchien_by_service(thuchien_after_load,sheet_name,1,selected_year,line_nv,selected_loaidoanhthu)
                            filtered_thuchien_service_kytruoc = module_view.filter_thuchien_by_service(thuchien_after_load,sheet_name,selected_months -1,selected_year,line_nv,selected_loaidoanhthu)
                            form_2_filter_service = module_view.filter_thuchien_by_service(thuchien_after_load,sheet_name,None,selected_year,line_nv,selected_loaidoanhthu)
                            form3_kehoach_after_filter_no_month_service = module_view.filter_kehoach_by_service(kehoach_after_load,nhanvien_after_load,line_nv,sheet_name,None,selected_year,selected_loaidoanhthu)
                            # FORM 1
                            df_form_service_form_1,kehoach_col_map, thuchien_col_map = module_view.create_thuchien_service_dataframe_form_1(selected_months,nhanvien_after_load)
                            df_form_service_form_1 = module_view.map_kehoach_to_thuchien_service(filtered_kehoach_service,df_form_service_form_1,kehoach_col_map,selected_months)
                            df_form_service_form_1 = module_view.map_thuchien_service_for_form(df_form_service_form_1[thuchien_col_map],filtered_thuchien_service_ht,df_form_service_form_1)
                            df_form_service_form_1 = module_view.map_thuchien_service_for_form(df_form_service_form_1["THỰC HIỆN T01"],filtered_thuchien_service_t1,df_form_service_form_1)
                            df_form_service_form_1 = module_view.map_thuchien_service_for_form(df_form_service_form_1["KỲ TRƯỚC"],filtered_thuchien_service_kytruoc,df_form_service_form_1)
                            df_form_service_form_1 = module_view.update_summary_rows_service(df_form_service_form_1)
                            df_form_service_form_1 = module_view.update_form_1_table(df_form_service_form_1,thuchien_col_map,kehoach_col_map)
                            # FORM 2
                            form2_data_service = module_view.create_data_service_form_2(nhanvien_after_load)
                            form2_data_service = module_view.map_thuchien_service_to_form_2(form_2_filter_service,form2_data_service)
                            form2_data_service = module_view.update_summary_rows_service(form2_data_service)
                            # FORM 3
                            form3_data_service = module_view.create_data_service_form_2(nhanvien_after_load)
                            form3_data_service = module_view.map_kehoach_to_form_3(form3_kehoach_after_filter_no_month_service,form3_data_service)
                            form3_data_service = module_view.update_summary_rows_service(form3_data_service)
                            
                            # if filtered_thuchien_service_ht.empty:
                            #     break
                            # else:
                            form3_data_service.iloc[:,2:]= form3_data_service.iloc[:,2:].astype(float).round(0) 
                            df_form_service_form_1.iloc[:, 2:] = (df_form_service_form_1.iloc[:, 2:].apply(pd.to_numeric, errors='coerce').fillna(0).round(0))
                            df_form_service_form_1['% THỰC HIỆN'] = df_form_service_form_1['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
                            df_form_service_form_1['% VỚI KỲ TRƯỚC'] = df_form_service_form_1['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
                            df_form_service_form_1['% THỰC HIỆN T01'] = df_form_service_form_1['% THỰC HIỆN T01'].apply(lambda x: '{:.0f}%'.format(x))
                            form2_data_service.iloc[:,2:]= form2_data_service.iloc[:,2:].astype(float).round(0)
                            form2_data_service=form2_data_service.fillna(0)
                            df_1,df_2,df_3 = df_form_service_form_1,form2_data_service,form3_data_service
                            file_name = f"Ketqua_tonghop_dichvu_{selected_months}_{selected_year}_{selected_loaidoanhthu}_{line_nv_for_file_down}.xlsx"

                        data = [df_3.columns.tolist()] + df_3.values.tolist()
                        for r_idx, row in enumerate(data, 1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx + 3, column=c_idx, value=value)
                                cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

                        data = [df_2.columns.tolist()] + df_2.values.tolist()
                        for r_idx, row in enumerate(data, 1):
                            for c_idx, value in enumerate(row, 1):
                                cell = ws.cell(row=r_idx + 3, column=c_idx + 15, value=value)
                                cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

                        data = [df_1.columns.tolist()] + df_1.values.tolist()
                        thin_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                        for r_idx, row in enumerate(data, start=1):
                            for c_idx, value in enumerate(row, start=1):
                                cell = ws.cell(row=r_idx + 3 + len(df_3) + 2, column=c_idx, value=value)
                                cell.font = Font(name='Times New Roman', size=12)
                                cell.border = thin_border
                                try:
                                    if c_idx in [8,11]: 
                                        val = float(value)
                                        if val != 0:
                                            cell.font = Font(color="00FF00" if val > 0 else "FF0000")
                                            cell.value = f"{'↑' if val > 0 else '↓'} {abs(val)}"
                                        else:
                                            cell.value = val

                                    elif c_idx in [5,7,10]:  
                                        val = float(value.strip('%'))
                                        if val != 0 and val != 100:
                                            cell.font = Font(color="00FF00" if (val - 100) > 0 else "FF0000")
                                            cell.value = f"{'↑' if (val - 100) > 0 else '↓'} {abs(val):.1f}%"
                                        else:
                                            cell.value = value
                                except (ValueError, AttributeError):
                                    pass
                        for col_range in [(3, 15), (16, 30)]:
                            start_col, end_col = col_range
                            start_cell = ws.cell(row=3, column=start_col)
                            end_cell = ws.cell(row=3, column=end_col)
                            ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
                            
                            # Thiết lập lại border cho các ô trong vùng merge
                            for row in ws.iter_rows(min_row=3, max_row=3, min_col=start_col, max_col=end_col):
                                for cell in row:
                                    cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))

                        # Gán giá trị cho ô merged và thiết lập font, đường viền
                        for col_idx, text in [(3, 'KẾ HOẠCH'), (16, 'THỰC HIỆN')]:
                            cell = ws.cell(row=3, column=col_idx, value=text)
                            cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                            cell.font = Font(name='Times New Roman', size=12, bold=True)  # Thiết lập font chữ và cỡ chữ
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                            # In đậm các dòng cần thiết
                        bold_rows_need_thuchien = [4, 4 + len(df_3) + 2]
                        for row_idx in bold_rows_need_thuchien:
                            for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_idx, max_row=row_idx):
                                for cell in col:
                                    cell.font = Font(bold=True)
                        # Tự động điều chỉnh độ rộng của các cột dựa trên nội dung của ô
                        for column_cells in ws.columns:
                            length = max(len(str(cell.value)) for cell in column_cells)
                            ws.column_dimensions[column_cells[0].column_letter].width = length + 2
                    if check_status_empty == False:
                        wb.save(output_all_tt)
                        output_all_tt.seek(0)
                        excel_base64 = base64.b64encode(output_all_tt.read()).decode()

                        # Create a download link using st.download_button
                        download_link_all_tt = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{file_name}.xlsx">Tải xuống kết quả</a>'
                        st.success("Tạo file thành công.")
                        st.markdown(download_link_all_tt, unsafe_allow_html=True)
                        
                    if check_status_empty:
                        st.warning("Không có dữ liệu để xuất file")
                        
             
    def download_excel_single(self,df_form1,df_form2, df_form3, file_name):
        check_status_empty = False
        output_all_tt = io.BytesIO()
        wb = Workbook()
        ws = wb.create_sheet(title="sheet 0")
        ws['A1'] = "Phòng"
        ws['B1'] = "DN3"
        ws['A2'] = "Line"
        ws['B2'] = line_after_load[line_after_load["ma_line"] == st.session_state.line_access]["ten_line"].values[0]
        if df_form1.empty:
            self.check_status_empty = True
        else:
            data = [df_form3.columns.tolist()] + df_form3.values.tolist()
            for r_idx, row in enumerate(data, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx + 3, column=c_idx, value=value)
                    cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                    cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

            data = [df_form2.columns.tolist()] + df_form2.values.tolist()
            for r_idx, row in enumerate(data, 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx + 3, column=c_idx + 15, value=value)
                    cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                    cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

            data = [df_form1.columns.tolist()] + df_form1.values.tolist()
            thin_border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
            for r_idx, row in enumerate(data, start=1):
                for c_idx, value in enumerate(row, start=1):
                    cell = ws.cell(row=r_idx + 3 + len(df_form3) + 2, column=c_idx, value=value)
                    cell.font = Font(name='Times New Roman', size=12)
                    cell.border = thin_border
                    try:
                        if c_idx in [8,11]: 
                            val = float(value)
                            if val != 0:
                                cell.font = Font(color="00FF00" if val > 0 else "FF0000")
                                cell.value = f"{'↑' if val > 0 else '↓'} {abs(val)}"
                            else:
                                cell.value = val

                        elif c_idx in [5,7,10]:  
                            val = float(value.strip('%'))
                            if val != 0 and val != 100:
                                cell.font = Font(color="00FF00" if (val - 100) > 0 else "FF0000")
                                cell.value = f"{'↑' if (val - 100) > 0 else '↓'} {abs(val):.1f}%"
                            else:
                                cell.value = value
                    except (ValueError, AttributeError):
                        pass
            for col_range in [(3, 15), (16, 30)]:
                start_col, end_col = col_range
                start_cell = ws.cell(row=3, column=start_col)
                end_cell = ws.cell(row=3, column=end_col)
                ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
                
                # Thiết lập lại border cho các ô trong vùng merge
                for row in ws.iter_rows(min_row=3, max_row=3, min_col=start_col, max_col=end_col):
                    for cell in row:
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))

            # Gán giá trị cho ô merged và thiết lập font, đường viền
            for col_idx, text in [(3, 'KẾ HOẠCH'), (16, 'THỰC HIỆN')]:
                cell = ws.cell(row=3, column=col_idx, value=text)
                cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                cell.font = Font(name='Times New Roman', size=12, bold=True)  # Thiết lập font chữ và cỡ chữ
                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                # In đậm các dòng cần thiết
            bold_rows_need_thuchien = [4,4+ len(df_form3) + 2] 
            for row_idx in bold_rows_need_thuchien:
                for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_idx, max_row=row_idx):
                    for cell in col:
                        cell.font = Font(bold=True)
            # Tự động điều chỉnh độ rộng của các cột dựa trên nội dung của ô
            for column_cells in ws.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                ws.column_dimensions[column_cells[0].column_letter].width = length + 2
        if check_status_empty == False:
            wb.save(output_all_tt)
            output_all_tt.seek(0)
            excel_base64 = base64.b64encode(output_all_tt.read()).decode()

            # Create a download link using st.download_button
            download_link_all_tt = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{file_name}.xlsx">Tải xuống kết quả</a>'
            st.success("Tạo file thành công.")
            st.markdown(download_link_all_tt, unsafe_allow_html=True)
            
        if check_status_empty:
            st.warning("Không có dữ liệu để xuất file")
        
# PART MAIN APP     
class MAIN_APP():
    def __init__(self):
        pass
    def run_view(self):
        with st.sidebar:
            module_config.show_expander_sidebar()
        ctn_head_select_option = st.container(key="container_head_select_option")
        with ctn_head_select_option:
            cols_head_select_option_revenue = ctn_head_select_option.columns([1,1])
            selected_menu_sidebar = cols_head_select_option_revenue[1].pills(" ",["Dạng Dashboard", "Dạng bảng"],default="Dạng Dashboard", key="menu_sidebar_view_button")
        return selected_menu_sidebar,cols_head_select_option_revenue
    def main(self,selected,cols_head_select_option_revenue):
        class_design_frontend = DESIGN_FRONTEND()
        if selected == "Dạng Dashboard":
            selected_option,selected_months,selected_year,selected_loaidoanhthu,line_nv,selected_data_kind = class_design_frontend.sidebar_option(cols_head_select_option_revenue)
            if selected_option == "👩‍💻 Nhân viên":
                class_design_frontend.employee_design_frontend(selected_months,line_nv, selected_year, selected_loaidoanhthu, selected_data_kind)
            else:
                class_design_frontend.service_design_frontend(selected_months,line_nv, selected_year, selected_loaidoanhthu, selected_data_kind)
        elif selected == "Dạng bảng":
            radio_type_view_table,export_excel_table_file,submit_view_option = class_design_frontend.ui_info_no_search()
            selected_option,selected_months,selected_year,selected_loaidoanhthu,line_nv,selected_data_kind = class_design_frontend.table_sidebar_option(radio_type_view_table,cols_head_select_option_revenue)
            class_design_frontend.table_employee_design(selected_option,selected_months,line_nv,selected_year,selected_loaidoanhthu,selected_data_kind,radio_type_view_table)
            if submit_view_option:
                BACKEND_TABLE_VIEW().table_generate_backend(selected_months,line_nv,selected_year,selected_loaidoanhthu,selected_data_kind,selected_option,radio_type_view_table,submit_view_option,False)
            if export_excel_table_file:
                BACKEND_TABLE_VIEW().table_generate_backend(selected_months,line_nv,selected_year,selected_loaidoanhthu,selected_data_kind,selected_option,radio_type_view_table,False,export_excel_table_file)


thuchien_after_load, kehoach_after_load, nhanvien_after_load, dichvu_after_load,line_after_load = module_view.load_data()     
design_frontend_class = DESIGN_FRONTEND()
main_app_class = MAIN_APP()
selected_menu_sidebar,cols_head_select_option_revenue = main_app_class.run_view()
main_app_class.main(selected_menu_sidebar,cols_head_select_option_revenue)
container_menu = st.sidebar.container(key="container_menu_view")
module_config.add_sidebar_footer()




