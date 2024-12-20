import streamlit as st
import pandas as pd
import mysql.connector
from mysql.connector import pooling, Error
from mysql.connector import OperationalError, InternalError
import base64
import datetime
from streamlit_option_menu import option_menu
from datetime import date
import os
import numpy as np
import io
from openpyxl.styles import Font, Border, Side
from openpyxl import Workbook
import openpyxl
import time
import xlsxwriter
import sys
from openpyxl.styles import PatternFill
import streamlit.components.v1 as components
import MY_LIBRARIES_VNPT as mvnpt
import bcrypt
import requests


st.set_page_config(layout='wide',page_title="THỐNG KÊ DOANH THU - KHDN3", page_icon='vnpt.ico', initial_sidebar_state='expanded')
def connect_to_mysql():
    try:
        # Thực hiện kết nối đến cơ sở dữ liệu MySQL
        conn = mysql.connector.connect(
            host='ABC',
            user='ABC',
            password='ABCD',
            database='samryvnc'
        )

        # Tạo đối tượng cursor từ kết nối
        cursor = conn.cursor()

        print("Kết nối đến cơ sở dữ liệu thành công.")
        
        # Trả về cả kết nối và cursor để có thể sử dụng chúng trong mã khác
        return conn, cursor
    except Exception as e:
        print("Lỗi khi kết nối đến cơ sở dữ liệu:", e)
        return None, None
def check_connection(conn):
    try:
        if conn.unread_result:
            conn.cursor().fetchall() 

        cursor = conn.cursor()
        cursor.execute("SELECT 1")
        cursor.fetchall() 
        cursor.close()
        return True  
    except IndexError as e:
        print(f"Lỗi IndexError khi truy vấn: {e}")
        return False
    except (OperationalError, InternalError) as e:
        print(f"Lỗi kết nối hoặc lỗi đọc kết quả: {e}")
        return False
    except Exception as e:
        print(f"Lỗi không xác định: {e}")
        return False
def reconnect_if_needed():
    if 'conn' not in st.session_state or st.session_state.conn is None:
        st.session_state.conn, st.session_state.cursor = connect_to_mysql()

    if not check_connection(st.session_state.conn):
        st.session_state.conn, st.session_state.cursor = connect_to_mysql()
def login():
    st.markdown("<h1 style='text-align: center;'>THỐNG KÊ DOANH THU</h1>", unsafe_allow_html=True)
    if 'is_logged_in' not in st.session_state:
        st.session_state.is_logged_in = False
        st.session_state.role_access_admin = False
    title_placeholder = st.empty()
    username_placeholder = st.empty()
    password_placeholder = st.empty()
    success_placeholder = st.empty()
    button_placeholder = st.empty()
    if not st.session_state.is_logged_in:
        title_placeholder.title("Đăng nhập")
        username = username_placeholder.text_input("Tên người dùng", placeholder="Enter user name", key="username_login")
        password = password_placeholder.text_input("Mật khẩu", type="password", placeholder="Enter password", key="password_login")
        if button_placeholder.button("Đăng nhập", use_container_width=True, help="Nhấn vào để đăng nhập!", key="login"):
            # os.environ.get('USERNAME_LOGIN')
            # os.environ.get('PASSWORD_LOGIN')
            conn, cursor = connect_to_mysql()
            st.session_state.__setitem__('conn', conn)
            st.session_state.__setitem__('cursor', cursor)
            user_role_mvnpt,image_url_profile = mvnpt.check_user_access(username,password,conn,cursor)
            if user_role_mvnpt and (user_role_mvnpt is not None):
                success_placeholder.success("Đăng nhập thành công!")
                time.sleep(2)  # Dừng 2 giây
                title_placeholder.empty()
                username_placeholder.empty()
                password_placeholder.empty()
                success_placeholder.empty()
                button_placeholder.empty()
                st.session_state.is_logged_in = True
                if user_role_mvnpt == 'admin':
                    st.session_state.role_access_admin = True
                    st.session_state.usernamevnpt = username
                    if image_url_profile is not None:
                        st.session_state.image_url_profile_123 = image_url_profile
                        image_url_profile_reponse = requests.get(image_url_profile)
                        if image_url_profile_reponse.status_code == 200:
                            st.session_state.image_profile = image_url_profile_reponse
                else:
                    st.session_state.role_access_admin = False
                    st.session_state.usernamevnpt = username
                    if image_url_profile is not None:
                        st.session_state.image_url_profile_123 = image_url_profile
                        image_url_profile_reponse = requests.get(image_url_profile)
                        if image_url_profile_reponse.status_code == 200:
                            st.session_state.image_profile = image_url_profile_reponse
            else:
                st.error("Tên người dùng hoặc mật khẩu không đúng!")

class DatabaseConnection:
    _instance = None

    def __new__(cls):
        if cls._instance is None:
            cls._instance = super(DatabaseConnection, cls).__new__(cls)
            # Thực hiện kết nối tới cơ sở dữ liệu ở đây
            cls._instance.conn, cls._instance.cursor = connect_to_mysql()
        return cls._instance


class CONTROL_DATABASE:
    def __init__(self):
        self.data = None
    # FUNCION QUERY DATA FROM DATABASE
    def query_kehoach_by_line_year(self,line_test, year_test,conn, cursor):
        
        # Thực hiện truy vấn SQL
        cursor.execute("SELECT * FROM kehoach WHERE Line=%s AND year=%s", (line_test, year_test))
        rows = cursor.fetchall()

        # Tạo DataFrame từ các hàng
        kq_kehoach = pd.DataFrame(rows, columns=[col[0] for col in cursor.description])

        return kq_kehoach
    def query_kehoach_by_line_year_luyke(self,line_test, year_test, conn, cursor):
        # Thực hiện truy vấn SQL
        cursor.execute("SELECT * FROM kehoach WHERE Line=%s AND year=%s", (line_test, year_test))
        rows = cursor.fetchall()

        # Tạo DataFrame từ các hàng
        kq_kehoach = pd.DataFrame(rows, columns=[col[0] for col in cursor.description])

        # Chọn các cột tháng
        thang_columns = kq_kehoach.columns[kq_kehoach.columns.str.startswith('Tháng')]

        # Tính tổng tích lũy cho các cột tháng
        kq_kehoach[thang_columns] = kq_kehoach[thang_columns].cumsum(axis=1)

        return kq_kehoach
    def get_data_thuchien(self,thang, nam, line,conn, cursor):
        
        query = "SELECT nhom_dv, doanh_thu FROM thuchien WHERE thang=%s AND nam=%s AND line=%s;"
        cursor.execute(query, (thang, nam, line))
        
        # Lấy tất cả kết quả từ truy vấn
        rows_thuchien = cursor.fetchall()
        
        # Tạo DataFrame từ kết quả
        kq_thuchien = pd.DataFrame(rows_thuchien, columns=["nhom_dv", "doanh_thu"])
        kq_thuchien['nhom_dv'] = kq_thuchien['nhom_dv'].str.strip()
        
        # Trả về DataFrame
        return kq_thuchien
    def get_ketqua_kehoach_tt(self, thang, year_value, line_value,conn, cursor):
        
        # Tạo tên cột động dựa trên giá trị của biến thang
        column_name = f'Tháng_{thang:02d}'

        # Tạo câu lệnh truy vấn để lấy cột động từ bảng revenue
        query = f"""SELECT Dịch_vụ, {column_name} FROM kehoach WHERE Line = %s AND year = %s;"""
        # Thực thi câu truy vấn và lấy kết quả
        cursor.execute(query, (line_value, year_value))
        rows = cursor.fetchall()

        # Tạo DataFrame từ các hàng
        kq_kehoach = pd.DataFrame(rows, columns=['Dịch_vụ', column_name])

        return kq_kehoach
    def get_data_thuchien_all(self,nam, line,conn, cursor):
        
        # Thực hiện truy vấn SQL để lấy dữ liệu từ bảng thuchien
        query = "SELECT nhom_dv, doanh_thu, thang, nam FROM thuchien WHERE nam=%s AND line=%s"
        cursor.execute(query, (nam, line))

        # Lấy tất cả kết quả từ truy vấn
        rows_thuchien = cursor.fetchall()

        # Tạo DataFrame từ kết quả
        kq_thuchien = pd.DataFrame(rows_thuchien, columns=["nhom_dv", "doanh_thu", "thang", "nam"])
        kq_thuchien['nhom_dv'] = kq_thuchien['nhom_dv'].str.strip()

        # Trả về DataFrame
        return kq_thuchien
    # FUNCTION INSERT DATA TO DATABASE
    def select_rows_kehoach(self,data):
        # Tạo danh sách các chỉ số hàng bạn muốn lấy
        rows_to_select = list(range(5, 14)) + list(range(15, 17)) + list(range(19, len(data)))

        # Sử dụng hàm iloc để lấy các hàng ngắt quãng
        subset = data.iloc[rows_to_select, 1:]

        return subset
    def insert_data_kehoach(self,line, year, data,conn, cursor):
        
    # Lặp qua từng hàng trong DataFrame
        for index, row in data.iterrows():
            # Câu lệnh SQL để insert dữ liệu
            sql = """INSERT INTO kehoach
                    (Dịch_vụ, Tháng_01, Tháng_02, Tháng_03, Tháng_04, Tháng_05, Tháng_06, Tháng_07, Tháng_08, Tháng_09, Tháng_10, Tháng_11, Tháng_12, LK_năm, Line, year)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            # Dữ liệu để insert, kết hợp với line và year
            data_to_insert = tuple(list(row[:14]) + [line, year])
            
            # Thực thi câu lệnh SQL
            cursor.execute(sql, data_to_insert)
        
        # Commit các thay đổi vào cơ sở dữ liệu
        conn.commit()
        return True
    def group_data_for_insert_thuchien(self,data):
        # Nhóm dữ liệu theo cột 'TEN_TO_CS' và 'NHOM DV', sau đó tính tổng doanh thu
        grouped_data = data.groupby(['TEN_TO_CS', 'NHOM DV']).agg({'DOANH THU': 'sum'}).reset_index()
        return grouped_data
    def insert_data_to_thuchien(self,thang, nam, group_data,conn, cursor):
        
        # Lặp qua từng hàng trong DataFrame
        for index, row in group_data.iterrows():
            # Câu lệnh SQL để insert dữ liệu
            sql = """INSERT INTO thuchien
                    (line, nhom_dv, doanh_thu, thang, nam)
                    VALUES (%s, %s, %s, %s, %s)"""
            # Dữ liệu để insert, kết hợp với thang và nam
            data_to_insert = (row[0], row[1], row[2], thang, nam)
            
            # Thực thi câu lệnh SQL
            cursor.execute(sql, data_to_insert)
        
        # Commit các thay đổi vào cơ sở dữ liệu
        conn.commit()
        return True
    
    # FUNCTION DELETE DATA FROM DATABASE
    def delete_from_kehoach(self,line, year,conn, cursor):
        # Kết nối đến cơ sở dữ liệu và tạo đối tượng cursor
        
        
        try:
            # Tạo câu lệnh SQL DELETE để xóa dữ liệu từ bảng `kehoach`
            query = "DELETE FROM kehoach WHERE Line = %s AND year = %s;"
            
            # Thực thi câu lệnh SQL với các tham số được chỉ định
            cursor.execute(query, (line, year))
            
            # Commit các thay đổi vào cơ sở dữ liệu
            conn.commit()
            
            st.success(f"Deleted from kehoach table: Line={line}, Year={year}")
            
        except mysql.connector.Error as e:
            st.error(f"Error deleting data from kehoach table: {e}")
    def delete_from_thuchien(self, nam, thang,conn, cursor):
        # Kết nối đến cơ sở dữ liệu và tạo đối tượng cursor
        
        
        try:
            # Tạo câu lệnh SQL DELETE để xóa dữ liệu từ bảng `thuchien`
            query = "DELETE FROM thuchien WHERE nam = %s AND thang = %s;"
            
            # Thực thi câu lệnh SQL với các tham số được chỉ định
            cursor.execute(query, (nam, thang))
            
            # Commit các thay đổi vào cơ sở dữ liệu
            conn.commit()
            
            st.success(f"Deleted from thuchien table: nam ={nam}, thang={thang}")
            
        except mysql.connector.Error as e:
            st.error(f"Error deleting data from thuchien table: {e}")
    def delete_from_thuchien_baocao(self, nam, thang,loaibaocao,conn, cursor):
        # Kết nối đến cơ sở dữ liệu và tạo đối tượng cursor
        
        
        try:
            # Tạo câu lệnh SQL DELETE để xóa dữ liệu từ bảng `thuchien`
            query = "DELETE FROM thuchien_baocao WHERE nam = %s AND thang = %s AND loaibaocao = %s;"
            
            # Thực thi câu lệnh SQL với các tham số được chỉ định
            cursor.execute(query, (nam, thang, loaibaocao))
            
            # Commit các thay đổi vào cơ sở dữ liệu
            conn.commit()
            
            st.success(f"Đã xóa dữ liệu thực hiện: năm ={nam}, tháng={thang}, loại báo cáo={loaibaocao}")
            
        except mysql.connector.Error as e:
            st.error(f"Error deleting data from thuchien table: {e}")

    def query_distinct_kehoach(self,conn, cursor):
        # Kết nối đến cơ sở dữ liệu và tạo đối tượng cursor
        
        
        try:
            # Tạo câu lệnh SQL để truy vấn các giá trị duy nhất của Line và Year từ bảng `kehoach`
            query = "SELECT DISTINCT Line, year FROM kehoach WHERE Line LIKE '%Line%';"
            
            # Thực thi câu lệnh SQL
            cursor.execute(query)
            
            # Lấy tất cả các hàng kết quả
            rows = cursor.fetchall()
            
            # Tạo DataFrame từ kết quả
            df = pd.DataFrame(rows, columns=['Line', 'Năm'])
            
            return df
            
        except mysql.connector.Error as e:
            print(f"Error querying data from kehoach table: {e}")
            return pd.DataFrame()
    def query_distinct_kehoach_tonghop(self,conn, cursor):
        # Kết nối đến cơ sở dữ liệu và tạo đối tượng cursor
        
        
        try:
            # Tạo câu lệnh SQL để truy vấn các giá trị duy nhất của Line và Year từ bảng `kehoach`
            query = "SELECT DISTINCT Line, year FROM kehoach WHERE Line NOT IN (SELECT Line FROM kehoach WHERE Line LIKE '%Line%');"
            
            # Thực thi câu lệnh SQL
            cursor.execute(query)
            
            # Lấy tất cả các hàng kết quả
            rows = cursor.fetchall()
            
            # Tạo DataFrame từ kết quả
            df = pd.DataFrame(rows, columns=['Loại báo cáo', 'Năm'])
            
            return df
            
        except mysql.connector.Error as e:
            print(f"Error querying data from kehoach table: {e}")
            return pd.DataFrame()
            
    def query_distinct_thuchien(self,conn, cursor):
        # Kết nối đến cơ sở dữ liệu và tạo đối tượng cursor
        
        
        try:
            # Tạo câu lệnh SQL để truy vấn các giá trị duy nhất của Line và Year từ bảng `kehoach`
            query = "SELECT DISTINCT line, nam,thang FROM thuchien;"
            
            # Thực thi câu lệnh SQL
            cursor.execute(query)
            
            # Lấy tất cả các hàng kết quả
            rows = cursor.fetchall()
            
            # Tạo DataFrame từ kết quả
            df = pd.DataFrame(rows, columns=['Line', 'Năm', 'Tháng'])
            
            return df
            
        except mysql.connector.Error as e:
            print(f"Error querying data from kehoach table: {e}")
            return pd.DataFrame()
            
    def query_distinct_thuchien_tonghop(self,conn, cursor):
        # Kết nối đến cơ sở dữ liệu và tạo đối tượng cursor
        
        
        try:
            # Tạo câu lệnh SQL để truy vấn các giá trị duy nhất của Line và Year từ bảng `kehoach`
            query = "SELECT DISTINCT loaibaocao, nam,thang FROM thuchien_baocao;"
            
            # Thực thi câu lệnh SQL
            cursor.execute(query)
            
            # Lấy tất cả các hàng kết quả
            rows = cursor.fetchall()
            
            # Tạo DataFrame từ kết quả
            df = pd.DataFrame(rows, columns=['Loại báo cáo', 'Năm', 'Tháng'])
            
            return df
            
        except mysql.connector.Error as e:
            print(f"Error querying data from kehoach table: {e}")
            return pd.DataFrame()
    # FUNCTION INSERT DATA BAOCAOTONGHOP TO DATABASE
    def select_rows_kehoach_baocao(self,data):
        # Tạo danh sách các chỉ số hàng bạn muốn lấy
        rows_to_select = list(range(4, 13)) + list(range(14, 16)) + list(range(18, len(data)))

        # Sử dụng hàm iloc để lấy các hàng ngắt quãng
        subset = data.iloc[rows_to_select, 1:]

        return subset
    def insert_data_thuchien_2023(self,loai, year, data, conn, cursor):
        # Lặp qua từng hàng trong DataFrame
        for index, row in data.iterrows():
            sql = """INSERT INTO thuchien_2023
                    (Dịch_vụ, Tháng_01, Tháng_02, Tháng_03, Tháng_04, Tháng_05, Tháng_06, Tháng_07, Tháng_08, Tháng_09, Tháng_10, Tháng_11, Tháng_12, LK_năm, loaibaocao, year)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)"""
            data_to_insert = tuple(list(row[:14]) + [loai, year])
            
            cursor.execute(sql, data_to_insert)
        
        conn.commit()
    def group_data_insert_baocaothuchien(self,data):
        # Nhóm dữ liệu theo cột 'TEN_TO_CS' và 'NHOM DV', sau đó tính tổng doanh thu
        grouped_data = data.groupby(['NHOM DV']).agg({'DOANH THU': 'sum'}).reset_index()
        return grouped_data
    def insert_data_to_thuchien_baocao(self,loaibaocao,thang, nam, group_data, conn, cursor):
        for index, row in group_data.iterrows():
            sql = """INSERT INTO thuchien_baocao
                    (loaibaocao, nhom_dv, doanh_thu, thang, nam)
                    VALUES (%s, %s, %s, %s, %s)"""
            data_to_insert = (loaibaocao,row.iloc[0], row.iloc[1], thang, nam)
            cursor.execute(sql, data_to_insert)
        conn.commit()
        return True
    # FUNCTION QUERY DATA BAOCAOTONGHOP FROM DATABASE
    def get_data_thuchien_baocao(self,thang, nam, line,conn, cursor):

        query = "SELECT nhom_dv, doanh_thu FROM thuchien_baocao WHERE thang=%s AND nam=%s AND loaibaocao=%s;"
        cursor.execute(query, (thang, nam, line))

        rows_thuchien = cursor.fetchall()

        kq_thuchien = pd.DataFrame(rows_thuchien, columns=["nhom_dv", "doanh_thu"])
        kq_thuchien['nhom_dv'] = kq_thuchien['nhom_dv'].str.strip()

        return kq_thuchien
    def get_ketqua_kehoach_baocao(self,thang, year_value, loaibaocao,conn, cursor):
        column_name = f'Tháng_{thang:02d}'
        query = f"""SELECT Dịch_vụ, {column_name},LK_năm FROM kehoach WHERE Line = %s AND year = %s;"""
        cursor.execute(query, (loaibaocao, year_value))
        rows = cursor.fetchall()

        kq_kehoach = pd.DataFrame(rows, columns=['Dịch_vụ', column_name,'LK_năm'])

        return kq_kehoach
    def get_ketqua_thuchien_2023(self,thang, loaibaocao,conn, cursor):
    
        column_name = f'Tháng_{thang:02d}'

        query = f"""SELECT Dịch_vụ, {column_name} FROM thuchien_2023 WHERE loaibaocao = %s AND year = %s;"""
        cursor.execute(query, (loaibaocao, 2023))
        rows = cursor.fetchall()

        kq_kehoach = pd.DataFrame(rows, columns=['Dịch_vụ', column_name])

        return kq_kehoach
    def get_data_form2_baocao(self,nam, loaibaocao,conn, cursor):
    
        # Thực hiện truy vấn SQL để lấy dữ liệu từ bảng thuchien
        query = "SELECT nhom_dv, doanh_thu, thang, nam FROM thuchien_baocao WHERE nam=%s AND loaibaocao=%s"
        cursor.execute(query, (nam, loaibaocao))

        # Lấy tất cả kết quả từ truy vấn
        rows_thuchien = cursor.fetchall()

        # Tạo DataFrame từ kết quả
        kq_thuchien = pd.DataFrame(rows_thuchien, columns=["nhom_dv", "doanh_thu", "thang", "nam"])
        kq_thuchien['nhom_dv'] = kq_thuchien['nhom_dv'].str.strip()

        # Trả về DataFrame
        return kq_thuchien
    def get_ketqua_thuchien_2023_luyke(self,thang, loaibaocao,conn, cursor):
    # Tạo danh sách các tên cột cho các tháng trước tháng hiện tại
        previous_month_cols = [f'Tháng_{i:02d}' for i in range(1, thang)]
        
        # Tạo danh sách các cột trong truy vấn SQL
        select_columns = ', '.join(previous_month_cols + ['Dịch_vụ', f'Tháng_{thang:02d}'])

        query = f"""SELECT {select_columns} FROM thuchien_2023 WHERE loaibaocao = %s AND year = %s;"""
        cursor.execute(query, (loaibaocao, 2023))
        rows = cursor.fetchall()

        # Tạo DataFrame từ kết quả truy vấn
        kq_kehoach = pd.DataFrame(rows, columns=previous_month_cols + ['Dịch_vụ', f'Tháng_{thang:02d}'])
        
        # Tính tổng lũy kế cho các tháng trước đó
        for col in previous_month_cols:
            kq_kehoach[f'Tháng_{thang:02d}'] += kq_kehoach[col]
            
        kq_kehoach = kq_kehoach.iloc[:, -2:]

        return kq_kehoach    
    def get_ketqua_kehoach_baocao_luyke(self, thang, year_value, loaibaocao, conn, cursor):
        # Tạo danh sách các tên cột cho các tháng trước tháng hiện tại
        previous_month_cols = [f'Tháng_{i:02d}' for i in range(1, thang)]
        
        # Tạo danh sách các cột trong truy vấn SQL
        select_columns = ', '.join(previous_month_cols + ['Dịch_vụ', f'Tháng_{thang:02d}', 'LK_năm'])
        
        # Truy vấn SQL để lấy dữ liệu từ bảng kehoach
        query = f"""
            SELECT {select_columns}
            FROM kehoach 
            WHERE Line = %s AND year = %s;
        """
        cursor.execute(query, (loaibaocao, year_value))
        rows = cursor.fetchall()

        # Tạo DataFrame từ kết quả truy vấn
        kq_kehoach = pd.DataFrame(rows, columns=previous_month_cols + ['Dịch_vụ', f'Tháng_{thang:02d}', 'LK_năm'])
        
        # Tính tổng lũy kế cho các tháng trước đó
        for col in previous_month_cols:
            kq_kehoach[f'Tháng_{thang:02d}'] += kq_kehoach[col]
            
        kq_kehoach = kq_kehoach.iloc[:, -3:]
        
        return kq_kehoach
    def get_data_form2_baocao_luyke(self,nam, loaibaocao, conn, cursor):
        # Thực hiện truy vấn SQL để lấy dữ liệu từ bảng thuchien
        query = "SELECT nhom_dv, doanh_thu, thang, nam FROM thuchien_baocao WHERE nam=%s AND loaibaocao=%s ORDER BY nhom_dv, thang;"
        cursor.execute(query, (nam, loaibaocao))

        # Lấy tất cả kết quả từ truy vấn
        rows_thuchien = cursor.fetchall()

        # Tạo DataFrame từ kết quả
        kq_thuchien = pd.DataFrame(rows_thuchien, columns=["nhom_dv", "doanh_thu", "thang", "nam"])
        kq_thuchien['nhom_dv'] = kq_thuchien['nhom_dv'].str.strip()

        # Tính tổng lũy kế cho mỗi dịch vụ (nhóm_dv)
        cumulative_sum_per_service = {}
        for index, row in kq_thuchien.iterrows():
            nhom_dv = row['nhom_dv']
            if nhom_dv not in cumulative_sum_per_service:
                cumulative_sum_per_service[nhom_dv] = []
            if len(cumulative_sum_per_service[nhom_dv]) == 0:
                cumulative_sum_per_service[nhom_dv].append(row['doanh_thu'])
            else:
                cumulative_sum_per_service[nhom_dv].append(cumulative_sum_per_service[nhom_dv][-1] + row['doanh_thu'])

        # Thêm cột tổng lũy kế cho mỗi dịch vụ vào DataFrame
        kq_thuchien['doanh_thu'] = [cumulative_sum_per_service[row['nhom_dv']].pop(0) for index, row in kq_thuchien.iterrows()]

        # Trả về DataFrame
        return kq_thuchien
    def get_data_thuchien_baocao_luyke(self,thang, nam, line, conn, cursor):
        query = """
            SELECT nhom_dv, doanh_thu 
            FROM thuchien_baocao
            WHERE thang <= %s AND nam = %s AND loaibaocao = %s
            ORDER BY nhom_dv;
        """
        cursor.execute(query, (thang, nam, line))
        rows_thuchien = cursor.fetchall()

        kq_thuchien = pd.DataFrame(rows_thuchien, columns=["nhom_dv", "doanh_thu"])
        kq_thuchien['nhom_dv'] = kq_thuchien['nhom_dv'].str.strip()

            # Group by 'nhom_dv' and calculate cumulative sum within each group
        if not kq_thuchien.empty:
            kq_thuchien['doanh_thu'] = kq_thuchien.groupby('nhom_dv')['doanh_thu'].cumsum()

        # Drop duplicate rows while keeping the first occurrence in each group
        kq_thuchien_unique = kq_thuchien.drop_duplicates(subset='nhom_dv', keep='last')

        return kq_thuchien_unique
class DATAFRAME:
    def __init__(self):
        self.data = None
    def create_dataframe_public(self):
        columns = ["STT", "Dịch vụ"] + [f"Tháng {i:02d}" for i in range(1, 13)] + ["LK năm"]

        data = {
            "STT": ["A", "1", "1.1", "1.1.1", "1.1.2", "1.1.3", "1.1.4", "1.1.5", "1.1.6", "1.1.7", "1.1.8", "1.1.9", "1.2", "1.2.1", "1.2.2", "2", "2.1", "2.1.1", "2.1.2", "2.1.3", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7"],
            "Dịch vụ": ["DOANH THU VTCNTT", "Công nghệ thông tin", "Dịch vụ số doanh nghiệp", "Hạ tầng CNTT", "Chính quyền số", "Y tế số", "Công nghệ nền tảng", "An toàn thông tin", "Giáo dục số", "Quản trị doanh nghiệp", "Vận tải và Logictic", "Phân phối bán lẻ", "Dịch vụ số cá nhân", "MyTV", "Dịch vụ tiện ích", "Dịch vụ Viễn thông", "Dịch vụ di động (5917)", "Di động trả trước", "Di động trả sau", "Gói GD,VP,Home", "Băng rộng", "Cố định", "Internet trực tiếp", "Truyền số liệu", "Dịch vụ GTGT khác", "Cho thuê Hạ tầng"],
        }

        # Tạo DataFrame
        kq_dataframe = pd.DataFrame(data, columns=columns)

        return kq_dataframe
    def map_kehoach_to_dataframe(self,kq_kehoach, kq_dataframe):
        for index, row in kq_kehoach.iterrows():
            service = row["Dịch_vụ"]

            # Tìm vị trí của dịch vụ trong kq_dataframe
            matching_rows = kq_dataframe[kq_dataframe["Dịch vụ"] == service]

            # Nếu có ít nhất một hàng khớp, ánh xạ dữ liệu
            if not matching_rows.empty:
                matching_index = matching_rows.index[0]
                kq_dataframe.iloc[matching_index, 2:] = row[1:-2].values
            else:
                print(f"Dịch vụ '{service}' không được tìm thấy trong DataFrame 'kq_dataframe'.")

        return kq_dataframe
    def update_summary_rows(self,df):
        df['STT'] = df['STT'].astype(str)

        # Lấy danh sách các cột tháng
        columns_to_sum = df.columns[2:]

        # Hàm để lấy tất cả các hàng con của một hàng tổng hợp
        def get_child_rows(df, parent_stt):
            return df[df['STT'].str.startswith(parent_stt + '.')]

        # Cập nhật các hàng tổng hợp
        for index, row in df.iterrows():
            if pd.isna(row[columns_to_sum]).all() or (row[columns_to_sum] == 0).all():
                # Lấy tất cả các hàng con của hàng tổng hợp hiện tại
                child_rows = get_child_rows(df, row['STT'])

                # Tính tổng các hàng con và cập nhật vào hàng tổng hợp
                if not child_rows.empty:
                    df.loc[index, columns_to_sum] = child_rows[columns_to_sum].sum()

        # Cập nhật hàng 'A' với tổng của các hàng '1' và '2'
        for index, row in df.iterrows():
            if row['STT'] == 'A':
                # Lấy giá trị của các hàng '1' và '2'
                row_1 = df[df['STT'] == '1'].iloc[:, 2:].sum()
                row_2 = df[df['STT'] == '2'].iloc[:, 2:].sum()

                # Tính tổng các hàng '1' và '2' để cập nhật vào hàng 'A'
                df.loc[index, columns_to_sum] = row_1 + row_2

        return df

    def create_dataframe_thuchien_tt(self,thang):
        ke_hoach_col = f"KẾ HOẠCH T{thang:02d}"
        thuc_hien_col = f"THỰC HIỆN T{thang:02d}"
        if thang == 1:
            thuc_hien_col = "THỰC HIỆN T1"
        columns = ["STT", "Dịch vụ", ke_hoach_col, thuc_hien_col, "% THỰC HIỆN", "KỲ TRƯỚC", "% VỚI KỲ TRƯỚC", "+/- VỚI KỲ TRƯỚC", "THỰC HIỆN T01", "% THỰC HIỆN T01", "+ /- VỚI T01"]

        # Dữ liệu
        data = {
            "STT": ["A", "1", "1.1", "1.1.1", "1.1.2", "1.1.3", "1.1.4", "1.1.5", "1.1.6", "1.1.7", "1.1.8", "1.1.9", "1.2", "1.2.1", "1.2.2", "2", "2.1", "2.1.1", "2.1.2", "2.1.3", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7"],
            "Dịch vụ": ["DOANH THU VTCNTT", "Công nghệ thông tin", "Dịch vụ số doanh nghiệp", "Hạ tầng CNTT", "Chính quyền số", "Y tế số", "Công nghệ nền tảng", "An toàn thông tin", "Giáo dục số", "Quản trị doanh nghiệp", "Vận tải và Logictic", "Phân phối bán lẻ", "Dịch vụ số cá nhân", "MyTV", "Dịch vụ tiện ích", "Dịch vụ Viễn thông", "Dịch vụ di động (5917)", "Di động trả trước", "Di động trả sau", "Gói GD,VP,Home", "Băng rộng", "Cố định", "Internet trực tiếp", "Truyền số liệu", "Dịch vụ GTGT khác", "Cho thuê Hạ tầng"],
        }

        # Tạo DataFrame với các giá trị NaN cho các cột còn lại
        # Sử dụng một dict comprehension để tạo ra các cột với giá trị NaN
        for col in columns[2:]:
            data[col] = [None] * len(data["STT"])

        # Tạo DataFrame
        kq_thuchien = pd.DataFrame(data, columns=columns)
        return kq_thuchien, ke_hoach_col, thuc_hien_col
    def map_kehoach_to_thuchien_tt(self,kq_kehoach,kq_thuchien,ke_hoach_col, column_name):
        for index, row in kq_kehoach.iterrows():
            service = row["Dịch_vụ"]

            # Tìm vị trí của dịch vụ trong kq_thuchien
            matching_rows = kq_thuchien[kq_thuchien["Dịch vụ"] == service]

            # Nếu có ít nhất một hàng khớp, ánh xạ dữ liệu
            if not matching_rows.empty:
                matching_index = matching_rows.index[0]
                kq_thuchien.loc[matching_index, ke_hoach_col] = row[column_name]
            else:
                print(f"Dịch vụ '{service}' không được tìm thấy trong DataFrame 'kq_thuchien'.")
        return kq_thuchien
    def update_column_data(self,column_data, kq_thuchien_tt, kq_thuchien):
        for index, row in kq_thuchien_tt.iterrows():
            service = row['nhom_dv']

            # Tìm vị trí của dịch vụ trong kq_thuchien
            matching_rows = kq_thuchien[kq_thuchien["Dịch vụ"] == service]

            # Nếu có ít nhất một hàng khớp, ánh xạ dữ liệu
            if not matching_rows.empty:
                matching_index = matching_rows.index[0]
                # Cập nhật giá trị tương ứng trong cột
                column_name = column_data.name
                kq_thuchien.at[matching_index, column_name] = row[1] / 1000000
            else:
                print(f"Dịch vụ '{service}' không được tìm thấy trong DataFrame 'kq_thuchien'.")
        
        return kq_thuchien
    def update_df_thuchien_kq2(self,kq_dataframe, thuc_hien_col, ke_hoach_col):
        # Tính cột '%thực hiện'
        kq_dataframe = kq_dataframe.fillna(0)
        kq_dataframe['% THỰC HIỆN'] = np.where(kq_dataframe[ke_hoach_col] != 0,
                                            kq_dataframe[thuc_hien_col] / kq_dataframe[ke_hoach_col] * 100,
                                            0)

        # Tính cột '%với kỳ trước'
        kq_dataframe['% VỚI KỲ TRƯỚC'] = np.where(kq_dataframe['KỲ TRƯỚC'] != 0,
                                                kq_dataframe[thuc_hien_col] / kq_dataframe['KỲ TRƯỚC'] * 100,
                                                0)

        # Tính cột '+- kỳ trước'
        kq_dataframe['+/- VỚI KỲ TRƯỚC'] = kq_dataframe[thuc_hien_col] - kq_dataframe['KỲ TRƯỚC']

        # Tính cột '%thực hiện T01'
        kq_dataframe['% THỰC HIỆN T01'] = np.where(kq_dataframe['THỰC HIỆN T01'] != 0,
                                                kq_dataframe[thuc_hien_col] / kq_dataframe['THỰC HIỆN T01'] * 100,
                                                0)

        # Tính cột '+- với T01'
        kq_dataframe['+ /- VỚI T01'] = kq_dataframe[thuc_hien_col] - kq_dataframe['THỰC HIỆN T01']

        # Replace NaN resulting from division by zero with 0


        return kq_dataframe
    def map_thuchienall_to_dataframe(self,kq_thuchien, kq_dataframe):
        for index, row in kq_thuchien.iterrows():
            service = row['nhom_dv']
            month = row['thang']
            revenue = row['doanh_thu']

            # Tìm vị trí của dịch vụ trong kq_dataframe
            matching_rows = kq_dataframe[kq_dataframe["Dịch vụ"] == service]

            # Nếu có ít nhất một hàng khớp, ánh xạ dữ liệu
            if not matching_rows.empty:
                matching_index = matching_rows.index[0]
                kq_dataframe.at[matching_index, f"Tháng {int(month):02d}"] = revenue/1000000
            else:
                print(f"Dịch vụ '{service}' không được tìm thấy trong DataFrame 'kq_dataframe'.")
        
        return kq_dataframe
    # DATAFRAME BAOCAOTONGHOP
    def create_dataframe_baocao_form1(self):
        columns = ["STT", "Dịch vụ","KH năm","KH tháng","TH tháng", "% với KH năm","% với KH tháng", "Tháng n-1", "+/- với tháng n-1", "Cùng kì", "% với cùng kì", "+/- với cùng kì"]

        data = {
            "STT": ["A", "1", "1.1", "1.1.1", "1.1.2", "1.1.3", "1.1.4", "1.1.5", "1.1.6", "1.1.7", "1.1.8", "1.1.9", "1.2", "1.2.1", "1.2.2", "2", "2.1", "2.1.1", "2.1.2", "2.1.3", "2.2", "2.3", "2.4", "2.5", "2.6", "2.7"],
            "Dịch vụ": ["DOANH THU VTCNTT", "Công nghệ thông tin", "Dịch vụ số doanh nghiệp", "Hạ tầng CNTT", "Chính quyền số", "Y tế số", "Công nghệ nền tảng", "An toàn thông tin", "Giáo dục số", "Quản trị doanh nghiệp", "Vận tải và Logictic", "Phân phối bán lẻ", "Dịch vụ số cá nhân", "MyTV", "Dịch vụ tiện ích", "Dịch vụ Viễn thông", "Dịch vụ di động (5917)", "Di động trả trước", "Di động trả sau", "Gói GD,VP,Home", "Băng rộng", "Cố định", "Internet trực tiếp", "Truyền số liệu", "Dịch vụ GTGT khác", "Cho thuê Hạ tầng"],
        }

        for col in columns[2:]:
            data[col] = [None] * len(data["STT"])
        kq_thuchien = pd.DataFrame(data, columns=columns)
        return kq_thuchien
    def update_df_baocao(self,kq_dataframe):
        # Tính cột '% KH năm
        kq_dataframe = kq_dataframe.fillna(0)
        kq_dataframe['% với KH năm'] = np.where(kq_dataframe['KH năm'] != 0,
                                            kq_dataframe['TH tháng'] / kq_dataframe['KH năm'] * 100,
                                            0)

        # Tính cột '%với KH tháng
        kq_dataframe['% với KH tháng'] = np.where(kq_dataframe['KH tháng'] != 0,
                                                kq_dataframe['TH tháng'] / kq_dataframe['KH tháng'] * 100,
                                                0)
            # Tính cột '%với cùng kì'
        kq_dataframe['% với cùng kì'] = np.where(kq_dataframe['Cùng kì'] != 0,
                                                kq_dataframe['TH tháng'] / kq_dataframe['Cùng kì'] * 100,
                                                0)

        # Tính cột '+- tháng n-1
        kq_dataframe['+/- với tháng n-1'] = kq_dataframe['TH tháng'] - kq_dataframe['Tháng n-1']

        # Tính cột '+- với cùng kì
        kq_dataframe['+/- với cùng kì'] = kq_dataframe['TH tháng'] - kq_dataframe['Cùng kì']
        
        return kq_dataframe
    def sum_hienhuu_moitrongnam(self,kq_thuchien_baocao_hienhuu, kq_thuchien_baocao_moitrongnam, kq_thuchien_thangtruoc_hienhuu, kq_thuchien_thangtruoc_moitrongnam, kq_kehoach_baocao_hienhuu, kq_kehoach_baocao_moitrongnam, kq_thuchien_baocao_cungki_hienhuu, kq_thuchien_baocao_cungki_moitrongnam):
        # Cộng từng cặp DataFrame tương ứng
        kq_thuchien_doanhthu = kq_thuchien_baocao_hienhuu.copy()
        kq_thuchien_doanhthu.iloc[:, 1:] += kq_thuchien_baocao_moitrongnam.iloc[:, 1:]

        kq_thangtruoc_doanhthu = kq_thuchien_thangtruoc_hienhuu.copy()
        kq_thangtruoc_doanhthu.iloc[:, 1:] += kq_thuchien_thangtruoc_moitrongnam.iloc[:, 1:]

        kq_kehoach_doanhthu = kq_kehoach_baocao_hienhuu.copy()
        kq_kehoach_doanhthu.iloc[:, 1:] += kq_kehoach_baocao_moitrongnam.iloc[:, 1:]

        kq_cungki_doanhthu = kq_thuchien_baocao_cungki_hienhuu.copy()
        kq_cungki_doanhthu.iloc[:, 1:] += kq_thuchien_baocao_cungki_moitrongnam.iloc[:, 1:]

        return kq_thuchien_doanhthu, kq_thangtruoc_doanhthu, kq_kehoach_doanhthu, kq_cungki_doanhthu
class MANAGERMENT_DATA:
    def __init__(self):
        self.data = None
        self.control_database = CONTROL_DATABASE()
    def get_ten_to(self,conn):
        cursor = conn.cursor(dictionary=True)
        cursor.execute("SELECT * FROM TEN_TO_CS")
        rows = cursor.fetchall()
        return rows
    def get_line(self, conn):
        cursor = conn.cursor()
        cursor.execute("SELECT Distinct(ten) FROM TEN_TO_CS")
        rows = cursor.fetchall()

        # Chỉ lấy giá trị 'ten' từ mỗi bản ghi và đưa vào một danh sách
        result = [row[0] for row in rows]

        return result

    # Hàm thêm tổ mới
    def add_ten_to(self,ten,conn, cursor):
        cursor.execute("INSERT INTO TEN_TO_CS (ten) VALUES (%s)", (ten,))
        conn.commit()

    # Hàm cập nhật tổ
    def update_ten_to(self,id, ten,conn,cursor):
        cursor.execute("UPDATE TEN_TO_CS SET ten = %s WHERE id = %s", (ten, id))
        conn.commit()

    # Hàm xóa tổ
    def delete_ten_to(self,id,conn,cursor):
        cursor.execute("DELETE FROM TEN_TO_CS WHERE id = %s", (id,))
        conn.commit()
class MAINCLASS:
    def __init__(self):
        self.data = None
        self.control_database = CONTROL_DATABASE()
        self.managerment_database = MANAGERMENT_DATA()
    def read_file_to_dataframe(self,file_path):
        # Đọc dữ liệu từ file Excel
        df = pd.read_excel(file_path)
        st.dataframe(df)
    def run(self, reponse_image_user_path):
        # Đường dẫn tới file ảnh
        file_path = "vnpt.png"
        current_dir = os.path.dirname(os.path.abspath(__file__))
        full_file_path = os.path.join(current_dir, file_path)

        # Đọc và mã hóa file ảnh
        
        if (reponse_image_user_path is not None):
            encoded_image = base64.b64encode(reponse_image_user_path.content).decode("utf-8")
                # st.session_state.encoded_image = encoded_image
        else:
            with open(full_file_path, "rb") as image_file:
                encoded_image = base64.b64encode(image_file.read()).decode("utf-8")
        st.sidebar.markdown(f"""
                <h1 style="text-align: center;position:fixed; top:5%;left:6%">PHÒNG KHDN3</h1>
                <div style="display: flex; justify-content: center; margin-bottom: 0;">
                    <img src='data:image/png;base64,{encoded_image}' alt='Team Image' width='40%' style='border-radius:50%; margin-bottom:10%;'>
                </div>
                """, unsafe_allow_html=True)

        # Thông tin chào mừng trong sidebar
        st.sidebar.markdown("""
    <div style="text-align: center;">
        <p style="font-size: 1.2em;">Cuộc sống đích thực!</p>
    </div>
    """, unsafe_allow_html=True)
        # st.sidebar.markdown("---")
    def documentation_text(self):
        with st.expander("📚 FORM MẪU NHẬP DỮ LIỆU"):
            tab1, tab2, tab3, tab4,tab5 = st.tabs(
                        ["📈 KẾ HOẠCH (HIỆN HỮU)", "📈 THỰC HIỆN (HIỆN HỮU)", "🗃 KẾ HOẠCH (TỔNG HỢP)", "🗃 THỰC HIỆN (TỔNG HỢP)","HƯỚNG DẪN SỬ"])
            with tab5:
                self.display_file_content("README.md")
            with tab1:
                with open('./form_import/form_mau_kehoachhienhuu.xlsx', "rb") as file:
                    btn = st.download_button(
                        label="Tải xuống tệp Excel",
                        data=file,
                        file_name="form_mau_kehoachhienhuu.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="form_mau_kehoachhienhuu"
                    )
                self.embed_image20("form_import/form_mau_kehoachhienhuu.png")
            with tab2:
                with open('./form_import/form_mau_thuchienhienhuu.xlsx', "rb") as file:
                    btn = st.download_button(
                        label="Tải xuống tệp Excel",
                        data=file,
                        file_name="form_mau_thuchienhienhuu.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="form_mau_thuchienhienhuu"
                    )
                self.embed_image20("form_import/form_mau_thuchienhienhuu.png")
            with tab3:
                with open('./form_import/form_mau_kehoachtonghop.xlsx', "rb") as file:
                    btn = st.download_button(
                        label="Tải xuống tệp Excel",
                        data=file,
                        file_name="form_mau_kehoachtonghop.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="form_mau_kehoachtonghop"
                    )
                self.embed_image20("form_import/form_mau_kehoachtonghop.png")
            with tab4:
                with open('./form_import/form_mau_thuchientonghop.xlsx', "rb") as file:
                    btn = st.download_button(
                        label="Tải xuống tệp Excel",
                        data=file,
                        file_name="form_mau_thuchientonghop.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="form_mau_thuchientonghop"
                    )
                self.embed_image20("form_import/form_mau_thuchientonghop.png")
    def streamlit_menu(self,role_access_admin):
        with st.sidebar:
            if role_access_admin is True:
                selected = option_menu(
                    menu_title=None,  # required
                    options=["Trang chủ", "Thêm dữ liệu","Xóa dữ liệu","Quản lý Line", "Người dùng","Quản Trị"],  # required
                    icons=["house", "plus","eraser", "gear","briefcase","list-task"],  # optional
                    menu_icon=None,  # optional
                    default_index=0,  # optional
                    styles={
                "container": {
                    "padding": "10px 0px", 
                    "margin": "0px auto",  
                    "border": "1px solid #000000",
                    "border-radius": "20px", 
                },
                "icon": {
                    "color": "#0069b4",  
                    "font-size": "16px"
                },
                "nav-link": {
                    "font-size": "16px", 
                    "text-align": "left",  
                    "--hover-color": "#B3D3F0"
                },
                "nav-link-selected": {
                    "background-color": "#0069b4", 
                    "font-size": "16px",
                    "font-family": "Tahoma, Geneva, sans-serif",
                },
            }
                )
            else:
                selected = option_menu(
                    menu_title=None,  # required
                    options=["Trang chủ", "Thêm dữ liệu","Xóa dữ liệu","Quản lý Line", "Người dùng"],  # required
                    icons=["house", "plus","eraser", "gear","briefcase"],  # optional
                    menu_icon=None,  # optional
                    default_index=0,  # optional
                    styles={
                "container": {
                    "padding": "10px 0px", 
                    "margin": "0px auto",  
                    "border": "1px solid #000000",
                    "border-radius": "20px", 
                },
                "icon": {
                    "color": "#0069b4",  
                    "font-size": "16px"
                },
                "nav-link": {
                    "font-size": "16px", 
                    "text-align": "left",  
                    "--hover-color": "#B3D3F0"
                },
                "nav-link-selected": {
                    "background-color": "#0069b4", 
                    "font-size": "16px",
                    "font-family": "Tahoma, Geneva, sans-serif",
                },
            }
                )
        return selected
    def main(self,selected,role_access_admin): 
        if role_access_admin is True:
            if selected == "Trang chủ":
                self.view()
                # self.datetimerefress()
            if selected == "Thêm dữ liệu":
                self.view_insert_data()
                if st.button("FORM MẪU NHẬP DỮ LIỆU", key="formmaunhapdulieu_expander_admin"):
                    self.documentation_text()
            if selected == "Quản lý Line":
                self.managerment_line()
            if selected == "Xóa dữ liệu":
                self.view_managerment_data()
            if selected == "Người dùng":
                self.layout_select_view()
            if selected == "Quản Trị": 
                self.info_contact()
        else:
            if selected == "Trang chủ":
                self.view()
                # self.datetimerefress()
            if selected == "Thêm dữ liệu":
                self.view_insert_data()
                if st.button("FORM MẪU NHẬP DỮ LIỆU", key="formmaunhapdulieu_expander"):
                    self.documentation_text()
            if selected == "Quản lý Line":
                self.managerment_line()
            if selected == "Xóa dữ liệu":
                self.view_managerment_data()
            if selected == "Người dùng":
                self.layout_select_view()

    def embed_image(self, file_path):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        full_file_path = os.path.join(current_dir, file_path)
        with open(full_file_path, "rb") as image_file:
            encoded_image = base64.b64encode(image_file.read()).decode("utf-8")
        html_code = f"""
        <div style="display: flex; justify-content: center;">
            <img src='data:image/jpeg;base64,{encoded_image}' alt='Ten_Hinh_Anh' width='100%' style='border-radius:60%; margin-bottom:5%;'>
        </div>
        """
        st.markdown(html_code, unsafe_allow_html=True)
    def embed_image20(self, file_path):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        full_file_path = os.path.join(current_dir, file_path)
        with open(full_file_path, "rb") as image_file:
            encoded_image = base64.b64encode(image_file.read()).decode("utf-8")
        html_code = f"""
        <div style="display: flex; justify-content: center;">
            <img src='data:image/jpeg;base64,{encoded_image}' alt='Ten_Hinh_Anh' width='70%'>
        </div>
        """
        st.markdown(html_code, unsafe_allow_html=True)
    def display_file_content(self, file_path):
        current_dir = os.path.dirname(os.path.abspath(__file__))
        full_file_path = os.path.join(current_dir, file_path)

        if os.path.exists(full_file_path):
            with open(full_file_path, "r", encoding="utf-8") as file:
                try:
                    lines = file.readlines()
                    content = "\n".join(lines).strip()
                    st.info(f"### Giới thiệu\n{content}")
                except UnicodeDecodeError:
                    st.error(
                        f"Tệp tin '{full_file_path}' không thể đọc với encoding utf-8.")
        else:
            st.error(f"Tệp tin '{full_file_path}' không tồn tại.")
    def download_excel(self,df, file_name):
            # Tạo một workbook mới với openpyxl
        workbook = Workbook()

        # Tạo một sheet trong workbook
        sheet = workbook.active
        sheet.title = "Data"

        # Ghi tiêu đề của DataFrame vào sheet
        for idx, col in enumerate(df.columns, start=1):
            sheet.cell(row=1, column=idx).value = col

        # Ghi dữ liệu từ DataFrame vào sheet
        for r_idx, row in enumerate(df.itertuples(), start=2):
            for c_idx, value in enumerate(row[1:], start=1):
                sheet.cell(row=r_idx, column=c_idx).value = value

        # Điều chỉnh phông chữ, cỡ chữ và đường viền
        font = Font(name='Times New Roman', size=12)
        border = Border(left=Side(border_style='thin'), 
                        right=Side(border_style='thin'), 
                        top=Side(border_style='thin'), 
                        bottom=Side(border_style='thin'))

        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            for cell in row:
                cell.font = font
                cell.border = border

        # Tự động fit nội dung theo chiều rộng của các cột
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(cell.value)
                except:
                    pass
            adjusted_width = (max_length + 2) * 1.2
            sheet.column_dimensions[column].width = adjusted_width

        # Tạo một bộ đệm để lưu trữ dữ liệu Excel
        excel_buffer = io.BytesIO()
        # Lưu workbook vào bộ đệm
        workbook.save(excel_buffer)
        # Lấy dữ liệu từ bộ đệm
        excel_data = excel_buffer.getvalue()

        # Tạo URL để tải xuống file Excel
        b64 = base64.b64encode(excel_data).decode()
        href = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{b64}" download="{file_name}">Tải xuống kết quả</a>'
        st.markdown(href, unsafe_allow_html=True)

    def view(self):
        if 'conn' not in st.session_state:
            st.session_state.conn, st.session_state.cursor = connect_to_mysql()
        conn = st.session_state.conn
        if conn is None:
            st.session_state.conn, st.session_state.cursor = connect_to_mysql()
            conn = st.session_state.conn
            cursor =conn.cursor()
        st.header("Thực hiện lựa chọn", anchor="Thực hiện lựa chọn",)
        col_view_1, col_view_2 = st.columns(2,gap="small")
        with col_view_1:
            self.option_select_report = st.radio("Chọn loại báo cáo", ('HIỆN HỮU', 'TỔNG HỢP'),horizontal=True,key="radio_chonloaibaocao_page_viewall",help="Chọn loại báo cáo cần xem")
            lines = self.managerment_database.get_line(st.session_state.conn)
            self.lines = st.empty()
            self.line_selected = self.lines.selectbox('Line', lines,key="Line_view_all",help="Chọn line cần xem")
            
            
        with col_view_2:
            loaibaocao = ['tổng doanh thu', 'hiện hữu', 'mới trong năm', 'mới trong tháng']
            
            # st.info(f"Chọn năm: {self.year}")
            if self.option_select_report == 'HIỆN HỮU':
                self.option = st.radio(
                                        "Chọn một tùy chọn",
                                        ('KẾ HOẠCH', 'THỰC HIỆN','TOÀN BỘ'), horizontal=True, key="option_view_all",help="Loại file muốn xem và tải về"
                                    )
            if self.option_select_report == 'TỔNG HỢP':
                self.lines.empty()
                self.option = st.radio(
                                        "Chọn một tùy chọn",
                                        ('KẾ HOẠCH TỔNG HỢP', 'THỰC HIỆN TỔNG HỢP', 'KẾT HỢP'), horizontal=True, key="baocao_view_all"
                                    )
                self.select_report_sum = st.selectbox("Chọn chế độ:", ('không lũy kế','lũy kế'), key="select_option_sum_luyke")
                if self.option != 'KẾT HỢP':
                    self.loaibaocao = st.selectbox("Chọn tùy chọn báo cáo",loaibaocao,key="loaibaocao_view_all")
            # self.year = st.text_input("Nhập năm:", "")
            self.year = st.selectbox("Chọn năm:",list(range(2000, date.today().year + 1)), index=date.today().year - 2000,key="select_year_view_all")
        if self.option == 'THỰC HIỆN' or self.option == 'TOÀN BỘ' or self.option == 'THỰC HIỆN TỔNG HỢP' or self.option == 'KẾT HỢP':
            months = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
            # Hiển thị các tháng trong một selectbox với giá trị hiển thị và giá trị thực tế
            self.month_select = st.selectbox("Chọn tháng:", months, key="select_month_view_all")
            self.column_name = f'Tháng_{int(self.month_select):02d}'
        if self.option == 'TOÀN BỘ':
            self.lines.empty()
        button_view_tt = st.empty()
        if button_view_tt.button('Xem', use_container_width=True, key="view_all_page"):
            control_database = CONTROL_DATABASE()
            control_dataframe = DATAFRAME()
            cursor =conn.cursor()
            if self.option == 'KẾ HOẠCH':
                df_kehoach_view = control_database.query_kehoach_by_line_year(self.line_selected, self.year,conn, cursor)
                df_kehoach_dataframe = control_dataframe.create_dataframe_public()
                df_kehoach_ketqua = control_dataframe.map_kehoach_to_dataframe(df_kehoach_view, df_kehoach_dataframe)
                df_kehoach_ketqua = control_dataframe.update_summary_rows(df_kehoach_ketqua)
                if (df_kehoach_view.empty):
                    st.warning("Không có dữ liệu kế hoạch cho năm và line đã chọn.")
                else:
                    df_kehoach_ketqua.iloc[:,2:]= df_kehoach_ketqua.iloc[:,2:].astype(float).round(0) 
                    file_name = f"kehoach_{self.line_selected}_{self.year}.xlsx"
                    self.download_excel(df_kehoach_ketqua, file_name)
                    st.dataframe(df_kehoach_ketqua,width=1400)
            if self.option == 'THỰC HIỆN':
                progress_bar = st.progress(0)
                # GET DATA FROM DATABASE FORM 1
                df_thuchien_hientai = control_database.get_data_thuchien(self.month_select, self.year, self.line_selected,conn, cursor)
                progress_bar.progress(25)
                df_thuchien_t01 = control_database.get_data_thuchien(1, self.year, self.line_selected,conn, cursor)
                progress_bar.progress(50)
                df_thuchien_thangtruoc = control_database.get_data_thuchien(int(self.month_select) - 1, self.year, self.line_selected,conn, cursor)
                progress_bar.progress(70)
                df_kehoach_hientai =control_database.get_ketqua_kehoach_tt(int(self.month_select), self.year, self.line_selected,conn, cursor)
                # MAP DATA TO DATAFRAME FORM 1
                df_thuchien_view, ke_hoach_col, thuc_hien_col = control_dataframe.create_dataframe_thuchien_tt(int(self.month_select))
                df_thuchien_view = control_dataframe.map_kehoach_to_thuchien_tt(df_kehoach_hientai, df_thuchien_view, ke_hoach_col, self.column_name)
                df_thuchien_view = control_dataframe.update_column_data(df_thuchien_view[thuc_hien_col],df_thuchien_hientai, df_thuchien_view)
                df_thuchien_view = control_dataframe.update_column_data(df_thuchien_view['THỰC HIỆN T01'],df_thuchien_t01, df_thuchien_view)
                df_thuchien_view = control_dataframe.update_column_data(df_thuchien_view['KỲ TRƯỚC'],df_thuchien_thangtruoc, df_thuchien_view)
                df_thuchien_view = control_dataframe.update_summary_rows(df_thuchien_view)
                
                df_thuchien_view = control_dataframe.update_df_thuchien_kq2(df_thuchien_view, thuc_hien_col, ke_hoach_col)
                # GET DATA FROM DATABASE FORM 2
                df_thuchien_all = control_database.get_data_thuchien_all(self.year, self.line_selected,conn, cursor)
                progress_bar.progress(90)
                # MAP DATA TO DATAFRAME FORM 2
                df_thuchien_all_view = control_dataframe.create_dataframe_public()
                df_thuchien_all_view = control_dataframe.map_thuchienall_to_dataframe(df_thuchien_all, df_thuchien_all_view)
                df_thuchien_all_view = control_dataframe.update_summary_rows(df_thuchien_all_view)
                progress_bar.empty()
                if (df_thuchien_hientai.empty):
                    st.warning("Không có dữ liệu thực hiện cho tháng, năm và line đã chọn.")
                else:
                    df_thuchien_view.iloc[:,2:]= df_thuchien_view.iloc[:,2:].astype(float).round(0)
                    df_thuchien_view['% THỰC HIỆN'] = df_thuchien_view['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
                    df_thuchien_view['% VỚI KỲ TRƯỚC'] = df_thuchien_view['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
                    df_thuchien_view['% THỰC HIỆN T01'] = df_thuchien_view['% THỰC HIỆN T01'].apply(lambda x: '{:.0f}%'.format(x))
                    df_thuchien_all_view.iloc[:,2:]= df_thuchien_all_view.iloc[:,2:].astype(float).round(0)
                    df_thuchien_all_view=df_thuchien_all_view.fillna(0)
                    file_name_form1 = f"thuchien_{self.line_selected}_{self.year}_{self.month_select}_form1.xlsx"
                    self.download_excel(df_thuchien_view, file_name_form1)
                    st.dataframe(df_thuchien_view)
                    
                    file_name_form2 = f"thuchien_{self.line_selected}_{self.year}_{self.month_select}_form2.xlsx"
                    self.download_excel(df_thuchien_all_view, file_name_form2)
                    st.dataframe(df_thuchien_all_view)
            if self.option == 'KẾ HOẠCH TỔNG HỢP':
                if self.select_report_sum == 'không lũy kế':
                    kq_kehoach_baocao = control_database.query_kehoach_by_line_year(self.loaibaocao, self.year,conn, cursor)
                else:
                    kq_kehoach_baocao = control_database.query_kehoach_by_line_year_luyke(self.loaibaocao, self.year,conn, cursor)
                dataframe_kehoach_baocao = control_dataframe.create_dataframe_public()
                dataframe_kehoach_baocao = control_dataframe.map_kehoach_to_dataframe(kq_kehoach_baocao, dataframe_kehoach_baocao)
                dataframe_kehoach_baocao = control_dataframe.update_summary_rows(dataframe_kehoach_baocao)
                if (kq_kehoach_baocao.empty):
                    st.warning("Không có dữ liệu kế hoạch cho năm và loại báo cáo đã chọn.")
                else:
                    dataframe_kehoach_baocao.iloc[:,2:]= dataframe_kehoach_baocao.iloc[:,2:].astype(float).round(0) 
                    file_name = f"kehoach_{self.loaibaocao}_{self.year}.xlsx"
                    self.download_excel(dataframe_kehoach_baocao, file_name)
                    st.dataframe(dataframe_kehoach_baocao)
            if self.option == 'THỰC HIỆN TỔNG HỢP':
                if self.loaibaocao != 'tổng doanh thu':
                    # FORM 1
                    dataframe_baocao_form1 = control_dataframe.create_dataframe_baocao_form1() 
                    if self.select_report_sum == 'không lũy kế':
                        kq_thuchien_hientai_baocao = control_database.get_data_thuchien_baocao(self.month_select, self.year, self.loaibaocao,conn, cursor)
                        kq_thuchien_thangtruoc_baocao = control_database.get_data_thuchien_baocao(int(self.month_select) - 1, self.year, self.loaibaocao,conn, cursor)
                        kq_kehoach_hientai_baocao = control_database.get_ketqua_kehoach_baocao(int(self.month_select), self.year, self.loaibaocao,conn, cursor)
                    else:
                        kq_thuchien_hientai_baocao = control_database.get_data_thuchien_baocao_luyke(self.month_select, self.year, self.loaibaocao,conn, cursor)
                        kq_thuchien_thangtruoc_baocao = control_database.get_data_thuchien_baocao_luyke(int(self.month_select) - 1, self.year, self.loaibaocao,conn, cursor)
                        kq_kehoach_hientai_baocao = control_database.get_ketqua_kehoach_baocao_luyke(int(self.month_select), self.year, self.loaibaocao,conn, cursor)
                    dataframe_baocao_form1 = control_dataframe.update_column_data(dataframe_baocao_form1['TH tháng'], kq_thuchien_hientai_baocao, dataframe_baocao_form1)
                    dataframe_baocao_form1 = control_dataframe.update_column_data(dataframe_baocao_form1['Tháng n-1'], kq_thuchien_thangtruoc_baocao, dataframe_baocao_form1)
                    dataframe_baocao_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_hientai_baocao, dataframe_baocao_form1, 'KH tháng', self.column_name)
                    dataframe_baocao_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_hientai_baocao, dataframe_baocao_form1, 'KH năm', 'LK_năm')
                    if self.year -1  == 2023:
                        if self.select_report_sum == 'không lũy kế':
                            kq_thuchien_cungki = control_database.get_ketqua_thuchien_2023(int(self.month_select), self.loaibaocao,conn, cursor)
                        else:
                            kq_thuchien_cungki = control_database.get_ketqua_thuchien_2023_luyke(int(self.month_select), self.loaibaocao,conn, cursor)
                        dataframe_baocao_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_thuchien_cungki, dataframe_baocao_form1, 'Cùng kì', self.column_name)
                    else:
                        if self.select_report_sum == 'không lũy kế':
                            kq_thuchien_cungki = control_database.get_data_thuchien_baocao(int(self.month_select), self.year-1, self.loaibaocao,conn, cursor)
                        else:
                            kq_thuchien_cungki = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year-1, self.loaibaocao,conn, cursor)
                        dataframe_baocao_form1 = control_dataframe.update_column_data(dataframe_baocao_form1['Cùng kì'], kq_thuchien_cungki, dataframe_baocao_form1)
                    dataframe_baocao_form1 = control_dataframe.update_summary_rows(dataframe_baocao_form1)
                    dataframe_baocao_form1 = control_dataframe.update_df_baocao(dataframe_baocao_form1)
                    # FORM 2
                    form2_dataframe_baocao = control_dataframe.create_dataframe_public()
                    if self.select_report_sum == 'không lũy kế':
                        kq_thuchien_all_baocao = control_database.get_data_form2_baocao(self.year,self.loaibaocao,conn, cursor)
                    else: 
                        kq_thuchien_all_baocao= control_database.get_data_form2_baocao_luyke(self.year,self.loaibaocao,conn, cursor)
                    form2_dataframe_baocao = control_dataframe.map_thuchienall_to_dataframe(kq_thuchien_all_baocao, form2_dataframe_baocao)
                    form2_dataframe_baocao = control_dataframe.update_summary_rows(form2_dataframe_baocao)
                    if (kq_thuchien_hientai_baocao.empty):
                        st.warning("Không có dữ liệu thực hiện cho tháng, năm và loại báo cáo đã chọn.")
                    else:
                        dataframe_baocao_form1.iloc[:,2:]= dataframe_baocao_form1.iloc[:,2:].astype(float).round(0)
                        form2_dataframe_baocao.iloc[:,2:]= form2_dataframe_baocao.iloc[:,2:].astype(float).round(0)
                        form2_dataframe_baocao=form2_dataframe_baocao.fillna(0)
                        dataframe_baocao_form1['% với KH năm'] = dataframe_baocao_form1['% với KH năm'].apply(lambda x: '{:.0f}%'.format(x))
                        dataframe_baocao_form1['% với KH tháng'] = dataframe_baocao_form1['% với KH tháng'].apply(lambda x: '{:.0f}%'.format(x))
                        dataframe_baocao_form1['% với cùng kì'] = dataframe_baocao_form1['% với cùng kì'].apply(lambda x: '{:.0f}%'.format(x))
                        file_name_form1 = f"thuchien_{self.loaibaocao}_{self.year}_{self.month_select}_form1.xlsx"
                        self.download_excel(dataframe_baocao_form1, file_name_form1)
                        st.dataframe(dataframe_baocao_form1)
                        
                        file_name_form2 = f"thuchien_{self.loaibaocao}_{self.year}_{self.month_select}_form2.xlsx"
                        self.download_excel(form2_dataframe_baocao, file_name_form2)
                        st.dataframe(form2_dataframe_baocao)
                else:
                    # FORM 1
                    dataframe_baocao_form1 = control_dataframe.create_dataframe_baocao_form1()
                    if self.select_report_sum == 'không lũy kế':
                        kq_thuchien_baocao_moitrongnam = control_database.get_data_thuchien_baocao(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                        kq_thuchien_thangtruoc_moitrongnam = control_database.get_data_thuchien_baocao(int(self.month_select) - 1, self.year, 'mới trong năm',conn, cursor)
                        kq_kehoach_baocao_moitrongnam = control_database.get_ketqua_kehoach_baocao(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                        kq_thuchien_baocao_hienhuu = control_database.get_data_thuchien_baocao(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                        kq_thuchien_thangtruoc_hienhuu = control_database.get_data_thuchien_baocao(int(self.month_select) - 1, self.year, 'hiện hữu',conn, cursor)
                        kq_kehoach_baocao_hienhuu = control_database.get_ketqua_kehoach_baocao(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                        if self.year - 1 ==2023:
                            kq_thuchien_baocao_cungki_moitrongnam = control_database.get_ketqua_thuchien_2023(int(self.month_select), 'mới trong năm',conn, cursor)
                            kq_thuchien_baocao_cungki_hienhuu = control_database.get_ketqua_thuchien_2023(int(self.month_select), 'hiện hữu',conn, cursor)
                        else:
                            kq_thuchien_baocao_cungki_moitrongnam = control_database.get_data_thuchien_baocao(int(self.month_select), self.year-1, 'mới trong năm',conn, cursor)
                            kq_thuchien_baocao_cungki_hienhuu = control_database.get_data_thuchien_baocao(int(self.month_select), self.year-1, 'hiện hữu',conn, cursor)
                                                # FORM 2 
                        kq_thuchien_all_baocao_hienhuu = control_database.get_data_form2_baocao(self.year,'hiện hữu',conn, cursor)
                        kq_thuchien_all_baocao_moitrongnam = control_database.get_data_form2_baocao(self.year,'mới trong năm',conn, cursor)
                    else:
                        kq_thuchien_baocao_moitrongnam = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                        kq_thuchien_thangtruoc_moitrongnam = control_database.get_data_thuchien_baocao_luyke(int(self.month_select) - 1, self.year, 'mới trong năm',conn, cursor)
                        kq_kehoach_baocao_moitrongnam = control_database.get_ketqua_kehoach_baocao_luyke(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                        kq_thuchien_baocao_hienhuu = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                        kq_thuchien_thangtruoc_hienhuu = control_database.get_data_thuchien_baocao_luyke(int(self.month_select) - 1, self.year, 'hiện hữu',conn, cursor)
                        kq_kehoach_baocao_hienhuu = control_database.get_ketqua_kehoach_baocao_luyke(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                        if self.year - 1 ==2023:
                            kq_thuchien_baocao_cungki_moitrongnam = control_database.get_ketqua_thuchien_2023_luyke(int(self.month_select), 'mới trong năm',conn, cursor)
                            kq_thuchien_baocao_cungki_hienhuu = control_database.get_ketqua_thuchien_2023_luyke(int(self.month_select), 'hiện hữu',conn, cursor)
                        else:
                            kq_thuchien_baocao_cungki_moitrongnam = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year-1, 'mới trong năm',conn, cursor)
                            kq_thuchien_baocao_cungki_hienhuu = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year-1, 'hiện hữu',conn, cursor)
                        # FORM 2
                        kq_thuchien_all_baocao_hienhuu = control_database.get_data_form2_baocao_luyke(self.year,'hiện hữu',conn, cursor)
                        kq_thuchien_all_baocao_moitrongnam = control_database.get_data_form2_baocao_luyke(self.year,'mới trong năm',conn, cursor)
                    kq_thuchien_all_tongdoanhthu = kq_thuchien_all_baocao_hienhuu.copy()
                    kq_thuchien_all_tongdoanhthu.iloc[:, 1] += kq_thuchien_all_baocao_moitrongnam.iloc[:, 1]
                    form2_dataframe_baocao = control_dataframe.create_dataframe_public()
                    form2_dataframe_baocao = control_dataframe.map_thuchienall_to_dataframe(kq_thuchien_all_tongdoanhthu, form2_dataframe_baocao)
                    form2_dataframe_baocao = control_dataframe.update_summary_rows(form2_dataframe_baocao)
                    if (kq_thuchien_baocao_moitrongnam.empty or kq_thuchien_baocao_hienhuu.empty):
                        st.warning("Không có dữ liệu thực hiện cho tháng, năm và loại báo cáo đã chọn.")
                    else:
                        kq_thuchien_doanhthu, kq_thangtruoc_doanhthu, kq_kehoach_doanhthu, kq_cungki_doanhthu = control_dataframe.sum_hienhuu_moitrongnam(kq_thuchien_baocao_hienhuu, kq_thuchien_baocao_moitrongnam, kq_thuchien_thangtruoc_hienhuu, kq_thuchien_thangtruoc_moitrongnam, kq_kehoach_baocao_hienhuu, kq_kehoach_baocao_moitrongnam, kq_thuchien_baocao_cungki_hienhuu, kq_thuchien_baocao_cungki_moitrongnam)
                        dataframe_baocao_form1 = control_dataframe.update_column_data(dataframe_baocao_form1['TH tháng'], kq_thuchien_doanhthu, dataframe_baocao_form1)
                        dataframe_baocao_form1 = control_dataframe.update_column_data(dataframe_baocao_form1['Tháng n-1'], kq_thangtruoc_doanhthu, dataframe_baocao_form1)
                        dataframe_baocao_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_doanhthu, dataframe_baocao_form1, 'KH tháng', self.column_name)
                        dataframe_baocao_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_doanhthu, dataframe_baocao_form1, 'KH năm', 'LK_năm')
                        if self.year - 1 ==2023:
                            dataframe_baocao_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_cungki_doanhthu, dataframe_baocao_form1, 'Cùng kì', self.column_name)
                        else:
                            dataframe_baocao_form1 = control_dataframe.update_column_data(dataframe_baocao_form1['Cùng kì'], kq_cungki_doanhthu, dataframe_baocao_form1)
                        dataframe_baocao_form1 = control_dataframe.update_summary_rows(dataframe_baocao_form1)
                        dataframe_baocao_form1 = control_dataframe.update_df_baocao(dataframe_baocao_form1)
                        dataframe_baocao_form1.iloc[:,2:]= dataframe_baocao_form1.iloc[:,2:].astype(float).round(0)
                        form2_dataframe_baocao.iloc[:,2:]= form2_dataframe_baocao.iloc[:,2:].astype(float).round(0)
                        form2_dataframe_baocao=form2_dataframe_baocao.fillna(0)
                        dataframe_baocao_form1['% với KH năm'] = dataframe_baocao_form1['% với KH năm'].apply(lambda x: '{:.0f}%'.format(x))
                        dataframe_baocao_form1['% với KH tháng'] = dataframe_baocao_form1['% với KH tháng'].apply(lambda x: '{:.0f}%'.format(x))
                        dataframe_baocao_form1['% với cùng kì'] = dataframe_baocao_form1['% với cùng kì'].apply(lambda x: '{:.0f}%'.format(x))
                        file_name_form1 = f"thuchien_{self.loaibaocao}_{self.year}_{self.month_select}_form1.xlsx"
                        self.download_excel(dataframe_baocao_form1, file_name_form1)
                        st.dataframe(dataframe_baocao_form1)
                        
                        file_name_form2 = f"thuchien_{self.loaibaocao}_{self.year}_{self.month_select}_form2.xlsx"
                        self.download_excel(form2_dataframe_baocao, file_name_form2)
                        st.dataframe(form2_dataframe_baocao)
                        
                
        if self.option == 'TOÀN BỘ':
            button_view_tt.empty()
            name_line_rut = ['IDC', 'SBN','SME1', 'SME2', 'SME3', 'BDS']
            control_database = CONTROL_DATABASE()
            control_dataframe = DATAFRAME()
            if st.button('Bắt đầu tạo', key="batdautaofileall"):
                cursor =conn.cursor()
                self.check_status_empty = False
                output_all_tt = io.BytesIO()
                wb = Workbook()
                for line, sheet_name in zip(lines, name_line_rut):
                    ws = wb.create_sheet(title=sheet_name)
                    ws['A1'] = "Phòng"
                    ws['B1'] = "DN3"
                    ws['A2'] = "Line"
                    ws['B2'] = line 
                    df_kehoach_view = control_database.query_kehoach_by_line_year(line, self.year,conn, cursor)
                    if df_kehoach_view.empty:
                        st.warning("Không có dữ liệu kế hoạch cho năm và line đã chọn.")
                        self.check_status_empty = True
                        break
                    else:
                        df_kehoach_dataframe = control_dataframe.create_dataframe_public()
                        df_kehoach_ketqua = control_dataframe.map_kehoach_to_dataframe(df_kehoach_view, df_kehoach_dataframe)
                        df_kehoach_ketqua = control_dataframe.update_summary_rows(df_kehoach_ketqua)
                        df_kehoach_ketqua.iloc[:,2:]= df_kehoach_ketqua.iloc[:,2:].astype(float).round(0)
                    
                    # GET DATA FROM DATABASE FORM 1
                    df_thuchien_hientai = control_database.get_data_thuchien(self.month_select, self.year, line,conn, cursor)
                    df_thuchien_t01 = control_database.get_data_thuchien(1, self.year, line,conn, cursor)
                    df_thuchien_thangtruoc = control_database.get_data_thuchien(int(self.month_select) - 1, self.year, line,conn, cursor)
                    df_kehoach_hientai =control_database.get_ketqua_kehoach_tt(int(self.month_select), self.year, line,conn, cursor)
                    if df_thuchien_hientai.empty:
                        st.warning("Không có dữ liệu thực hiện cho tháng, năm và line đã chọn.")
                        self.check_status_empty = True
                        break
                    else:
                    # MAP DATA TO DATAFRAME FORM 1
                        df_thuchien_view, ke_hoach_col, thuc_hien_col = control_dataframe.create_dataframe_thuchien_tt(int(self.month_select))
                        df_thuchien_view = control_dataframe.map_kehoach_to_thuchien_tt(df_kehoach_hientai, df_thuchien_view, ke_hoach_col, self.column_name)
                        df_thuchien_view = control_dataframe.update_column_data(df_thuchien_view[thuc_hien_col],df_thuchien_hientai, df_thuchien_view)
                        df_thuchien_view = control_dataframe.update_column_data(df_thuchien_view['THỰC HIỆN T01'],df_thuchien_t01, df_thuchien_view)
                        df_thuchien_view = control_dataframe.update_column_data(df_thuchien_view['KỲ TRƯỚC'],df_thuchien_thangtruoc, df_thuchien_view)
                        df_thuchien_view = control_dataframe.update_summary_rows(df_thuchien_view)
                        df_thuchien_view = control_dataframe.update_df_thuchien_kq2(df_thuchien_view, thuc_hien_col, ke_hoach_col)
                    # GET DATA FROM DATABASE FORM 2
                    df_thuchien_all = control_database.get_data_thuchien_all(self.year, line,conn, cursor)
                    if df_thuchien_all.empty:
                        st.warning("Không có dữ liệu thực hiện cho năm và line đã chọn.")
                        self.check_status_empty = True
                        break
                    else:
                    # MAP DATA TO DATAFRAME FORM 2
                        df_thuchien_all_view = control_dataframe.create_dataframe_public()
                        df_thuchien_all_view = control_dataframe.map_thuchienall_to_dataframe(df_thuchien_all, df_thuchien_all_view)
                        df_thuchien_all_view = control_dataframe.update_summary_rows(df_thuchien_all_view)
                        df_thuchien_view.iloc[:,2:]= df_thuchien_view.iloc[:,2:].astype(float).round(0)
                        df_thuchien_view['% THỰC HIỆN'] = df_thuchien_view['% THỰC HIỆN'].apply(lambda x: '{:.0f}%'.format(x))
                        df_thuchien_view['% VỚI KỲ TRƯỚC'] = df_thuchien_view['% VỚI KỲ TRƯỚC'].apply(lambda x: '{:.0f}%'.format(x))
                        df_thuchien_view['% THỰC HIỆN T01'] = df_thuchien_view['% THỰC HIỆN T01'].apply(lambda x: '{:.0f}%'.format(x))
                        df_thuchien_all_view.iloc[:,2:]= df_thuchien_all_view.iloc[:,2:].astype(float).round(0)
                        df_thuchien_all_view=df_thuchien_all_view.fillna(0)
                    # Tạo sheet mới với tên từ 'line'
                            # Ghi dữ liệu từ DataFrame vào sheet
                    data = [df_kehoach_ketqua.columns.tolist()] + df_kehoach_ketqua.values.tolist()
                    for r_idx, row in enumerate(data, 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx + 3, column=c_idx, value=value)
                            cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

                    data = [df_thuchien_all_view.columns.tolist()] + df_thuchien_all_view.values.tolist()
                    for r_idx, row in enumerate(data, 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx + 3, column=c_idx + 15, value=value)
                            cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền

                    data = [df_thuchien_view.columns.tolist()] + df_thuchien_view.values.tolist()
                    for r_idx, row in enumerate(data, 1):
                        for c_idx, value in enumerate(row, 1):
                            cell = ws.cell(row=r_idx + 3 + 29, column=c_idx, value=value)
                            cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))  # Thiết lập đường viền
                    for col_range in [(3, 15), (16, 30)]:
                        start_col, end_col = col_range
                        start_cell = ws.cell(row=3, column=start_col)
                        end_cell = ws.cell(row=3, column=end_col)
                        ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
                        
                        # Thiết lập lại border cho các ô trong vùng merge
                        for row in ws.iter_rows(min_row=3, max_row=3, min_col=start_col, max_col=end_col):
                            for cell in row:
                                cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                    top=Side(border_style='thin'), bottom=Side(border_style='thin'))

                    # Gán giá trị cho ô merged và thiết lập font, đường viền
                    for col_idx, text in [(3, 'KẾ HOẠCH'), (16, 'THỰC HIỆN')]:
                        cell = ws.cell(row=3, column=col_idx, value=text)
                        cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                        cell.font = Font(name='Times New Roman', size=12, bold=True)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                     # In đậm các dòng cần thiết
                    bold_rows_need_thuchien = [4,5,6,7,17,20, 21,25,26,27,28,29,30,33,34,35,36,46,49,50,54,55,56,57,58,59] 
                    for row_idx in bold_rows_need_thuchien:
                        for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_idx, max_row=row_idx):
                            for cell in col:
                                cell.font = Font(bold=True)
                    # Tự động điều chỉnh độ rộng của các cột dựa trên nội dung của ô
                    for column_cells in ws.columns:
                        length = max(len(str(cell.value)) for cell in column_cells)
                        ws.column_dimensions[column_cells[0].column_letter].width = length + 2
                if self.check_status_empty == False:
                    wb.save(output_all_tt)
                    output_all_tt.seek(0)
                    excel_base64 = base64.b64encode(output_all_tt.read()).decode()

                    # Create a download link using st.download_button
                    file_name_download_all_tt = f"kq_thuchien_{self.year}_thang_{self.month_select}.xlsx"
                    download_link_all_tt = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{file_name_download_all_tt}.xlsx">Tải xuống kết quả</a>'
                    st.success("Tạo file thành công.")
                    st.markdown(download_link_all_tt, unsafe_allow_html=True)
        if self.option == 'KẾT HỢP':
            button_view_tt.empty()
            control_database = CONTROL_DATABASE()
            control_dataframe = DATAFRAME()
            if st.button('Bắt đầu tạo',key="batdautaofilekethop"):
                cursor =conn.cursor()
                self.check_status_empty = False
                output_all_baocao = io.BytesIO()
                wb = Workbook()
                if self.select_report_sum == 'không lũy kế':
                    ws = wb.create_sheet(title='báo cáo kết hợp')
                else:
                    ws = wb.create_sheet(title='báo cáo kết hợp lũy kế')
            # NGOẠI TRỪ TỔNG DOANH THU
                #form1
                        # Tạo các dataframe để lưu dữ liệu
                dataframe_baocao_form1_moitrongthang = control_dataframe.create_dataframe_baocao_form1()
                dataframe_baocao_form1_moitrongnam = control_dataframe.create_dataframe_baocao_form1()
                dataframe_baocao_form1_hienhuu = control_dataframe.create_dataframe_baocao_form1()
                        # END Tạo các dataframe để lưu dữ liệu
                        # Get ket qua
                if self.select_report_sum == 'không lũy kế':
                    kq_thuchien_baocao_moitrongthang = control_database.get_data_thuchien_baocao(int(self.month_select), self.year, 'mới trong tháng',conn, cursor)
                    kq_thuchien_thangtruoc_moitrongthang = control_database.get_data_thuchien_baocao(int(self.month_select) - 1, self.year, 'mới trong tháng',conn, cursor)
                    kq_kehoach_baocao_moitrongthang = control_database.get_ketqua_kehoach_baocao(int(self.month_select), self.year, 'mới trong tháng',conn, cursor)
                    kq_thuchien_baocao_moitrongnam = control_database.get_data_thuchien_baocao(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                    kq_thuchien_thangtruoc_moitrongnam = control_database.get_data_thuchien_baocao(int(self.month_select) - 1, self.year, 'mới trong năm',conn, cursor)
                    kq_kehoach_baocao_moitrongnam = control_database.get_ketqua_kehoach_baocao(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                    kq_thuchien_baocao_hienhuu = control_database.get_data_thuchien_baocao(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                    kq_thuchien_thangtruoc_hienhuu = control_database.get_data_thuchien_baocao(int(self.month_select) - 1, self.year, 'hiện hữu',conn, cursor)
                    kq_kehoach_baocao_hienhuu = control_database.get_ketqua_kehoach_baocao(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                    if self.year - 1 ==2023:
                        kq_thuchien_baocao_cungki_moitrongthang = control_database.get_ketqua_thuchien_2023(int(self.month_select), 'mới trong tháng',conn, cursor)
                        kq_thuchien_baocao_cungki_moitrongnam = control_database.get_ketqua_thuchien_2023(int(self.month_select), 'mới trong năm',conn, cursor)
                        kq_thuchien_baocao_cungki_hienhuu = control_database.get_ketqua_thuchien_2023(int(self.month_select), 'hiện hữu',conn, cursor)
                    else:
                        kq_thuchien_baocao_cungki_moitrongthang = control_database.get_data_thuchien_baocao(int(self.month_select), self.year-1, 'mới trong tháng',conn, cursor)
                        kq_thuchien_baocao_cungki_moitrongnam = control_database.get_data_thuchien_baocao(int(self.month_select), self.year-1, 'mới trong năm',conn, cursor)
                        kq_thuchien_baocao_cungki_hienhuu = control_database.get_data_thuchien_baocao(int(self.month_select), self.year-1, 'hiện hữu',conn, cursor)
                else:
                    kq_thuchien_baocao_moitrongthang = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year, 'mới trong tháng',conn, cursor)
                    kq_thuchien_thangtruoc_moitrongthang = control_database.get_data_thuchien_baocao_luyke(int(self.month_select) - 1, self.year, 'mới trong tháng',conn, cursor)
                    kq_kehoach_baocao_moitrongthang = control_database.get_ketqua_kehoach_baocao_luyke(int(self.month_select), self.year, 'mới trong tháng',conn, cursor)
                    kq_thuchien_baocao_moitrongnam = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                    kq_thuchien_thangtruoc_moitrongnam = control_database.get_data_thuchien_baocao_luyke(int(self.month_select) - 1, self.year, 'mới trong năm',conn, cursor)
                    kq_kehoach_baocao_moitrongnam = control_database.get_ketqua_kehoach_baocao_luyke(int(self.month_select), self.year, 'mới trong năm',conn, cursor)
                    kq_thuchien_baocao_hienhuu = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                    kq_thuchien_thangtruoc_hienhuu = control_database.get_data_thuchien_baocao_luyke(int(self.month_select) - 1, self.year, 'hiện hữu',conn, cursor)
                    kq_kehoach_baocao_hienhuu = control_database.get_ketqua_kehoach_baocao_luyke(int(self.month_select), self.year, 'hiện hữu',conn, cursor)
                    if self.year - 1 ==2023:
                        kq_thuchien_baocao_cungki_moitrongthang = control_database.get_ketqua_thuchien_2023_luyke(int(self.month_select), 'mới trong tháng',conn, cursor)
                        kq_thuchien_baocao_cungki_moitrongnam = control_database.get_ketqua_thuchien_2023_luyke(int(self.month_select), 'mới trong năm',conn, cursor)
                        kq_thuchien_baocao_cungki_hienhuu = control_database.get_ketqua_thuchien_2023_luyke(int(self.month_select), 'hiện hữu',conn, cursor)
                    else:
                        kq_thuchien_baocao_cungki_moitrongthang = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year-1, 'mới trong tháng',conn, cursor)
                        kq_thuchien_baocao_cungki_moitrongnam = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year-1, 'mới trong năm',conn, cursor)
                        kq_thuchien_baocao_cungki_hienhuu = control_database.get_data_thuchien_baocao_luyke(int(self.month_select), self.year-1, 'hiện hữu',conn, cursor)
                        # End get ket qua
                if (kq_thuchien_baocao_moitrongthang.empty or kq_thuchien_baocao_moitrongnam.empty or kq_thuchien_baocao_hienhuu.empty):
                    st.warning("Không có dữ liệu thực hiện cho tháng, năm và loại báo cáo đã chọn.")
                    self.check_status_empty = True
                    sys.exit()
                        # Update data to dataframe
                dataframe_baocao_form1_moitrongthang = control_dataframe.update_column_data(dataframe_baocao_form1_moitrongthang['TH tháng'], kq_thuchien_baocao_moitrongthang, dataframe_baocao_form1_moitrongthang)
                dataframe_baocao_form1_moitrongthang= control_dataframe.update_column_data(dataframe_baocao_form1_moitrongthang['Tháng n-1'], kq_thuchien_thangtruoc_moitrongthang, dataframe_baocao_form1_moitrongthang)
                dataframe_baocao_form1_moitrongthang = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_baocao_moitrongthang, dataframe_baocao_form1_moitrongthang, 'KH tháng', self.column_name)
                dataframe_baocao_form1_moitrongthang = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_baocao_moitrongthang, dataframe_baocao_form1_moitrongthang, 'KH năm', 'LK_năm')
                dataframe_baocao_form1_moitrongnam = control_dataframe.update_column_data(dataframe_baocao_form1_moitrongnam['TH tháng'], kq_thuchien_baocao_moitrongnam, dataframe_baocao_form1_moitrongnam)
                dataframe_baocao_form1_moitrongnam= control_dataframe.update_column_data(dataframe_baocao_form1_moitrongnam['Tháng n-1'], kq_thuchien_thangtruoc_moitrongnam, dataframe_baocao_form1_moitrongnam)
                dataframe_baocao_form1_moitrongnam = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_baocao_moitrongnam, dataframe_baocao_form1_moitrongnam, 'KH tháng', self.column_name)
                dataframe_baocao_form1_moitrongnam = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_baocao_moitrongnam, dataframe_baocao_form1_moitrongnam, 'KH năm', 'LK_năm')
                dataframe_baocao_form1_hienhuu = control_dataframe.update_column_data(dataframe_baocao_form1_hienhuu['TH tháng'], kq_thuchien_baocao_hienhuu, dataframe_baocao_form1_hienhuu)
                dataframe_baocao_form1_hienhuu= control_dataframe.update_column_data(dataframe_baocao_form1_hienhuu['Tháng n-1'], kq_thuchien_thangtruoc_hienhuu, dataframe_baocao_form1_hienhuu)
                dataframe_baocao_form1_hienhuu = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_baocao_hienhuu, dataframe_baocao_form1_hienhuu, 'KH tháng', self.column_name)
                dataframe_baocao_form1_hienhuu = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_baocao_hienhuu, dataframe_baocao_form1_hienhuu, 'KH năm', 'LK_năm')
                if self.year - 1 ==2023:
                    dataframe_baocao_form1_moitrongthang = control_dataframe.map_kehoach_to_thuchien_tt(kq_thuchien_baocao_cungki_moitrongthang, dataframe_baocao_form1_moitrongthang, 'Cùng kì', self.column_name)
                    dataframe_baocao_form1_moitrongnam = control_dataframe.map_kehoach_to_thuchien_tt(kq_thuchien_baocao_cungki_moitrongnam, dataframe_baocao_form1_moitrongnam, 'Cùng kì', self.column_name)
                    dataframe_baocao_form1_hienhuu = control_dataframe.map_kehoach_to_thuchien_tt(kq_thuchien_baocao_cungki_hienhuu, dataframe_baocao_form1_hienhuu, 'Cùng kì', self.column_name)
                else:
                    dataframe_baocao_form1_moitrongthang = control_dataframe.update_column_data(dataframe_baocao_form1_moitrongthang['Cùng kì'], kq_thuchien_baocao_cungki_moitrongthang, dataframe_baocao_form1_moitrongthang)
                    dataframe_baocao_form1_moitrongnam = control_dataframe.update_column_data(dataframe_baocao_form1_moitrongnam['Cùng kì'], kq_thuchien_baocao_cungki_moitrongnam, dataframe_baocao_form1_moitrongnam)
                    dataframe_baocao_form1_hienhuu = control_dataframe.update_column_data(dataframe_baocao_form1_hienhuu['Cùng kì'], kq_thuchien_baocao_cungki_hienhuu, dataframe_baocao_form1_hienhuu)
                        # End update data to dataframe
                        # Update summary rows and format data
                dataframe_baocao_form1_moitrongthang = control_dataframe.update_summary_rows(dataframe_baocao_form1_moitrongthang)
                dataframe_baocao_form1_moitrongnam = control_dataframe.update_summary_rows(dataframe_baocao_form1_moitrongnam)
                dataframe_baocao_form1_hienhuu = control_dataframe.update_summary_rows(dataframe_baocao_form1_hienhuu)
                dataframe_baocao_form1_moitrongthang =  control_dataframe.update_df_baocao(dataframe_baocao_form1_moitrongthang)
                dataframe_baocao_form1_moitrongnam =  control_dataframe.update_df_baocao(dataframe_baocao_form1_moitrongnam)
                dataframe_baocao_form1_hienhuu =  control_dataframe.update_df_baocao(dataframe_baocao_form1_hienhuu)
                        # End update summary rows and format data
                #form2
                        # Tạo các dataframe để lưu dữ liệu
                form2_dataframe_baocao_moitrongthang =  control_dataframe.create_dataframe_public()
                form2_dataframe_baocao_moitrongnam =  control_dataframe.create_dataframe_public()
                form2_dataframe_baocao_hienhuu =  control_dataframe.create_dataframe_public()
                        # End Tạo các dataframe để lưu dữ liệu
                        # Get ket qua
                if self.select_report_sum == 'không lũy kế':
                    kq_thuchien_all_baocao_moitrongthang = control_database.get_data_form2_baocao(self.year,'mới trong tháng',conn, cursor)
                    kq_thuchien_all_baocao_moitrongnam = control_database.get_data_form2_baocao(self.year,'mới trong năm',conn, cursor)
                    kq_thuchien_all_baocao_hienhuu = control_database.get_data_form2_baocao(self.year,'hiện hữu',conn, cursor)
                else:
                    kq_thuchien_all_baocao_moitrongthang = control_database.get_data_form2_baocao_luyke(self.year,'mới trong tháng',conn, cursor)
                    kq_thuchien_all_baocao_moitrongnam = control_database.get_data_form2_baocao_luyke(self.year,'mới trong năm',conn, cursor)
                    kq_thuchien_all_baocao_hienhuu = control_database.get_data_form2_baocao_luyke(self.year,'hiện hữu',conn, cursor)
                        # End get ket qua
                if (kq_thuchien_all_baocao_moitrongthang.empty or kq_thuchien_all_baocao_moitrongnam.empty or kq_thuchien_all_baocao_hienhuu.empty):
                    st.warning("Không có dữ liệu thực hiện cho năm và loại báo cáo đã chọn.")
                    self.check_status_empty = True
                    sys.exit()
                        # Update data to dataframe
                form2_dataframe_baocao_moitrongthang = control_dataframe.map_thuchienall_to_dataframe(kq_thuchien_all_baocao_moitrongthang, form2_dataframe_baocao_moitrongthang)
                form2_dataframe_baocao_moitrongnam = control_dataframe.map_thuchienall_to_dataframe(kq_thuchien_all_baocao_moitrongnam, form2_dataframe_baocao_moitrongnam)
                form2_dataframe_baocao_hienhuu = control_dataframe.map_thuchienall_to_dataframe(kq_thuchien_all_baocao_hienhuu, form2_dataframe_baocao_hienhuu)
                        # End update data to dataframe
                        # Update summary rows and format data
                form2_dataframe_baocao_moitrongthang = control_dataframe.update_summary_rows(form2_dataframe_baocao_moitrongthang)
                form2_dataframe_baocao_moitrongnam = control_dataframe.update_summary_rows(form2_dataframe_baocao_moitrongnam)
                form2_dataframe_baocao_hienhuu = control_dataframe.update_summary_rows(form2_dataframe_baocao_hienhuu)
                #format data all 
                dataframe_baocao_form1_moitrongthang.iloc[:,2:]= dataframe_baocao_form1_moitrongthang.iloc[:,2:].astype(float).round(0)
                dataframe_baocao_form1_moitrongnam.iloc[:,2:]= dataframe_baocao_form1_moitrongnam.iloc[:,2:].astype(float).round(0)
                dataframe_baocao_form1_hienhuu.iloc[:,2:]= dataframe_baocao_form1_hienhuu.iloc[:,2:].astype(float).round(0)
                form2_dataframe_baocao_moitrongthang.iloc[:,2:]= form2_dataframe_baocao_moitrongthang.iloc[:,2:].astype(float).round(0)
                form2_dataframe_baocao_moitrongnam.iloc[:,2:]= form2_dataframe_baocao_moitrongnam.iloc[:,2:].astype(float).round(0)
                form2_dataframe_baocao_hienhuu.iloc[:,2:]= form2_dataframe_baocao_hienhuu.iloc[:,2:].astype(float).round(0)

                form2_dataframe_baocao_moitrongthang=form2_dataframe_baocao_moitrongthang.fillna(0)
                form2_dataframe_baocao_moitrongnam=form2_dataframe_baocao_moitrongnam.fillna(0)
                form2_dataframe_baocao_hienhuu=form2_dataframe_baocao_hienhuu.fillna(0)

                dataframe_baocao_form1_moitrongthang['% với KH năm'] = dataframe_baocao_form1_moitrongthang['% với KH năm'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_moitrongthang['% với KH tháng'] = dataframe_baocao_form1_moitrongthang['% với KH tháng'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_moitrongthang['% với cùng kì'] = dataframe_baocao_form1_moitrongthang['% với cùng kì'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_moitrongnam['% với KH năm'] = dataframe_baocao_form1_moitrongnam['% với KH năm'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_moitrongnam['% với KH tháng'] = dataframe_baocao_form1_moitrongnam['% với KH tháng'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_moitrongnam['% với cùng kì'] = dataframe_baocao_form1_moitrongnam['% với cùng kì'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_hienhuu['% với KH năm'] = dataframe_baocao_form1_hienhuu['% với KH năm'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_hienhuu['% với KH tháng'] = dataframe_baocao_form1_hienhuu['% với KH tháng'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_form1_hienhuu['% với cùng kì'] = dataframe_baocao_form1_hienhuu['% với cùng kì'].apply(lambda x: '{:.0f}%'.format(x))
            # TỔNG DOANH THU
                #form1
                        # Tạo các dataframe để lưu dữ liệu
                dataframe_baocao_tongdoanhthu_form1 = control_dataframe.create_dataframe_baocao_form1()
                form2_dataframe_baocao_tongdoanhthu = control_dataframe.create_dataframe_public()
                        # END Tạo các dataframe để lưu dữ liệu
                kq_thuchien_doanhthu, kq_thangtruoc_doanhthu, kq_kehoach_doanhthu, kq_cungki_doanhthu = control_dataframe.sum_hienhuu_moitrongnam(kq_thuchien_baocao_hienhuu,kq_thuchien_baocao_moitrongnam, kq_thuchien_thangtruoc_hienhuu,kq_thuchien_thangtruoc_moitrongnam,kq_kehoach_baocao_hienhuu,kq_kehoach_baocao_moitrongnam ,kq_thuchien_baocao_cungki_hienhuu, kq_thuchien_baocao_cungki_moitrongnam)
                dataframe_baocao_tongdoanhthu_form1 = control_dataframe.update_column_data(dataframe_baocao_tongdoanhthu_form1['TH tháng'], kq_thuchien_doanhthu, dataframe_baocao_tongdoanhthu_form1)
                dataframe_baocao_tongdoanhthu_form1= control_dataframe.update_column_data(dataframe_baocao_tongdoanhthu_form1['Tháng n-1'], kq_thangtruoc_doanhthu, dataframe_baocao_tongdoanhthu_form1)
                dataframe_baocao_tongdoanhthu_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_doanhthu, dataframe_baocao_tongdoanhthu_form1, 'KH tháng', self.column_name)
                dataframe_baocao_tongdoanhthu_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_kehoach_doanhthu, dataframe_baocao_tongdoanhthu_form1, 'KH năm', 'LK_năm')
                if self.year - 1 ==2023:
                    dataframe_baocao_tongdoanhthu_form1 = control_dataframe.map_kehoach_to_thuchien_tt(kq_cungki_doanhthu, dataframe_baocao_tongdoanhthu_form1, 'Cùng kì', self.column_name)
                else:
                    dataframe_baocao_tongdoanhthu_form1 = control_dataframe.update_column_data(dataframe_baocao_tongdoanhthu_form1['Cùng kì'], kq_cungki_doanhthu, dataframe_baocao_tongdoanhthu_form1)
                dataframe_baocao_tongdoanhthu_form1 = control_dataframe.update_summary_rows(dataframe_baocao_tongdoanhthu_form1)
                dataframe_baocao_tongdoanhthu_form1 =  control_dataframe.update_df_baocao(dataframe_baocao_tongdoanhthu_form1)
                #form2
                kq_thuchien_all_tongdoanhthu = kq_thuchien_all_baocao_hienhuu.copy()
                kq_thuchien_all_tongdoanhthu.iloc[:, 1] += kq_thuchien_all_baocao_moitrongnam.iloc[:, 1]
                form2_dataframe_baocao_tongdoanhthu = control_dataframe.create_dataframe_public()
                form2_dataframe_baocao_tongdoanhthu = control_dataframe.map_thuchienall_to_dataframe(kq_thuchien_all_tongdoanhthu, form2_dataframe_baocao_tongdoanhthu)
                form2_dataframe_baocao_tongdoanhthu = control_dataframe.update_summary_rows(form2_dataframe_baocao_tongdoanhthu)
                #format all 
                dataframe_baocao_tongdoanhthu_form1.iloc[:,2:]= dataframe_baocao_tongdoanhthu_form1.iloc[:,2:].astype(float).round(0)
                form2_dataframe_baocao_tongdoanhthu.iloc[:,2:]= form2_dataframe_baocao_tongdoanhthu.iloc[:,2:].astype(float).round(0)
                form2_dataframe_baocao_tongdoanhthu=form2_dataframe_baocao_tongdoanhthu.fillna(0)
                dataframe_baocao_tongdoanhthu_form1['% với KH năm'] = dataframe_baocao_tongdoanhthu_form1['% với KH năm'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_tongdoanhthu_form1['% với KH tháng'] = dataframe_baocao_tongdoanhthu_form1['% với KH tháng'].apply(lambda x: '{:.0f}%'.format(x))
                dataframe_baocao_tongdoanhthu_form1['% với cùng kì'] = dataframe_baocao_tongdoanhthu_form1['% với cùng kì'].apply(lambda x: '{:.0f}%'.format(x))
            # PHẦN TẠO WORDBOOK
                data = [dataframe_baocao_tongdoanhthu_form1.columns.tolist()] + dataframe_baocao_tongdoanhthu_form1.values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3, column=c_idx, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                data = [dataframe_baocao_form1_hienhuu.iloc[:,2:].columns.tolist()] + dataframe_baocao_form1_hienhuu.iloc[:,2:].values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3, column=c_idx + 12, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin')) 
                data = [dataframe_baocao_form1_moitrongnam.iloc[:,2:].columns.tolist()] + dataframe_baocao_form1_moitrongnam.iloc[:,2:].values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3, column=c_idx + 22, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                data = [dataframe_baocao_form1_moitrongthang.iloc[:,2:].columns.tolist()] + dataframe_baocao_form1_moitrongthang.iloc[:,2:].values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3, column=c_idx + 32, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin')) 
                        
                data = [form2_dataframe_baocao_tongdoanhthu.columns.tolist()] + form2_dataframe_baocao_tongdoanhthu.values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3 + 29, column=c_idx, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                data = [form2_dataframe_baocao_hienhuu.iloc[:,2:].columns.tolist()] + form2_dataframe_baocao_hienhuu.iloc[:,2:].values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3 + 29, column=c_idx + 15, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                data = [form2_dataframe_baocao_moitrongnam.iloc[:,2:].columns.tolist()] + form2_dataframe_baocao_moitrongnam.iloc[:,2:].values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3 + 29, column=c_idx + 28, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                data = [form2_dataframe_baocao_moitrongthang.iloc[:,2:].columns.tolist()] + form2_dataframe_baocao_moitrongthang.iloc[:,2:].values.tolist()
                for r_idx, row in enumerate(data, 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = ws.cell(row=r_idx + 3 + 29, column=c_idx + 41, value=value)
                        cell.font = Font(name='Times New Roman', size=12)  # Thiết lập font chữ và cỡ chữ
                        cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                            top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                for col_range in [(3, 12), (13, 22),(23,32), (33,42)]:
                    start_col, end_col = col_range
                    start_cell = ws.cell(row=3, column=start_col)
                    end_cell = ws.cell(row=3, column=end_col)
                    ws.merge_cells(start_row=3, start_column=start_col, end_row=3, end_column=end_col)
                    
                    # Thiết lập lại border cho các ô trong vùng merge
                    for row in ws.iter_rows(min_row=3, max_row=3, min_col=start_col, max_col=end_col):
                        for cell in row:
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                for col_range in [(3, 15), (16, 28),(29,41), (42,54)]:
                    start_col, end_col = col_range
                    start_cell = ws.cell(row=32, column=start_col)
                    end_cell = ws.cell(row=32, column=end_col)
                    ws.merge_cells(start_row=32, start_column=start_col, end_row=32, end_column=end_col)
                    
                    # Thiết lập lại border cho các ô trong vùng merge
                    for row in ws.iter_rows(min_row=32, max_row=32, min_col=start_col, max_col=end_col):
                        for cell in row:
                            cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                                top=Side(border_style='thin'), bottom=Side(border_style='thin'))


                # Gán giá trị cho ô merged và thiết lập font, đường viền
                for col_idx, text,color in [(3, 'TỔNG KẾ HOẠCH','2F75B5'), (13, 'HIỆN HỮU','8932E5'), (23, 'MỚI TRONG NĂM','548235'), (33, 'MỚI TRONG THÁNG','F4B084')]:
                    cell = ws.cell(row=3, column=col_idx, value=text)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    cell.font = Font(name='Times New Roman', size=12, bold=True)  # Thiết lập font chữ và cỡ chữ
                    cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type = "solid")
                # Gán giá trị cho ô merged và thiết lập font, đường viền
                for col_idx, text,color in [(3, 'TỔNG KẾ HOẠCH','2F75B5'), (16, 'HIỆN HỮU','8932E5'), (29, 'MỚI TRONG NĂM','548235'), (42, 'MỚI TRONG THÁNG','F4B084')]:
                    cell = ws.cell(row=32, column=col_idx, value=text)
                    cell.alignment = openpyxl.styles.Alignment(horizontal='center')
                    cell.font = Font(name='Times New Roman', size=12, bold=True)  # Thiết lập font chữ và cỡ chữ
                    cell.border = Border(left=Side(border_style='thin'), right=Side(border_style='thin'), 
                                        top=Side(border_style='thin'), bottom=Side(border_style='thin'))
                    cell.fill = PatternFill(start_color=color, end_color=color, fill_type = "solid")
                    # In đậm các dòng cần thiết
                bold_rows_need_thuchien = [4,5,6,7,17,20, 21,25,26,27,28,29,30,33,34,35,36,46,49,50,54,55,56,57,58,59] 
                for row_idx in bold_rows_need_thuchien:
                    for col in ws.iter_cols(min_col=1, max_col=ws.max_column, min_row=row_idx, max_row=row_idx):
                        for cell in col:
                            cell.font = Font(bold=True)
                
                # Tự động điều chỉnh độ rộng của các cột dựa trên nội dung của ô
                for column_cells in ws.columns:
                    length = max(len(str(cell.value)) for cell in column_cells)
                    ws.column_dimensions[column_cells[0].column_letter].width = length + 2
                if self.check_status_empty == False:
                    wb.save(output_all_baocao)
                    output_all_baocao.seek(0)
                    excel_base64 = base64.b64encode(output_all_baocao.read()).decode()

                    # Create a download link using st.download_button
                    if self.select_report_sum == 'không lũy kế':
                        file_name_download_all_baocao = f"BaoCaoDoanhThu_{self.year}_Thang_{self.month_select}.xlsx"
                    else:
                        file_name_download_all_baocao = f"BaoCaoDoanhThu_{self.year}_Thang_{self.month_select}_LuyKe.xlsx"
                    download_link_all_baocao = f'<a href="data:application/vnd.openxmlformats-officedocument.spreadsheetml.sheet;base64,{excel_base64}" download="{file_name_download_all_baocao}.xlsx">Tải xuống kết quả</a>'
                    st.success("Tạo file thành công.")
                    st.markdown(download_link_all_baocao, unsafe_allow_html=True)
    def view_insert_data(self):
        if 'conn' not in st.session_state:
            st.session_state.conn, st.session_state.cursor = connect_to_mysql()
        conn = st.session_state.conn
        
        if conn is None:
            st.session_state.conn, st.session_state.cursor = connect_to_mysql()
            conn = st.session_state.conn
            cursor =conn.cursor()

        control_database = CONTROL_DATABASE()
        control_dataframe = DATAFRAME()
        # PHẦN LỰA CHỌN
        st.header("Thực hiện thêm dữ liệu")
        col_view_1, col_view_2 = st.columns(2)
        with col_view_1:
            self.select_insert_baocao = st.radio("Chọn loại báo cáo: ", ("HIỆN HỮU", "TỔNG HỢP"),key="radio_chonloaibaocao_page_insertdata")
            lines = self.managerment_database.get_line(conn)
            self.lines_insert_empty_thuchien = st.empty()
            self.lines_insert = self.lines_insert_empty_thuchien.selectbox('Line', lines, key="selectbox_line_page_insertdata")
            
        with col_view_2:
            if self.select_insert_baocao == 'HIỆN HỮU':
                self.option_insert = st.radio(
                            "Chọn một tùy chọn",
                            ('KẾ HOẠCH', 'THỰC HIỆN'), key="radio_tuychonloaixem_page_insertdata"
                        )
            else:
                self.option_insert = st.radio(
                            "Chọn một tùy chọn",
                            ('KẾ HOẠCH TỔNG HỢP', 'THỰC HIỆN TỔNG HỢP'),key="radio_tuychonloaixemtonghop_page_insertdata"
                        )
            # self.year = st.text_input("Nhập năm:", "")
            self.year_insert = st.selectbox("Chọn năm:", list(range(2000, date.today().year + 1)), index=date.today().year - 2000, key="selectbox_nam_page_insertdata")
            if self.option_insert == 'KẾ HOẠCH TỔNG HỢP':
                self.baocaotonghop_option_empty = st.empty()
                self.baocaotonghop_option = self.baocaotonghop_option_empty.selectbox("Chọn loại: ", ["tổng doanh thu","mới trong tháng", "mới trong năm", "hiện hữu"],key="selectbox_loaibaocao_page_insertdata")
            if self.option_insert == 'THỰC HIỆN TỔNG HỢP':
                self.baocaotonghop_option_empty = st.empty()
                self.baocaotonghop_option = self.baocaotonghop_option_empty.selectbox("Chọn loại: ", ["mới trong tháng", "mới trong năm", "hiện hữu"],key="selectbox_loaibaocao2_page_insertdata")
        if self.option_insert == 'THỰC HIỆN TỔNG HỢP' or self.option_insert == 'KẾ HOẠCH TỔNG HỢP' or self.option_insert == 'THỰC HIỆN':
            self.lines_insert_empty_thuchien.empty()
        if self.option_insert == 'THỰC HIỆN' or self.option_insert == 'THỰC HIỆN TỔNG HỢP':
            months = ['1', '2', '3', '4', '5', '6', '7', '8', '9', '10', '11', '12']
            # Hiển thị các tháng trong một selectbox với giá trị hiển thị và giá trị thực tế
            self.month_select_insert = st.selectbox("Chọn tháng:", months, key="selectbox_thang_page_insertdata")
            self.column_name = f'Tháng_{int(self.month_select_insert):02d}'
        
            
            
        self.insert_data_todatabase = self.upload_excel()
        # PHẦN THAO TÁC VỚI DỮ LIỆU
        col_button_insert_layout1, col_button_insert_layout2 = st.columns(2)
        with col_button_insert_layout1:
            if st.button('Thêm dữ liệu vào cơ sở dữ liệu',use_container_width=True,key="themvaocsdldata"):
                cursor =conn.cursor()
                
                if self.insert_data_todatabase is None:
                    st.warning("Vui lòng tải lên file Excel trước khi thêm dữ liệu.")
                else:
                    if self.option_insert == 'KẾ HOẠCH':
                        if (self.lines_insert,self.year_insert):
                            check_data_from_kehoach = control_database.query_kehoach_by_line_year(self.lines_insert, self.year_insert,conn, cursor)
                            if (check_data_from_kehoach.empty):
                                data_insert = control_database.select_rows_kehoach(self.insert_data_todatabase)
                                status_insert = control_database.insert_data_kehoach(self.lines_insert, self.year_insert, data_insert,conn, cursor)
                                if status_insert == True:
                                    content_action_check = "vừa thêm dữ liệu kế hoạch cho line " + str(self.lines_insert) + "với năm " + str(self.year_insert) + " vào csdl!"
                                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                                    st.success("Thêm dữ liệu thành công.")
                                else:
                                    st.error("Thêm dữ liệu thất bại.")
                            else:
                                st.warning("Dữ liệu đã tồn tại, vui lòng kiểm tra lại.")
                                
                        else:
                            st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
                    if self.option_insert == 'THỰC HIỆN':
                        if (self.lines_insert,self.year_insert,self.month_select_insert):
                            check_data_from_thuchien = control_database.get_data_thuchien(self.month_select_insert, self.year_insert, self.lines_insert,conn, cursor)
                            if (check_data_from_thuchien.empty):
                                group_data = control_database.group_data_for_insert_thuchien(self.insert_data_todatabase)
                                status_insert_thuc_hien = control_database.insert_data_to_thuchien(self.month_select_insert, self.year_insert, group_data,conn, cursor)
                                if status_insert_thuc_hien == True:
                                    content_action_check = "vừa thêm dữ liệu thực hiện vào tháng " + str(self.month_select_insert) + " năm " + str(self.year_insert) + " vào csdl!"
                                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                                    st.success("Thêm dữ liệu thành công.")
                                else:
                                    st.error("Thêm dữ liệu thất bại.")
                            else:
                                st.warning("Dữ liệu đã tồn tại, vui lòng kiểm tra lại.")
                        else:
                            st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
                    if self.option_insert == 'KẾ HOẠCH TỔNG HỢP':
                        if (self.baocaotonghop_option,self.year_insert):
                            check_data_from_kehoach_tonghop = control_database.query_kehoach_by_line_year(self.baocaotonghop_option, self.year_insert,conn, cursor)
                            if (check_data_from_kehoach_tonghop.empty):
                                data_insert = control_database.select_rows_kehoach_baocao(self.insert_data_todatabase)
                                status_insert = control_database.insert_data_kehoach(self.baocaotonghop_option, self.year_insert, data_insert,conn, cursor)
                                if status_insert == True:
                                    content_action_check = "vừa thêm dữ liệu kế hoạch tổng hợp vào năm " + str(self.year_insert) + "với loại báo cáo: " + str(self.baocaotonghop_option) + " vào csdl!"
                                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                                    st.success("Thêm dữ liệu thành công.")
                                else:
                                    st.error("Thêm dữ liệu thất bại.")
                            else:
                                st.warning("Dữ liệu đã tồn tại, vui lòng kiểm tra lại.")
                        else:
                            st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
                    if self.option_insert == 'THỰC HIỆN TỔNG HỢP':
                        if (self.baocaotonghop_option,self.year_insert,self.month_select_insert):
                            check_data_from_thuchien_tonghop = control_database.get_data_thuchien_baocao(self.month_select_insert, self.year_insert, self.baocaotonghop_option,conn, cursor)
                            if (check_data_from_thuchien_tonghop.empty):
                                group_data = control_database.group_data_insert_baocaothuchien(self.insert_data_todatabase)
                                status_insert_thuc_hien = control_database.insert_data_to_thuchien_baocao(self.baocaotonghop_option,self.month_select_insert, self.year_insert, group_data,conn, cursor)
                                if status_insert_thuc_hien == True:
                                    content_action_check = "vừa thêm dữ liệu thực hiện tổng hợp vào tháng " + str(self.month_select_insert) + " năm " + str(self.year_insert) + "với loại báo cáo: " + str(self.baocaotonghop_option) + " vào csdl!"
                                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                                    st.success("Thêm dữ liệu thành công.")
                                else:
                                    st.error("Thêm dữ liệu thất bại.")
                            else:
                                st.warning("Dữ liệu đã tồn tại, vui lòng kiểm tra lại.")
                        else:
                            st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
        with col_button_insert_layout2:
            if st.button('Xem trước dữ liệu',use_container_width=True,key="xemtruocdulieu_pageinsert"):
                cursor =conn.cursor()
                if self.insert_data_todatabase is None:
                    st.warning("Vui lòng tải lên file Excel trước khi xem dữ liệu.")
                else:
                    st.dataframe(self.insert_data_todatabase)
    def managerment_line(self):
        db_connection = DatabaseConnection()
        # Sử dụng kết nối và con trỏ để thao tác với cơ sở dữ liệu
        conn = db_connection.conn
        cursor = db_connection.cursor
        # Hiển thị bảng danh sách các tổ
        with st.expander("Quản lý các Line", expanded=True):
            data = self.managerment_database.get_ten_to(conn)
            for idx,row  in enumerate(data):
                col1, col2, col3 = st.columns([2, 0.5, 0.5],gap="small")
                with col1:
                    st.write(f"{row['ten']}")
                with col2:
                    if st.button("Xóa", key=f"delete_{row['id']}",use_container_width=True):
                        self.managerment_database.delete_ten_to(row['id'], conn, cursor)
                        # Phần action
                        content_action_check = "vừa xóa line có tên: " + row['ten'] + " khỏi csdl!"
                        mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                        # End action
                        st.success("Đã xóa tổ")
                        # Chỉ cập nhật lại trạng thái, không tải lại toàn bộ trang
                        st.session_state['rerun'] = True
                with col3:
                    if st.button("Sửa", key=f"edit_{row['id']}",use_container_width=True):
                        st.session_state.edit_id = row['id']
                        st.session_state.edit_ten = row['ten']
                        # Chỉ cập nhật lại trạng thái, không tải lại toàn bộ trang
                        st.session_state['rerun'] = True
                if idx < len(data) - 1:
                                st.markdown("<hr style='margin-top:0; margin-bottom: 0;'>", unsafe_allow_html=True)

        if 'rerun' in st.session_state and st.session_state['rerun']:
            st.session_state['rerun'] = False
            st.rerun()

        if "edit_id" in st.session_state and "edit_ten" in st.session_state:
            st.subheader("Chỉnh sửa tổ")
            new_ten = st.text_input("Tên tổ mới", st.session_state.edit_ten, key="edit_ten_input_page_managermentline")
            if st.button("Lưu", key="save_editpagemanagermentline"):
                self.managerment_database.update_ten_to(st.session_state.edit_id, new_ten, conn, cursor)
                # Phần action
                content_action_check = "vừa sửa line có tên: " + st.session_state.edit_ten + "thành tên: " + new_ten
                mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                # End action
                st.success("Đã cập nhật tổ")
                st.session_state.pop("edit_id")
                st.session_state.pop("edit_ten")
                st.rerun()

    # Thêm tổ mới
        with st.expander("Thêm Line mới"):
            st.subheader("Thêm Line mới")

            # Sử dụng session_state để lưu trữ giá trị của trường input
            if 'new_ten_value' not in st.session_state:
                st.session_state['new_ten_value'] = ""

            new_ten = st.text_input("Tên tổ mới", value=st.session_state['new_ten_value'], key="new_ten_input_pagemanagerment_lien")

            if st.button("Thêm", key="add_new_linepagemanagermentline"):
                self.managerment_database.add_ten_to(new_ten, conn, cursor)
                content_action_check = "vừa thêm line có tên: " + new_ten + " vào csdl!"
                mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                st.success("Đã thêm tổ mới")
                st.session_state['new_ten_value'] = "" 
                new_ten = ""  
                st.rerun()
            else:
                st.session_state['new_ten_value'] = new_ten 
    def action_managerment(self):
        db_connection = DatabaseConnection()
        # Sử dụng kết nối và con trỏ để thao tác với cơ sở dữ liệu
        conn = db_connection.conn
        cursor = db_connection.cursor
        # Thanh công cụ
        st.subheader("Quản lý thao tác người dùng:")

        # Các điều kiện lọc
        col_action_check_1, col_action_check_2 = st.columns(2)
        with col_action_check_1:
            find_action = st.text_input("Nhập nội dung thao tác:",placeholder="Nhập nội dung", key="content_action_check")  # Tìm theo nội dung
        with col_action_check_2:
            date_input_filter_action = st.date_input("Chọn ngày thực hiện:", value=None, key="date_input_filter_action")  # Tìm theo ngày

        # Xử lý hiển thị bảng dữ liệu theo các trường hợp
        if find_action:  # Trường hợp 1: Có điều kiện tìm kiếm nội dung
            rows = mvnpt.search_action_check(find_action,conn, cursor)
            if rows:
                df_action_check_view = pd.DataFrame(rows, columns=["ID", "Username", "Action", "Time"])
                st.dataframe(df_action_check_view,use_container_width=True)  # Hiển thị dữ liệu tìm thấy
            else:
                st.warning("Không tìm thấy hành động nào khớp với nội dung tìm kiếm.")
        
        elif date_input_filter_action:  # Trường hợp 2: Có điều kiện lọc theo ngày
            rows = mvnpt.select_action_check_by_filter( date_input_filter_action,conn, cursor)
            if rows:
                df_action_check_view = pd.DataFrame(rows, columns=["ID", "Username", "Action", "Time"])
                st.dataframe(df_action_check_view,use_container_width=True)  # Hiển thị dữ liệu tìm thấy
            else:
                st.warning("Không tìm thấy hành động nào cho ngày đã chọn.")
        
        else:  # Trường hợp 3: Hiển thị tất cả
            rows = mvnpt.select_action_check(conn, cursor)
            if rows:
                df_action_check_view = pd.DataFrame(rows, columns=["ID", "Username", "Action", "Time"])
                st.dataframe(df_action_check_view,use_container_width=True)# Hiển thị tất cả các hành động
            else:
                st.warning("Không có dữ liệu để hiển thị.")
    def user_setting_vnpt(self):
        db_connection = DatabaseConnection()
        # Sử dụng kết nối và con trỏ để thao tác với cơ sở dữ liệu
        conn = db_connection.conn
        cursor = db_connection.cursor
        # Thanh công cụ
        st.subheader("Cài đặt người dùng :")
        username_select,password_select = mvnpt.select_info_user(st.session_state.usernamevnpt,conn, cursor)
        # Các điều kiện lọc
        username_edit = st.text_input("Nhập tên người dùng:",username_select, key="username_edit_pagemanagermentuser")  
        password_edit = st.text_input("Nhập mật khẩu:",placeholder="Nhập mật khẩu mới", key="password_edit_pagemanagermentuser")  
        uploaded_image = st.file_uploader("Chọn ảnh để upload", type=["jpg", "png", "jpeg"],key="file_upload_image_pagemanagermentuser")
        
        if st.button("Lưu",key="save_userpagemanagermentuser"):
            check_duplicate_username = mvnpt.check_duplicate_user(username_edit,conn, cursor)
            if check_duplicate_username == True and username_edit != st.session_state.usernamevnpt:
                st.warning("Tên người dùng đã tồn tại, vui lòng chọn tên khác.")
            else:
                if uploaded_image is not None:
                    bucket_name = "vnpt-bucket-manager" 
                    image_url_edit = mvnpt.upload_image_to_gcs(uploaded_image, bucket_name)
                else:
                    image_url_edit = st.session_state.image_url_profile_123
                if password_edit:
                    update_status = mvnpt.update_user(st.session_state.usernamevnpt,username_edit,password_edit,image_url_edit,conn, cursor)
                    if update_status is True:
                        image_url_edit_reponse = requests.get(image_url_edit)
                        if image_url_edit_reponse.status_code == 200:
                            st.session_state.image_profile = image_url_edit_reponse
                            st.rerun()
                        st.success("Đã cập nhật thông tin người dùng.")
                    else:
                        st.error("Cập nhật thông tin người dùng thất bại.")
                else:
                    st.warning("Vui lòng nhập mật khẩu.")            
    def layout_select_view(self):
        # Sử dụng session_state để lưu trữ trạng thái của check_status_user
        if 'check_status_user' not in st.session_state:
            st.session_state.check_status_user = 'action_user'

        layout_action_user_1, layout_action_user_2 = st.columns(2)

        with layout_action_user_1:
            if st.button("Quản lý thao tác người dùng", key="action_userpagemanagerment_user", use_container_width=True):
                st.session_state.check_status_user = 'action_user'

        with layout_action_user_2:
            if st.button("Cài đặt người dùng", key="setting_user_page_managerment_user", use_container_width=True):
                st.session_state.check_status_user = 'setting_user'

        # Kiểm tra trạng thái lưu trong session_state để gọi các hàm tương ứng
        if st.session_state.check_status_user == 'action_user':
            self.action_managerment()
        elif st.session_state.check_status_user == 'setting_user':
            self.user_setting_vnpt()

            
    
    def upload_excel(self):
        # Cho phép người dùng tải lên file
        uploaded_file = st.file_uploader("Chọn file Excel", type=["xlsx", "xls"], key="file_upload_excel")

        # Nếu người dùng đã tải lên file
        if uploaded_file is not None:
            try:
                # Đọc dữ liệu từ file Excel
                df = pd.read_excel(uploaded_file, sheet_name=0, engine='openpyxl')
                return df

            except Exception as e:
                st.error(f"Đã xảy ra lỗi: {str(e)}")
    # @st.cache_resource
    def view_managerment_data(self):
        if 'conn' not in st.session_state:
            st.session_state.conn, st.session_state.cursor = connect_to_mysql()
        conn = st.session_state.conn
        cursor =conn.cursor()
        if conn is None:
            st.session_state.conn, st.session_state.cursor = connect_to_mysql()
            conn = st.session_state.conn
            cursor =conn.cursor()
        st.header("Thực hiện xóa dữ liệu")
        contrl_database = CONTROL_DATABASE()
        col_delete_3, col_delete_4 = st.columns(2)
        with col_delete_3:
            self.select_option_report = st.radio(
                                    "Chọn một loại báo cáo",
                                    ('HIỆN HỮU', 'TỔNG HỢP'), key="radio_chonloaibaocao_page_managerment_data"
                                )
        with col_delete_4:
            if self.select_option_report == 'HIỆN HỮU':
                self.option_delete = st.radio(
                                        "Chọn một tùy chọn",
                                        ('KẾ HOẠCH', 'THỰC HIỆN'), key="radio_tuychonloaixem_page_managerment_data"
                                    )
            else:
                self.option_delete = st.radio(
                                        "Chọn một tùy chọn",
                                        ('KẾ HOẠCH TỔNG HỢP', 'THỰC HIỆN TỔNG HỢP'), key="radio_tuychonloaixemtonghop_page_managerment_data"
                                    )
        if self.option_delete == 'KẾ HOẠCH':
            distinct_data = contrl_database.query_distinct_kehoach(conn, cursor)
        if self.option_delete == 'THỰC HIỆN':
            distinct_data = contrl_database.query_distinct_thuchien(conn, cursor)
        if self.option_delete == 'KẾ HOẠCH TỔNG HỢP':
            distinct_data = contrl_database.query_distinct_kehoach_tonghop(conn, cursor)
        if self.option_delete == 'THỰC HIỆN TỔNG HỢP':
            distinct_data = contrl_database.query_distinct_thuchien_tonghop(conn, cursor)
        col_delete_1, col_delete_2 = st.columns(2)
        with col_delete_1:
            self.year_delete = st.selectbox("Chọn năm:", distinct_data['Năm'].unique(), key="selectbox_nam_page_managerment_data")
        with col_delete_2:
            if self.option_delete == 'KẾ HOẠCH':
                lines_delete = st.selectbox("Chọn line:", distinct_data['Line'].unique(), key="selectbox_line_page_managerment_data")
        if self.option_delete == 'THỰC HIỆN' or self.option_delete == 'THỰC HIỆN TỔNG HỢP':
            self.month_delete = st.selectbox("Chọn tháng:", distinct_data['Tháng'].unique(), key="selectbox_thang_page_managerment_data")
        if self.option_delete == 'KẾ HOẠCH TỔNG HỢP' or self.option_delete == 'THỰC HIỆN TỔNG HỢP':
            self.loaibaocao_delete = st.selectbox("Chọn loại báo cáo:", distinct_data['Loại báo cáo'].unique(), key="selectbox_loaibaocao_page_managerment_data")
        if st.button('Xem dữ liệu',key="xemdulieu_page_managerment_data"):
            if not distinct_data.empty:
                st.dataframe(distinct_data)
            else:
                st.error("Không có dữ liệu.")
                
        if st.button('Xóa dữ liệu',key="xoadulieu_page_managerment_data"):
            
            if self.option_delete == 'KẾ HOẠCH':
                if (lines_delete,self.year_delete):
                    status_delete = contrl_database.delete_from_kehoach(lines_delete, int(self.year_delete),conn, cursor)
                    content_action_check = "vừa xóa dữ liệu kế hoạch cho line " + str(lines_delete) + "với năm " + str(self.year_delete) + " khỏi csdl!"
                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                else:
                    st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
            if self.option_delete == 'THỰC HIỆN':
                if (self.year_delete,self.month_delete):
                    status_delete = contrl_database.delete_from_thuchien(int(self.year_delete), self.month_delete,conn, cursor)
                    content_action_check = "vừa xóa dữ liệu thực hiện vào tháng " + str(self.month_delete) + " năm " + str(self.year_delete) + " khỏi csdl!"
                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                else:
                    st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
            if self.option_delete == 'KẾ HOẠCH TỔNG HỢP':
                if (self.year_delete,self.loaibaocao_delete):
                    status_delete = contrl_database.delete_from_kehoach(self.loaibaocao_delete, int(self.year_delete),conn, cursor)
                    content_action_check = "vừa xóa dữ liệu kế hoạch tổng hợp vào năm " + str(self.year_delete) + "với loại báo cáo " + str(self.loaibaocao_delete) + " khỏi csdl!"
                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                else:
                    st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
            if self.option_delete == 'THỰC HIỆN TỔNG HỢP':
                if (self.year_delete,self.month_delete,self.loaibaocao_delete):
                    status_delete = contrl_database.delete_from_thuchien_baocao(int(self.year_delete), self.month_delete, self.loaibaocao_delete,conn, cursor)
                    content_action_check = "vừa xóa dữ liệu thực hiện tổng hợp vào tháng " + str(self.month_delete) + " năm " + str(self.year_delete) + "với loại báo cáo " + str(self.loaibaocao_delete) + " khỏi csdl!"
                    mvnpt.insert_action_check(st.session_state.usernamevnpt, content_action_check, conn, cursor)
                else:
                    st.warning("Bạn nhập thiếu thông tin, vui lòng kiểm tra lại.")
                    
    def info_contact(self):
        st.markdown("""
        - **Địa chỉ:** 142 Điện Biên Phủ, Phường Đakao, Quận 1, TP.Hồ Chí Minh
        - **Số điện thoại:** 18001166
        - **Email:** sales@vnpttphcm.com.vn
        """)
        
    def footer_info(_self):
        st.sidebar.markdown("""
            ---
            """)
        st.sidebar.info(
            "Created and designed by [VNPT-HCM](https://vnpttphcm.net.vn/)")

app = MAINCLASS()

reconnect_if_needed()
# Thực thi hàm main
if __name__ == '__main__':
    
    # Login
    login()
    # Tải giao diện
    if 'image_profile' in st.session_state:
        image_inter_face =  st.session_state.image_profile
        app.run(image_inter_face)
    else:
        image_inter_face = None
        app.run(image_inter_face)
    # Check login
    if st.session_state.is_logged_in:

        selected = app.streamlit_menu(st.session_state.role_access_admin)

        sidebar_expand_function_1, sidebar_expand_function_2 = st.columns(2)
        with sidebar_expand_function_1:
            if st.sidebar.button('Kết nối lại database', key='reconnect',use_container_width=True,help='Bạn kết nối lại database nếu bị mất kết nối'):
                st.session_state.conn, st.session_state.cursor = connect_to_mysql()
        with sidebar_expand_function_2:
            if st.sidebar.button('Đăng xuất', key='logout',use_container_width=True,help='Đăng xuất khỏi hệ thống'):
                st.session_state.is_logged_in = False
                st.session_state.role_access_admin = False
                st.session_state.image_profile = None
                st.experimental_rerun()
        app.main(selected,st.session_state.role_access_admin)
app.footer_info()